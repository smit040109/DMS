"""
GO OIL — Distributor Price List importer.

Parses the official "AUTHORIZED DISTRIBUTOR PRICE CIRCULAR" format (Excel / CSV /
PDF) where CATEGORIES are shown as full-width header rows and products sit under
them, with columns:

  MATERIAL DESCRIPTION | GRADE/ SPECS | PACK SIZE | MRP | DLP |
  DISTRIBUTOR MARGINE | CASH COUPON | FOC BENEFITS | MONTHLY GIFT | TRADE DISCOUNT

Returns a normalized ordered list of items:
  { "type": "category", "name": "MCO - SUPER CATEGORY" }
  { "type": "product",  "material_description": "...", "grade_specs": "...",
    "pack_size": "...", "mrp": 1150.0, "dlp": 845.0,
    "distributor_margin_pct": 9.0, "cash_coupon": "50 TO 100",
    "foc_benefits": "", "monthly_gift": "AVAILABLE", "trade_discount": "" }
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional


# ── header aliases (normalized lower-case, non-alnum stripped) ───────────────
_ALIASES = {
    "material_description": ["material description", "materialdescription", "material", "description", "product", "item"],
    "grade_specs": ["grade specs", "grade/ specs", "grade/specs", "grade / specs", "grade", "specs", "specification"],
    "pack_size": ["pack size", "packsize", "pack", "size", "packing"],
    "mrp": ["mrp", "m.r.p", "mrp rs"],
    "dlp": ["dlp", "d.l.p", "dlp rs", "dealer price", "distributor price"],
    "distributor_margin_pct": ["distributor margine", "distributor margin", "distributormargine", "margin", "margine", "dist margin"],
    "cash_coupon": ["cash coupon", "cashcoupon", "coupon"],
    "foc_benefits": ["foc benefits", "focbenefits", "foc benefit", "foc"],
    "monthly_gift": ["monthly gift", "monthlygift", "gift"],
    "trade_discount": ["trade discount", "tradediscount", "discount", "td"],
}


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def _key(s: Any) -> str:
    """Aggressive normalization used for matching header labels."""
    return re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower())


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # keep digits, dot, minus — drop ₹, commas, % etc.
    s = s.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _build_colmap(header_cells: List[str]) -> Optional[Dict[str, int]]:
    """Return {field: col_index} if this row looks like a header, else None."""
    norm = [_key(c) for c in header_cells]
    colmap: Dict[str, int] = {}
    for field, aliases in _ALIASES.items():
        for idx, cell in enumerate(norm):
            cellk = re.sub(r"\s+", " ", cell).strip()
            if not cellk:
                continue
            for a in aliases:
                if cellk == a or (len(a) >= 4 and a in cellk) or (len(cellk) >= 4 and cellk in a):
                    colmap[field] = idx
                    break
            if field in colmap:
                break
    # It's a header only if we found the anchor columns
    if "material_description" in colmap and ("dlp" in colmap or "mrp" in colmap):
        return colmap
    return None


def _get(row: List[Any], colmap: Dict[str, int], field: str) -> str:
    idx = colmap.get(field)
    if idx is None or idx >= len(row):
        return ""
    return _norm(row[idx])


# ── raw row extraction per source type ───────────────────────────────────────
def _rows_from_xlsx(raw: bytes) -> List[List[List[Any]]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    tables: List[List[List[Any]]] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if rows:
            tables.append(rows)
    return tables


def _rows_from_csv(raw: bytes) -> List[List[List[Any]]]:
    import csv
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [[list(r) for r in reader]]


def _rows_from_pdf(raw: bytes) -> List[List[List[Any]]]:
    import pdfplumber
    tables: List[List[List[Any]]] = []
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            page_tables = page.extract_tables() or []
            for t in page_tables:
                if t:
                    tables.append(t)
            if not page_tables:
                # fallback: split each text line by 2+ spaces
                txt = page.extract_text() or ""
                line_rows = [re.split(r"\s{2,}", ln.strip()) for ln in txt.splitlines() if ln.strip()]
                if line_rows:
                    tables.append(line_rows)
    return tables


def parse_price_list(raw: bytes, filename: str) -> Dict[str, Any]:
    fn = (filename or "").lower()
    if fn.endswith(".pdf"):
        tables = _rows_from_pdf(raw)
        src = "pdf"
    elif fn.endswith(".csv"):
        tables = _rows_from_csv(raw)
        src = "csv"
    elif fn.endswith(".xlsx") or fn.endswith(".xlsm") or fn.endswith(".xls"):
        tables = _rows_from_xlsx(raw)
        src = "xlsx"
    else:
        # try excel first, then pdf
        try:
            tables = _rows_from_xlsx(raw)
            src = "xlsx"
        except Exception:
            tables = _rows_from_pdf(raw)
            src = "pdf"

    items: List[Dict[str, Any]] = []
    warnings: List[str] = []
    colmap: Optional[Dict[str, int]] = None
    header_found = False

    PRICING_FIELDS = ["mrp", "dlp", "distributor_margin_pct", "cash_coupon",
                      "foc_benefits", "monthly_gift", "trade_discount", "grade_specs", "pack_size"]

    for table in tables:
        for raw_row in table:
            row = list(raw_row) if raw_row is not None else []
            cells = [_norm(c) for c in row]
            if not any(cells):
                continue

            # detect / re-detect header row
            maybe = _build_colmap(cells)
            if maybe:
                colmap = maybe
                header_found = True
                continue

            if not colmap:
                continue

            md = _get(row, colmap, "material_description")
            if not md:
                # sometimes category text lands in a different first non-empty cell
                first = next((c for c in cells if c), "")
                grade = _get(row, colmap, "grade_specs")
                dlp = _get(row, colmap, "dlp")
                mrp = _get(row, colmap, "mrp")
                if first and not grade and not dlp and not mrp:
                    items.append({"type": "category", "name": first})
                continue

            grade = _get(row, colmap, "grade_specs")
            pack = _get(row, colmap, "pack_size")
            dlp = _to_float(_get(row, colmap, "dlp"))
            mrp = _to_float(_get(row, colmap, "mrp"))

            # skip repeated column-title rows
            if _key(md) in ("material description", "material", "description"):
                continue

            has_pricing = any(_get(row, colmap, f) for f in PRICING_FIELDS) or dlp is not None or mrp is not None
            if not has_pricing:
                # full-width row → category header
                items.append({"type": "category", "name": md})
                continue

            margin = _to_float(_get(row, colmap, "distributor_margin_pct"))
            items.append({
                "type": "product",
                "material_description": md,
                "grade_specs": grade,
                "pack_size": pack,
                "mrp": mrp or 0.0,
                "dlp": dlp if dlp is not None else (mrp or 0.0),
                "distributor_margin_pct": margin or 0.0,
                "cash_coupon": _get(row, colmap, "cash_coupon"),
                "foc_benefits": _get(row, colmap, "foc_benefits"),
                "monthly_gift": _get(row, colmap, "monthly_gift"),
                "trade_discount": _get(row, colmap, "trade_discount"),
            })

    if not header_found:
        warnings.append(
            "Could not find a header row containing MATERIAL DESCRIPTION and DLP/MRP. "
            "Please make sure the first table row has those column titles.")

    products = [i for i in items if i["type"] == "product"]
    cats = [i for i in items if i["type"] == "category"]
    return {
        "source": src,
        "items": items,
        "product_count": len(products),
        "category_count": len(cats),
        "header_found": header_found,
        "warnings": warnings,
    }


def make_sku(material_description: str, pack_size: str, existing: set) -> str:
    """Generate a stable, unique-ish SKU from description + pack."""
    base = re.sub(r"[^A-Z0-9]+", "-", (material_description or "").upper()).strip("-")
    pack = re.sub(r"[^A-Z0-9]+", "", (pack_size or "").upper())
    sku = f"{base[:22]}-{pack}" if pack else base[:26]
    sku = sku.strip("-") or "PRODUCT"
    candidate = sku
    n = 2
    while candidate in existing:
        candidate = f"{sku}-{n}"
        n += 1
    existing.add(candidate)
    return candidate


def build_template_xlsx() -> bytes:
    """A ready-to-fill sample matching the GO OIL circular format."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"
    headers = ["MATERIAL DESCRIPTION", "GRADE/ SPECS", "PACK SIZE", "MRP", "DLP",
               "DISTRIBUTOR MARGINE", "CASH COUPON", "FOC BENEFITS",
               "MONTHLY GIFT", "TRADE DISCOUNT"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="C9A227")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    catfill = PatternFill("solid", fgColor="FFF3CD")

    def add_cat(name):
        ws.append([name])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
            c.fill = catfill

    def add_prod(row):
        ws.append(row)

    add_cat("MCO - SUPER CATEGORY")
    add_prod(["POWER 4T 15W50", "SN", "2.5 ltr", 1150, 845, "9%", "50 TO 100", "", "AVAILABLE", ""])
    add_prod(["POWER 4T 20W40", "SN", "1 ltr", 498, 334, "9%", "", "FOC 9+1", "AVAILABLE", ""])
    add_cat("GEAR OIL - GL5")
    add_prod(["GEAR GUARD 80W90", "GL5", "1 ltr", 442, 290, "9%", "", "", "", 50])

    # widths
    widths = [26, 12, 12, 8, 8, 16, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
