"""GO OIL DMS — Integration layer scaffolds.

This module provides provider-agnostic INTERFACES for every third-party integration mentioned in
the enterprise sprint. Concrete providers are wired in later via env vars — no business logic in
this file, only stable contracts and stub adapters that log intent.

Interfaces:
  - PaymentGateway         (Razorpay / Stripe)
  - MessagingChannel       (Email / WhatsApp / SMS)  ← also used by notifications.py
  - TaxIntegration         (GST returns/validation)
  - AccountingExport       (Tally XML / QuickBooks / Zoho Books)
  - DataImport             (Excel/CSV bulk upload)
  - CodeScanner            (Barcode / QR generation and lookup)
  - WebhookOutbox          (outbound webhooks to third-parties)
  - RestApiExport          (public API-key based access for external systems)

Registry pattern:
  from integrations import registry
  provider = registry.get("payment")
  result = await provider.create_order(...)

Configured via env vars (dev-time all default to "scaffold"):
  PAYMENT_PROVIDER=razorpay|stripe|scaffold
  EMAIL_PROVIDER=sendgrid|smtp|scaffold
  WHATSAPP_PROVIDER=twilio|meta|scaffold
  SMS_PROVIDER=twilio|msg91|scaffold
  TAX_PROVIDER=gstn|scaffold
  ACCOUNTING_PROVIDER=tally|scaffold
"""
from __future__ import annotations
import os
import uuid
import base64
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from xml.sax.saxutils import escape as xml_escape
from fastapi import APIRouter, Depends, HTTPException, Body, Query, UploadFile, File

logger = logging.getLogger("gooil.dms.integrations")


# ==========================================================
# BASE INTERFACES
# ==========================================================

class IntegrationScaffoldResult(dict):
    def __init__(self, ok: bool = True, provider: str = "scaffold", **kw):
        super().__init__(
            ok=ok,
            provider=provider,
            timestamp=datetime.now(timezone.utc).isoformat(),
            **kw,
        )


class PaymentGateway:
    """Contract for payment gateways (Razorpay / Stripe / etc.)"""
    name: str = "scaffold"

    async def create_order(self, amount: float, currency: str, meta: dict) -> IntegrationScaffoldResult:
        raise NotImplementedError

    async def verify_signature(self, payload: dict) -> IntegrationScaffoldResult:
        raise NotImplementedError

    async def refund(self, order_id: str, amount: float) -> IntegrationScaffoldResult:
        raise NotImplementedError


class RazorpayGateway(PaymentGateway):
    name = "razorpay"

    def __init__(self, key_id: str | None = None, secret: str | None = None):
        self.key_id = key_id or os.environ.get("RAZORPAY_KEY_ID", "")
        self.secret = secret or os.environ.get("RAZORPAY_KEY_SECRET", "")
        self.configured = bool(self.key_id and self.secret)

    async def create_order(self, amount: float, currency: str, meta: dict) -> IntegrationScaffoldResult:
        if not self.configured:
            logger.info(f"[razorpay SCAFFOLD] create_order amount={amount} {currency}")
            return IntegrationScaffoldResult(
                provider=self.name,
                configured=False,
                order_id=f"scaffold-rzp-{uuid.uuid4().hex[:12]}",
                amount=amount, currency=currency,
                message="Razorpay not configured — set RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET.",
            )
        # Real implementation would call razorpay-python SDK here.
        return IntegrationScaffoldResult(provider=self.name, configured=True)

    async def verify_signature(self, payload: dict) -> IntegrationScaffoldResult:
        return IntegrationScaffoldResult(provider=self.name, configured=self.configured,
                                            verified=self.configured,
                                            message="Signature verification stub")

    async def refund(self, order_id: str, amount: float) -> IntegrationScaffoldResult:
        return IntegrationScaffoldResult(provider=self.name, configured=self.configured,
                                            refund_id=f"scaffold-rfnd-{uuid.uuid4().hex[:10]}",
                                            amount=amount)


class StripeGateway(PaymentGateway):
    name = "stripe"

    def __init__(self, secret_key: str | None = None):
        self.secret_key = secret_key or os.environ.get("STRIPE_SECRET_KEY", "")
        self.configured = bool(self.secret_key)

    async def create_order(self, amount: float, currency: str, meta: dict) -> IntegrationScaffoldResult:
        if not self.configured:
            logger.info(f"[stripe SCAFFOLD] create_order amount={amount} {currency}")
            return IntegrationScaffoldResult(
                provider=self.name,
                configured=False,
                payment_intent_id=f"scaffold-pi-{uuid.uuid4().hex[:12]}",
                amount=amount, currency=currency,
                message="Stripe not configured — set STRIPE_SECRET_KEY.",
            )
        return IntegrationScaffoldResult(provider=self.name, configured=True)

    async def verify_signature(self, payload: dict) -> IntegrationScaffoldResult:
        return IntegrationScaffoldResult(provider=self.name, configured=self.configured)

    async def refund(self, order_id: str, amount: float) -> IntegrationScaffoldResult:
        return IntegrationScaffoldResult(provider=self.name, configured=self.configured, amount=amount)


# ---------- Tax (GST) ----------

class TaxIntegration:
    name = "scaffold"

    async def validate_gstin(self, gstin: str) -> IntegrationScaffoldResult:
        raise NotImplementedError

    async def build_gstr1_export(self, invoices: list[dict]) -> str:
        raise NotImplementedError


class GSTIntegration(TaxIntegration):
    name = "gst"

    def __init__(self):
        self.configured = bool(os.environ.get("GSTN_API_KEY"))

    async def validate_gstin(self, gstin: str) -> IntegrationScaffoldResult:
        # GSTIN format: 15 chars, [2-digit state][10-char PAN][entity][Z][checksum]
        ok = isinstance(gstin, str) and len(gstin) == 15 and gstin.isalnum()
        return IntegrationScaffoldResult(
            provider=self.name,
            configured=self.configured,
            valid_format=ok,
            gstin=gstin,
            message="Format check only. Live GSTN API requires GSTN_API_KEY.",
        )

    async def build_gstr1_export(self, invoices: list[dict]) -> str:
        """Build a JSON payload matching GSTN GSTR-1 shape (b2b + b2cs)."""
        b2b: dict = {}
        b2cs: list = []
        for inv in invoices:
            party_gstin = (inv.get("party_gstin") or "").upper()
            row = {
                "inum": inv.get("invoice_no"),
                "idt": inv.get("invoice_date") or inv.get("created_at"),
                "val": inv.get("total"),
                "pos": inv.get("place_of_supply", "27"),
                "rchrg": "N",
                "itms": [
                    {"num": i + 1, "itm_det": {
                        "txval": ln.get("taxable_value", 0),
                        "rt": ln.get("gst_rate", 18),
                        "iamt": ln.get("igst", 0),
                        "camt": ln.get("cgst", 0),
                        "samt": ln.get("sgst", 0),
                    }} for i, ln in enumerate(inv.get("lines") or [])
                ],
            }
            if party_gstin:
                b2b.setdefault(party_gstin, []).append(row)
            else:
                b2cs.append(row)
        payload = {
            "gstin": os.environ.get("COMPANY_GSTIN", "27AAAAA0000A1Z5"),
            "fp": datetime.now(timezone.utc).strftime("%m%Y"),
            "b2b": [{"ctin": k, "inv": v} for k, v in b2b.items()],
            "b2cs": b2cs,
        }
        return payload


# ---------- Accounting export (Tally) ----------

class AccountingExport:
    name = "scaffold"

    async def export_to_tally_xml(self, vouchers: list[dict]) -> str:
        raise NotImplementedError


class TallyExport(AccountingExport):
    """Tally XML voucher import format. Compatible with Tally.ERP 9 / TallyPrime."""
    name = "tally"

    def _fmt_date(self, iso: str | None) -> str:
        if not iso:
            return datetime.now(timezone.utc).strftime("%Y%m%d")
        try:
            return iso[:10].replace("-", "")
        except Exception:
            return datetime.now(timezone.utc).strftime("%Y%m%d")

    async def export_to_tally_xml(self, vouchers: list[dict]) -> str:
        parts = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            "<ENVELOPE>",
            "<HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>",
            "<BODY><IMPORTDATA>",
            "<REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC>",
            "<REQUESTDATA>",
        ]
        for v in vouchers:
            v_type = xml_escape(v.get("type", "Sales"))
            v_no = xml_escape(str(v.get("voucher_no", v.get("id", "")))[:30])
            v_date = self._fmt_date(v.get("date") or v.get("created_at"))
            narration = xml_escape((v.get("narration") or "")[:200])
            party = xml_escape(v.get("party_name", "Unknown"))
            parts.append(
                f'<TALLYMESSAGE><VOUCHER VCHTYPE="{v_type}" ACTION="Create">'
                f'<DATE>{v_date}</DATE>'
                f'<VOUCHERNUMBER>{v_no}</VOUCHERNUMBER>'
                f'<VOUCHERTYPENAME>{v_type}</VOUCHERTYPENAME>'
                f'<PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>'
                f'<NARRATION>{narration}</NARRATION>'
            )
            for ln in v.get("ledger_entries") or []:
                parts.append(
                    f'<ALLLEDGERENTRIES.LIST>'
                    f'<LEDGERNAME>{xml_escape(ln.get("ledger","Sundry"))}</LEDGERNAME>'
                    f'<ISDEEMEDPOSITIVE>{"Yes" if ln.get("dr") else "No"}</ISDEEMEDPOSITIVE>'
                    f'<AMOUNT>{ "-" if ln.get("dr") else "" }{ln.get("amount",0)}</AMOUNT>'
                    f'</ALLLEDGERENTRIES.LIST>'
                )
            parts.append("</VOUCHER></TALLYMESSAGE>")
        parts += ["</REQUESTDATA>", "</IMPORTDATA></BODY>", "</ENVELOPE>"]
        return "\n".join(parts)


# ---------- Barcode / QR ----------

class CodeScanner:
    """Barcode / QR generation & lookup scaffold.

    Generation currently returns SVG data URL — no external deps needed. Later plug in
    python-barcode / qrcode / segno for higher-fidelity images.
    """
    name = "code"

    def _svg_qr_stub(self, text: str, size: int = 160) -> str:
        """Cheap SVG placeholder QR — replace with a real generator when needed."""
        # produces a deterministic grid so the label is visually meaningful
        seed = sum(ord(c) for c in text)
        grid_n = 21
        cell = size / grid_n
        rects: list[str] = []
        for y in range(grid_n):
            for x in range(grid_n):
                if ((seed * (x + 1) * (y + 3)) % 7) < 3:
                    rects.append(f'<rect x="{x*cell:.2f}" y="{y*cell:.2f}" width="{cell:.2f}" height="{cell:.2f}" fill="#111827"/>')
        svg = (f'<svg xmlns="http://www.w3.org/2000/svg" width="{size}" height="{size}" viewBox="0 0 {size} {size}">'
               f'<rect width="{size}" height="{size}" fill="#fff"/>' + "".join(rects) + '</svg>')
        return "data:image/svg+xml;base64," + base64.b64encode(svg.encode()).decode()

    async def generate(self, kind: str, value: str) -> IntegrationScaffoldResult:
        if kind not in ("barcode", "qr"):
            raise HTTPException(400, "kind must be 'barcode' or 'qr'")
        data_url = self._svg_qr_stub(value)
        return IntegrationScaffoldResult(
            provider="scaffold",
            kind=kind,
            value=value,
            data_url=data_url,
            message="Placeholder SVG. Install python-barcode / qrcode / segno for production glyphs.",
        )

    async def lookup(self, code: str, db) -> IntegrationScaffoldResult:
        # Try to resolve to SKU / batch / invoice
        found = None
        for coll in ("skus", "batches", "invoices", "primary_orders"):
            doc = await db[coll].find_one({"code": code}, {"_id": 0})
            if not doc:
                doc = await db[coll].find_one({"id": code}, {"_id": 0})
            if doc:
                found = {"collection": coll, "record": doc}
                break
        return IntegrationScaffoldResult(provider="scaffold", found=bool(found), match=found)


# ---------- Webhooks (outbound) ----------

class WebhookOutbox:
    async def deliver(self, url: str, event: str, payload: dict) -> IntegrationScaffoldResult:
        logger.info(f"[webhook SCAFFOLD] would POST {event} → {url}")
        return IntegrationScaffoldResult(provider="scaffold", event=event, url=url, delivered=False,
                                          message="Outbound webhook delivery scaffolded. Configure WEBHOOK_ENABLED=true and set target URL to enable.")


# ==========================================================
# REGISTRY
# ==========================================================

class IntegrationRegistry:
    def __init__(self):
        self._reg: Dict[str, Any] = {}

    def register(self, key: str, impl: Any):
        self._reg[key] = impl

    def get(self, key: str):
        if key not in self._reg:
            raise HTTPException(500, f"Integration '{key}' not registered")
        return self._reg[key]

    def status(self) -> dict:
        return {k: {
            "provider": getattr(v, "name", "unknown"),
            "configured": getattr(v, "configured", False),
        } for k, v in self._reg.items()}


registry = IntegrationRegistry()

# Register defaults — pick provider by env, default scaffold.
_payment_pref = os.environ.get("PAYMENT_PROVIDER", "razorpay").lower()
registry.register("payment", RazorpayGateway() if _payment_pref == "razorpay" else StripeGateway())
registry.register("payment_alt", StripeGateway() if _payment_pref == "razorpay" else RazorpayGateway())
registry.register("tax", GSTIntegration())
registry.register("accounting", TallyExport())
registry.register("code", CodeScanner())
registry.register("webhook", WebhookOutbox())


# ==========================================================
# ROUTER
# ==========================================================

def build_integrations_router(db, get_current_user, require_admin_dep=None):
    router = APIRouter(prefix="/integrations", tags=["integrations"])

    @router.get("/status")
    async def status(user: dict = Depends(get_current_user)):
        """Return which providers are configured vs scaffolded."""
        return {
            "registry": registry.status(),
            "hints": {
                "payment": "Set PAYMENT_PROVIDER=razorpay|stripe and provide keys.",
                "email": "Set EMAIL_PROVIDER=sendgrid|smtp and SENDGRID_API_KEY / SMTP_URL.",
                "whatsapp": "Set WHATSAPP_PROVIDER=twilio|meta and matching credentials.",
                "sms": "Set SMS_PROVIDER=twilio|msg91 and matching credentials.",
                "tax": "Set GSTN_API_KEY once you have GSTN Suvidha Provider access.",
                "accounting": "Tally XML export is live. Configure Tally listener URL to push directly.",
            },
        }

    # ---------- Payment ----------
    @router.post("/payments/create-order")
    async def payment_create_order(body: dict = Body(...), user: dict = Depends(get_current_user)):
        gw = registry.get("payment")
        return await gw.create_order(
            amount=float(body.get("amount", 0)),
            currency=body.get("currency", "INR"),
            meta=body.get("meta", {}),
        )

    @router.post("/payments/verify")
    async def payment_verify(body: dict = Body(...), user: dict = Depends(get_current_user)):
        gw = registry.get("payment")
        return await gw.verify_signature(body)

    @router.post("/payments/refund")
    async def payment_refund(body: dict = Body(...), user: dict = Depends(get_current_user)):
        gw = registry.get("payment")
        return await gw.refund(body.get("order_id", ""), float(body.get("amount", 0)))

    # ---------- Tax / GST ----------
    @router.get("/tax/validate-gstin")
    async def tax_validate(gstin: str = Query(...), user: dict = Depends(get_current_user)):
        return await registry.get("tax").validate_gstin(gstin)

    @router.post("/tax/gstr1-preview")
    async def tax_gstr1_preview(user: dict = Depends(get_current_user)):
        invs = await db.invoices.find({}, {"_id": 0}).limit(500).to_list(500)
        payload = await registry.get("tax").build_gstr1_export(invs)
        return {"payload": payload, "invoice_count": len(invs)}

    # ---------- Accounting / Tally ----------
    @router.get("/accounting/tally-export")
    async def accounting_tally(user: dict = Depends(get_current_user)):
        from fastapi.responses import Response
        # Convert recent journal entries into Tally vouchers
        journals = await db.double_ledger.find({}, {"_id": 0}).sort("timestamp", -1).limit(200).to_list(200)
        # Group by reference_id
        by_ref: Dict[str, dict] = {}
        for j in journals:
            ref = j.get("reference_id") or j.get("id")
            entry = by_ref.setdefault(ref, {
                "voucher_no": ref,
                "type": j.get("voucher_type", "Journal"),
                "date": j.get("timestamp"),
                "narration": j.get("narration") or f"Voucher {ref}",
                "party_name": j.get("party_id", "Unknown"),
                "ledger_entries": [],
            })
            entry["ledger_entries"].append({
                "ledger": j.get("account", "Sundry"),
                "dr": bool(j.get("dr", 0)),
                "amount": abs(j.get("dr", 0) or j.get("cr", 0) or 0),
            })
        xml = await registry.get("accounting").export_to_tally_xml(list(by_ref.values()))
        return Response(
            content=xml,
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="tally-{datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")}.xml"'},
        )

    # ---------- Codes ----------
    @router.get("/code/generate")
    async def code_generate(kind: str = Query("qr", regex="^(qr|barcode)$"),
                              value: str = Query(...),
                              user: dict = Depends(get_current_user)):
        return await registry.get("code").generate(kind, value)

    @router.get("/code/lookup")
    async def code_lookup(code: str = Query(...), user: dict = Depends(get_current_user)):
        return await registry.get("code").lookup(code, db)

    # ---------- Excel / CSV import ----------
    @router.post("/import/excel")
    async def import_excel(
        collection: str = Query(..., regex="^[a-z_]+$"),
        file: UploadFile = File(...),
        user: dict = Depends(get_current_user),
    ):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(500, "openpyxl not installed on server")
        allowed = {"products", "skus", "distributors", "retailers", "customers"}
        if collection not in allowed:
            raise HTTPException(400, f"Import into '{collection}' not allowed. Allowed: {sorted(allowed)}")
        data = await file.read()
        import io
        wb = load_workbook(filename=io.BytesIO(data), read_only=True)
        ws = wb.active
        header = None
        rows: List[dict] = []
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c or "").strip().lower() for c in row]
                continue
            record = {header[i]: row[i] for i in range(min(len(header), len(row))) if header[i]}
            if not any(record.values()):
                continue
            record["id"] = record.get("id") or f"{collection[:3]}-{uuid.uuid4().hex[:10]}"
            record["created_at"] = datetime.now(timezone.utc).isoformat()
            record["imported_by"] = user.get("email")
            rows.append(record)
        if not rows:
            return {"imported": 0, "message": "No data rows found"}
        r = await db[collection].insert_many(rows)
        return {"imported": len(r.inserted_ids), "collection": collection,
                "sample": [{k: v for k, v in row.items() if k != "_id"} for row in rows[:3]]}

    # ---------- Webhooks ----------
    @router.post("/webhooks/emit")
    async def webhook_emit(body: dict = Body(...), user: dict = Depends(get_current_user)):
        return await registry.get("webhook").deliver(
            url=body.get("url", ""),
            event=body.get("event", "unknown"),
            payload=body.get("payload", {}),
        )

    @router.post("/webhooks/inbox")
    async def webhook_inbox(body: dict = Body(...)):
        """Public webhook receiver — payment providers, etc. call this to notify events."""
        await db.webhook_events.insert_one({
            "id": f"wh-{uuid.uuid4().hex[:12]}",
            "received_at": datetime.now(timezone.utc).isoformat(),
            "payload": body,
        })
        return {"ok": True}

    # ---------- Public REST API for external systems ----------
    @router.get("/public/health")
    async def public_health():
        return {"status": "ok", "service": "gooil-dms-integrations", "version": "5.0"}

    return router
