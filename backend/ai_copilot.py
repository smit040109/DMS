"""GO OIL DMS — AI Business Copilot.

A business analyst assistant that answers analytical questions by:
  1. Detecting the user's intent (executive / sales / finance / inventory / party / general).
  2. Fetching the appropriate structured context from analytics endpoints (already computed in Phase 4).
  3. Passing the context + question to the LLM (Emergent Universal Key by default).
  4. Returning: {answer, sources: [{endpoint, key_numbers}], model, tokens, session_id}.

Sessions:
  - Multi-turn conversation history stored in `ai_copilot_sessions` keyed by session_id.
  - Frontend controls session_id lifetime (new session per page, or per user).

Providers:
  - Default provider/model: openai / gpt-5.4 (per verified playbook).
  - Selectable via env vars AI_PROVIDER + AI_MODEL, or per-request 'model' param.
  - Requires EMERGENT_LLM_KEY in env. If missing, returns 503 with a helpful message.

Endpoints (prefix /api/ai/copilot):
  POST /ask               synchronous ask (non-streaming) → recommended for MVP
  GET  /suggestions       list of pre-canned executive questions
  GET  /sessions/{id}     fetch history
  DELETE /sessions/{id}   reset session
"""
from __future__ import annotations
import os
import json
import uuid
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Body
from pydantic import BaseModel

logger = logging.getLogger("gooil.dms.ai_copilot")

# Emergent SDK (optional at import — degrade if missing)
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    _EI_OK = True
except Exception as e:
    _EI_OK = False
    _EI_IMPORT_ERROR = str(e)


DEFAULT_PROVIDER = os.environ.get("AI_PROVIDER", "openai")
DEFAULT_MODEL = os.environ.get("AI_MODEL", "gpt-5.4")

SYSTEM_PROMPT = """You are the GO OIL DMS Assistant — an in-app helper for a Distribution
Management System covering oil & lubricant sales across owner, distributors, retailers,
salespersons and managers.

WHO YOU ARE TALKING TO:
- User: {user_name} (role: {user_role}). Current date: {today}.
- The CONTEXT block below contains ONLY this user's own data, already scoped to what their
  role is allowed to see. Answer strictly from it. Never reveal or infer other users' private data.

RULES:
1. Be helpful and friendly, like a smart colleague. Plain language. If the user writes in Hindi
   or Hinglish, reply in the same style.
2. Two kinds of questions:
   a) "How do I…/what is…" app-usage questions → give clear step-by-step guidance about the app.
   b) "My numbers / reports" questions → answer using ONLY the numbers in CONTEXT.
3. Never invent numbers. If CONTEXT lacks something, say so plainly and point to the exact page/report
   in the app where they can find it.
4. Money in Indian Rupees (₹) with commas (e.g. ₹1,25,000). Percentages to one decimal.
5. When asked for a "report" or "summary", produce a compact, well-structured report: a short
   headline, then bullet KPIs, then a small ranked list if relevant. Keep it accurate to CONTEXT.
6. Use bullet lists for scans; short paragraphs for explanations.
7. Close data answers with a one-line "Next best action:" suggestion. Skip that for pure how-to answers.
"""

# Canned executive suggestions the frontend can display.
SUGGESTIONS = [
    "Give me today's summary",
    "How do I add a new sale?",
    "Which retailers haven't ordered recently?",
    "What is my outstanding amount?",
    "Show my top selling products",
    "How do I download a report?",
]

# Role-aware starter questions (mix of "how do I" + "my data").
ROLE_SUGGESTIONS = {
    "owner": [
        "Give me today's business summary",
        "Which distributor has the highest outstanding?",
        "Which products are low on stock?",
        "How much did we collect this month?",
        "How do I create a new price circular?",
    ],
    "owner_accountant": [
        "What is total outstanding across distributors?",
        "Show this month's expenses summary",
        "How do I record a payment?",
        "Which cheques are pending?",
    ],
    "distributor": [
        "Kitna maal maine is mahine order kiya?",
        "Mera outstanding kitna hai?",
        "Which of my retailers order the most?",
        "How do I add a sale bill?",
        "How do I update my bank / UPI details?",
    ],
    "distributor_accountant": [
        "Show my primary and secondary ledger summary",
        "What is my current outstanding?",
        "How do I record a retailer payment?",
    ],
    "retailer": [
        "What is my wallet balance?",
        "Show my recent orders",
        "How do I place a new order?",
        "How do I redeem my coupons?",
    ],
    "salesperson": [
        "Aaj maine kitni visit ki?",
        "Show my attendance for today",
        "How much did I collect today?",
        "How do I add a new retailer?",
        "Which retailers should I visit next?",
    ],
    "team_leader": [
        "How is my team performing today?",
        "Which salesperson has the best sales?",
        "Show today's attendance of my team",
    ],
    "regional_manager": [
        "Give me my region's performance summary",
        "Which team leader is underperforming?",
        "Show attendance across my region",
    ],
}

# Intent → analytics-context to fetch.
INTENT_MATRIX = {
    "sales":     ["sales", "executive"],
    "finance":   ["finance", "executive"],
    "inventory": ["inventory", "executive"],
    "expiry":    ["inventory"],
    "returns":   ["executive"],
    "approvals": ["executive"],
    "executive": ["executive"],
    "general":   ["executive"],
}

_INTENT_KEYWORDS = {
    "sales":     ["sales", "revenue", "top skus", "best selling", "primary order", "secondary order", "growth"],
    "finance":   ["outstanding", "receivable", "aging", "collection", "cash", "wallet", "coupon", "cashback", "ledger", "reconcile"],
    "inventory": ["stock", "inventory", "reorder", "batch", "warehouse"],
    "expiry":    ["expiry", "expiring", "near expiry", "shelf life"],
    "returns":   ["return", "damage", "claim", "replacement", "credit note"],
    "approvals": ["approval", "pending", "waiting", "escalation"],
}


def _detect_intent(q: str) -> str:
    ql = q.lower()
    scores = {intent: sum(1 for kw in kws if kw in ql) for intent, kws in _INTENT_KEYWORDS.items()}
    best = max(scores.items(), key=lambda kv: kv[1])
    return best[0] if best[1] > 0 else "executive"


def _summarise_context(ctx: dict) -> dict:
    """Extract the top 12 numbers from a context blob so the LLM sees only what matters."""
    summary: dict = {}
    if not isinstance(ctx, dict):
        return {"note": "empty"}
    scope = ctx.get("scope")
    if scope == "executive":
        summary["kpis"] = ctx.get("summary", {})
        summary["alerts_by_severity"] = ctx.get("alerts_summary", {})
        summary["top_alerts"] = ctx.get("recent_alerts", [])[:5]
    elif scope == "sales":
        summary["totals"] = ctx.get("totals") or ctx.get("summary") or {}
        summary["top_skus"] = (ctx.get("top_skus") or ctx.get("top") or [])[:5]
        summary["top_distributors"] = (ctx.get("top_distributors") or [])[:5]
        summary["timeseries_tail"] = (ctx.get("timeseries") or ctx.get("series") or [])[-7:]
    elif scope == "finance":
        summary["ar_aging"] = ctx.get("ar_aging") or ctx.get("aging") or {}
        summary["totals"] = ctx.get("totals") or ctx.get("summary") or {}
        summary["top_outstanding"] = (ctx.get("top_outstanding") or [])[:5]
    elif scope == "inventory":
        summary["buckets"] = ctx.get("buckets") or ctx.get("totals") or {}
        summary["expiring"] = (ctx.get("expiring") or ctx.get("near_expiry") or [])[:8]
        summary["low_stock"] = (ctx.get("low_stock") or [])[:8]
    else:
        summary = ctx
    return summary


class AskIn(BaseModel):
    question: str
    session_id: Optional[str] = None
    provider: Optional[str] = None
    model: Optional[str] = None


def build_ai_copilot_router(db, get_current_user, analytics_router, dms_router=None):
    router = APIRouter(prefix="/ai/copilot", tags=["ai-copilot"])

    # Locate the ai-context callable on the analytics router — cheaper than HTTP loopback.
    _ai_context_endpoint = None
    for r in analytics_router.routes:
        if getattr(r, "path", "").endswith("/ai-context/{scope}"):
            _ai_context_endpoint = r.endpoint
            break

    # Locate DMS dashboard callables so the copilot can answer using the logged-in
    # user's OWN role-scoped DMS data (accurate, reuses tested dashboard logic).
    _dms_dash = {}
    if dms_router is not None:
        for r in dms_router.routes:
            p = getattr(r, "path", "")
            if "/dashboard/" in p:
                _dms_dash[p.split("/dashboard/")[-1]] = r.endpoint

    # role -> list of dashboard keys to pull for that user
    _ROLE_DASH = {
        "owner": ["owner", "finance-snapshot"],
        "owner_accountant": ["owner", "finance-snapshot"],
        "super_admin": ["owner", "finance-snapshot"],
        "company_admin": ["owner", "finance-snapshot"],
        "distributor": ["distributor"],
        "distributor_accountant": ["distributor"],
        "retailer": ["retailer"],
        "salesperson": ["salesperson"],
        "team_leader": ["team-leader"],
        "regional_manager": ["regional-manager"],
    }

    async def _gather_dms_context(user: dict) -> List[dict]:
        """Pull the logged-in user's own role-scoped DMS snapshot(s)."""
        collected: List[dict] = []
        role = user.get("role", "")
        for key in _ROLE_DASH.get(role, ["owner"]):
            fn = _dms_dash.get(key)
            if not fn:
                continue
            try:
                data = await fn(user=user)
                collected.append({
                    "scope": f"dms:{key}",
                    "endpoint": f"/api/dms/dashboard/{key}",
                    "context": data,
                })
            except Exception as e:
                logger.warning(f"dms dashboard/{key} failed: {e}")
        return collected

    async def _gather_context(intents: List[str], user: dict) -> List[dict]:
        collected: List[dict] = []
        for intent in intents:
            scope = intent if intent in ("executive", "sales", "finance", "inventory") else "executive"
            try:
                if _ai_context_endpoint:
                    ctx = await _ai_context_endpoint(scope=scope, user=user)
                    collected.append({
                        "scope": scope,
                        "endpoint": f"/api/analytics/ai-context/{scope}",
                        "context": ctx,
                    })
            except Exception as e:
                logger.warning(f"ai-context/{scope} failed: {e}")
        return collected

    def _extract_sources(gathered: List[dict]) -> List[dict]:
        out = []
        for g in gathered:
            summ = _summarise_context(g.get("context") or {})
            key_numbers = {}
            # try to pull 3-5 headline numbers
            if "kpis" in summ and isinstance(summ["kpis"], dict):
                for k, v in list(summ["kpis"].items())[:5]:
                    if isinstance(v, (int, float, str)):
                        key_numbers[k] = v
            elif "totals" in summ and isinstance(summ["totals"], dict):
                for k, v in list(summ["totals"].items())[:5]:
                    if isinstance(v, (int, float, str)):
                        key_numbers[k] = v
            out.append({
                "endpoint": g["endpoint"],
                "scope": g["scope"],
                "key_numbers": key_numbers,
            })
        return out

    async def _load_history(session_id: str) -> List[dict]:
        s = await db.ai_copilot_sessions.find_one({"id": session_id}, {"_id": 0})
        return (s or {}).get("history", [])

    async def _save_history(session_id: str, user_id: str, history: List[dict]) -> None:
        now = datetime.now(timezone.utc).isoformat()
        await db.ai_copilot_sessions.update_one(
            {"id": session_id},
            {"$set": {"history": history, "user_id": user_id, "updated_at": now},
             "$setOnInsert": {"id": session_id, "created_at": now}},
            upsert=True,
        )

    @router.get("/suggestions")
    async def suggestions(user: dict = Depends(get_current_user)):
        return {"data": ROLE_SUGGESTIONS.get(user.get("role", ""), SUGGESTIONS)}

    @router.get("/status")
    async def status(user: dict = Depends(get_current_user)):
        key_set = bool(os.environ.get("EMERGENT_LLM_KEY"))
        return {
            "sdk_available": _EI_OK,
            "key_configured": key_set,
            "provider": DEFAULT_PROVIDER,
            "model": DEFAULT_MODEL,
            "ready": _EI_OK and key_set,
            "message": (
                "Ready" if _EI_OK and key_set
                else "AI Copilot requires EMERGENT_LLM_KEY in backend/.env "
                     "(get one from Emergent → Profile → Universal Key)."
            ),
        }

    @router.get("/sessions/{session_id}")
    async def get_session(session_id: str, user: dict = Depends(get_current_user)):
        s = await db.ai_copilot_sessions.find_one({"id": session_id, "user_id": user["id"]}, {"_id": 0})
        if not s:
            return {"id": session_id, "history": []}
        return s

    @router.delete("/sessions/{session_id}")
    async def delete_session(session_id: str, user: dict = Depends(get_current_user)):
        r = await db.ai_copilot_sessions.delete_one({"id": session_id, "user_id": user["id"]})
        return {"ok": True, "deleted": r.deleted_count}

    @router.post("/ask")
    async def ask(body: AskIn, user: dict = Depends(get_current_user)):
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not _EI_OK:
            raise HTTPException(
                503,
                f"emergentintegrations SDK unavailable: {globals().get('_EI_IMPORT_ERROR', 'unknown')}",
            )
        if not api_key:
            raise HTTPException(
                503,
                "EMERGENT_LLM_KEY not configured. Please add it to backend/.env "
                "and restart. Get your key from Emergent → Profile → Universal Key.",
            )

        question = (body.question or "").strip()
        if not question:
            raise HTTPException(400, "Question is required")

        session_id = body.session_id or f"sess-{uuid.uuid4().hex[:12]}"
        provider = (body.provider or DEFAULT_PROVIDER).lower()
        model = body.model or DEFAULT_MODEL

        # 1. Gather structured context — prefer the user's OWN DMS data.
        intent = _detect_intent(question)
        gathered = []
        if _dms_dash:
            gathered = await _gather_dms_context(user)
        if not gathered:
            scopes = INTENT_MATRIX.get(intent, INTENT_MATRIX["executive"])
            gathered = await _gather_context(scopes, user)
        context_bundle = {
            "user": {"name": user.get("name"), "role": user.get("role"),
                        "email": user.get("email")},
            "date": datetime.now(timezone.utc).isoformat(),
            "detected_intent": intent,
            "contexts": [{"scope": g["scope"], "data": _summarise_context(g["context"])}
                          for g in gathered],
        }

        # 2. Build the LLM chat
        system_message = SYSTEM_PROMPT.format(
            user_role=user.get("role", "executive"),
            user_name=user.get("name", "there"),
            today=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        )
        try:
            chat = LlmChat(
                api_key=api_key,
                session_id=session_id,
                system_message=system_message,
            ).with_model(provider, model)
        except Exception as e:
            raise HTTPException(500, f"Failed to initialise LlmChat ({provider}/{model}): {e}")

        # 3. Load prior history and replay (LlmChat is stateless across calls)
        history = await _load_history(session_id)
        # LlmChat replay is done implicitly through session id in some SDKs;
        # to be safe we prepend prior turns to the current user message.

        prior_transcript = ""
        for turn in history[-6:]:  # last 3 exchanges
            role = turn.get("role", "user")
            prior_transcript += f"\n[{role.upper()}]: {turn.get('content','')}\n"

        composed = (
            f"{prior_transcript}\n"
            f"CONTEXT (structured business snapshot for your reference):\n"
            f"{json.dumps(context_bundle, default=str)[:8000]}\n\n"
            f"USER QUESTION: {question}"
        )

        # 4. Non-streaming call (per user pref "lean & real")
        try:
            resp = await chat.send_message(UserMessage(text=composed))
        except AttributeError:
            # Some SDK versions expose only stream_message — fall back to collecting text.
            try:
                from emergentintegrations.llm.chat import TextDelta, StreamDone
                chunks: List[str] = []
                async for ev in chat.stream_message(UserMessage(text=composed)):
                    if isinstance(ev, TextDelta):
                        chunks.append(ev.content)
                    elif isinstance(ev, StreamDone):
                        break
                resp = "".join(chunks)
            except Exception as e:
                raise HTTPException(500, f"LLM call failed: {e}")
        except Exception as e:
            raise HTTPException(500, f"LLM call failed: {e}")

        # LlmChat.send_message returns either a string, or a response object with .content.
        if isinstance(resp, str):
            answer_text = resp
            usage = None
        else:
            answer_text = getattr(resp, "content", None) or getattr(resp, "text", None) or str(resp)
            usage = getattr(resp, "usage", None)

        # 5. Persist history
        history = history + [
            {"role": "user", "content": question,
             "timestamp": datetime.now(timezone.utc).isoformat()},
            {"role": "assistant", "content": answer_text,
             "timestamp": datetime.now(timezone.utc).isoformat(),
             "model": f"{provider}/{model}"},
        ]
        await _save_history(session_id, user["id"], history[-40:])  # cap history

        return {
            "session_id": session_id,
            "answer": answer_text,
            "sources": _extract_sources(gathered),
            "intent": intent,
            "model": f"{provider}/{model}",
            "usage": usage,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

    @router.post("/export")
    async def export_report(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        """Turn any AI answer/report into a downloadable PDF or Excel file."""
        from fastapi.responses import StreamingResponse
        import io, re as _re

        fmt = (body.get("format") or "pdf").lower()
        title = (body.get("title") or "GO OIL DMS — AI Report").strip()
        content = (body.get("content") or "").strip()
        if not content:
            raise HTTPException(400, "content is required")

        # strip simple markdown so exports read cleanly
        def _plain(s: str) -> str:
            s = _re.sub(r"\*\*(.+?)\*\*", r"\1", s)
            s = _re.sub(r"[*`_#>]", "", s)
            return s.strip()

        lines = [_plain(ln) for ln in content.splitlines()]
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        who = f"{user.get('name','')} ({user.get('role','')})"

        if fmt in ("excel", "xlsx"):
            from openpyxl import Workbook
            from openpyxl.styles import Font
            wb = Workbook(); ws = wb.active; ws.title = "AI Report"
            ws["A1"] = title; ws["A1"].font = Font(size=14, bold=True)
            ws["A2"] = f"Generated for {who} · {stamp}"; ws["A2"].font = Font(size=9, italic=True)
            r = 4
            for ln in lines:
                if ln.strip():
                    ws.cell(row=r, column=1, value=ln)
                r += 1
            ws.column_dimensions["A"].width = 100
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": 'attachment; filename="ai_report.xlsx"'},
            )

        # default: PDF
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm,
                                leftMargin=18*mm, rightMargin=18*mm)
        styles = getSampleStyleSheet()
        h = ParagraphStyle("h", parent=styles["Title"], fontSize=16, textColor="#8a6600")
        meta = ParagraphStyle("meta", parent=styles["Normal"], fontSize=8, textColor="#888888")
        body_s = ParagraphStyle("b", parent=styles["Normal"], fontSize=10.5, leading=15)
        story = [Paragraph(title, h), Paragraph(f"Generated for {who} &middot; {stamp}", meta), Spacer(1, 8)]
        for ln in lines:
            if ln.strip():
                story.append(Paragraph(ln.replace("&", "&amp;"), body_s))
            else:
                story.append(Spacer(1, 6))
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="ai_report.pdf"'},
        )

    return router
