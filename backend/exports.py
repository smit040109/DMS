"""GO OIL DMS — Enterprise export layer (CSV / Excel / PDF / Print HTML).

Design:
- Single provider-agnostic `render_export(rows, format, columns, title)` function.
- Endpoint `POST /api/exports/render` accepts arbitrary tabular data + format and returns a file stream.
- Endpoint `GET /api/exports/{resource}` renders any collection to the requested format on-the-fly.
- Reused by Reports Hub, Analytics pages, DataTable component (via a shared UI button).

Formats:
    csv   — text/csv                   (rfc 4180)
    xlsx  — application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    pdf   — application/pdf            (reportlab landscape table)
    print — text/html                  (print-friendly styled HTML)
"""
from __future__ import annotations
import io
import csv
import html
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Body, Query, Response
from fastapi.responses import StreamingResponse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _XLSX_OK = True
except Exception:
    _XLSX_OK = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    _PDF_OK = True
except Exception:
    _PDF_OK = False


# ---------- helpers ----------

def _infer_columns(rows: List[Dict[str, Any]], max_cols: int = 12) -> List[str]:
    seen: List[str] = []
    for r in rows[:20]:
        for k in r.keys():
            if k not in seen and not k.startswith("_"):
                seen.append(k)
    return seen[:max_cols]


def _fmt_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        try:
            import json
            return json.dumps(v, default=str)[:200]
        except Exception:
            return str(v)[:200]
    if isinstance(v, float):
        return f"{v:,.2f}"
    return str(v)


def to_csv(rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow(columns)
    for row in rows:
        w.writerow([_fmt_value(row.get(c)) for c in columns])
    return buf.getvalue().encode("utf-8-sig")  # utf-8 BOM so Excel reads unicode


def to_xlsx(rows: List[Dict[str, Any]], columns: List[str], title: str = "Export") -> bytes:
    if not _XLSX_OK:
        raise HTTPException(500, "openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Export"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    ws.append(columns)
    for i, _ in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append([_fmt_value(row.get(c)) for c in columns])
    # auto-fit columns (approx)
    for i, col in enumerate(columns, start=1):
        max_len = max([len(str(col))] + [len(_fmt_value(r.get(col))) for r in rows[:500]])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(rows: List[Dict[str, Any]], columns: List[str], title: str = "Export",
             subtitle: str = "") -> bytes:
    if not _PDF_OK:
        raise HTTPException(500, "reportlab not installed")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title,
                              leftMargin=10 * mm, rightMargin=10 * mm,
                              topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16,
                                    textColor=colors.HexColor("#111827"))
    subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9,
                                     textColor=colors.HexColor("#6B7280"))
    elements = [Paragraph(html.escape(title), title_style)]
    if subtitle:
        elements.append(Paragraph(html.escape(subtitle), subtitle_style))
    elements.append(Spacer(1, 6))
    data: List[List[str]] = [columns[:]]
    for row in rows[:2000]:  # cap PDF rows
        data.append([_fmt_value(row.get(c)) for c in columns])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tbl)
    footer = ParagraphStyle("Foot", parent=styles["Normal"], fontSize=8,
                             textColor=colors.HexColor("#9CA3AF"), alignment=1)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"GO OIL DMS · Generated {datetime.now(timezone.utc).isoformat()}", footer))
    doc.build(elements)
    return buf.getvalue()


def to_print_html(rows: List[Dict[str, Any]], columns: List[str], title: str = "Export",
                    subtitle: str = "") -> bytes:
    parts: List[str] = ['<!doctype html><html><head><meta charset="utf-8">']
    parts.append(f"<title>{html.escape(title)}</title>")
    parts.append("""
<style>
  body { font-family: -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
         margin: 24px; color:#111827; }
  h1 { font-size: 20px; margin: 0 0 4px 0; }
  .sub { font-size: 12px; color: #6B7280; margin-bottom: 16px; }
  table { width: 100%; border-collapse: collapse; font-size: 12px; }
  thead th { background: #1F2937; color: white; text-align: left; padding: 6px 8px; }
  tbody td { padding: 6px 8px; border-bottom: 1px solid #E5E7EB; }
  tbody tr:nth-child(even) td { background: #F9FAFB; }
  .footer { margin-top: 20px; font-size: 10px; color: #9CA3AF; text-align: center; }
  @media print {
    body { margin: 12mm; }
    .no-print { display: none; }
    thead { display: table-header-group; }
    tr { page-break-inside: avoid; }
  }
  .print-btn { position: fixed; top: 16px; right: 16px; padding: 8px 16px;
               background: #1F2937; color: white; border: none; border-radius: 6px;
               cursor: pointer; }
</style>
</head><body>""")
    parts.append('<button class="print-btn no-print" onclick="window.print()">Print</button>')
    parts.append(f'<h1>{html.escape(title)}</h1>')
    if subtitle:
        parts.append(f'<div class="sub">{html.escape(subtitle)}</div>')
    parts.append('<table><thead><tr>')
    for c in columns:
        parts.append(f"<th>{html.escape(str(c))}</th>")
    parts.append("</tr></thead><tbody>")
    for row in rows:
        parts.append("<tr>")
        for c in columns:
            parts.append(f"<td>{html.escape(_fmt_value(row.get(c)))}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    parts.append(f'<div class="footer">GO OIL DMS · Generated {datetime.now(timezone.utc).isoformat()}</div>')
    parts.append("</body></html>")
    return "".join(parts).encode("utf-8")


def render_export(rows: List[Dict[str, Any]], fmt: str, columns: Optional[List[str]] = None,
                    title: str = "Export", subtitle: str = "") -> Response:
    cols = columns or _infer_columns(rows)
    fmt = (fmt or "csv").lower()
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in title.lower())[:60] or "export"
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    if fmt == "csv":
        data = to_csv(rows, cols)
        return Response(content=data, media_type="text/csv",
                          headers={"Content-Disposition": f'attachment; filename="{safe}-{stamp}.csv"'})
    if fmt == "xlsx":
        data = to_xlsx(rows, cols, title=title)
        return Response(content=data,
                          media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          headers={"Content-Disposition": f'attachment; filename="{safe}-{stamp}.xlsx"'})
    if fmt == "pdf":
        data = to_pdf(rows, cols, title=title, subtitle=subtitle)
        return Response(content=data, media_type="application/pdf",
                          headers={"Content-Disposition": f'inline; filename="{safe}-{stamp}.pdf"'})
    if fmt in ("html", "print"):
        data = to_print_html(rows, cols, title=title, subtitle=subtitle)
        return Response(content=data, media_type="text/html")
    raise HTTPException(400, f"Unsupported export format: {fmt}")


# ---------- Router ----------

# Map of exportable collections → (mongo collection name, human title)
EXPORT_COLLECTIONS = {
    "products": ("products", "Products"),
    "skus": ("skus", "SKUs"),
    "batches": ("batches", "Batches"),
    "distributors": ("distributors", "Distributors"),
    "retailers": ("retailers", "Retailers"),
    "customers": ("customers", "Customers"),
    "primary-orders": ("primary_orders", "Primary Orders"),
    "secondary-orders": ("secondary_orders", "Secondary Orders"),
    "customer-orders": ("customer_orders", "Customer Orders"),
    "invoices": ("invoices", "Invoices"),
    "dispatches": ("dispatches", "Dispatches"),
    "grns": ("grns", "Goods Received Notes"),
    "payments": ("payments", "Payments"),
    "outstanding": ("outstanding", "Outstanding"),
    "ledger": ("double_ledger", "Double-Entry Ledger"),
    "returns": ("returns", "Returns"),
    "damage": ("damage", "Damage Records"),
    "claims": ("claims", "Claims"),
    "credit-notes": ("credit_notes", "Credit Notes"),
    "debit-notes": ("debit_notes", "Debit Notes"),
    "replacements": ("replacements", "Replacements"),
    "expiry-records": ("expiry_records", "Expiry Records"),
    "exceptions": ("exceptions", "Exceptions"),
    "approval-requests": ("approval_requests", "Approval Requests"),
    "audit-log": ("audit_log", "Audit Log"),
    "expenses": ("expenses", "Expenses"),
    "notifications": ("notifications", "Notifications"),
    "coupons": ("coupons", "Coupons"),
    "cashback-rules": ("cashback_rules", "Cashback Rules"),
    "cashback-transactions": ("cashback_transactions", "Cashback Transactions"),
    "wallets": ("wallets", "Wallets"),
    "stock-ledger": ("stock_ledger", "Stock Ledger"),
    "company-inventory": ("company_inventory", "Company Inventory"),
    "distributor-inventory": ("distributor_inventory", "Distributor Inventory"),
    "retailer-inventory": ("retailer_inventory", "Retailer Inventory"),
}


def build_exports_router(db, get_current_user):
    router = APIRouter(prefix="/exports", tags=["exports"])

    @router.get("/collections")
    async def list_exportable(user: dict = Depends(get_current_user)):
        """List every collection exportable via GET /exports/{resource}."""
        return {"data": [{"key": k, "title": v[1]} for k, v in EXPORT_COLLECTIONS.items()]}

    @router.get("/{resource}")
    async def export_resource(
        resource: str,
        format: str = Query("csv", regex="^(csv|xlsx|pdf|print|html)$"),
        limit: int = Query(5000, ge=1, le=50000),
        user: dict = Depends(get_current_user),
    ):
        if resource not in EXPORT_COLLECTIONS:
            raise HTTPException(404, "Unknown resource")
        coll_name, title = EXPORT_COLLECTIONS[resource]
        rows = await db[coll_name].find({}, {"_id": 0}).limit(limit).to_list(limit)
        subtitle = f"{len(rows)} rows · exported by {user.get('email', '?')}"
        return render_export(rows, format, title=title, subtitle=subtitle)

    @router.post("/render")
    async def render_generic(
        body: Dict[str, Any] = Body(...),
        user: dict = Depends(get_current_user),
    ):
        """Render arbitrary tabular data (already-computed) into the requested format.

        Body: { rows: [...], columns?: [...], format: "csv|xlsx|pdf|print", title?: str, subtitle?: str }
        """
        rows = body.get("rows") or []
        if not isinstance(rows, list):
            raise HTTPException(400, "rows must be an array")
        fmt = body.get("format", "csv")
        columns = body.get("columns")
        title = body.get("title") or "Report"
        subtitle = body.get("subtitle") or ""
        return render_export(rows, fmt, columns=columns, title=title, subtitle=subtitle)

    return router
