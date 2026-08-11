"""
Simple DMS — Distribution Management System.
Fresh, minimal router. Only two workflows:
  1) Primary Sales   (Owner  ↔ Distributor)
  2) Secondary Sales (Distributor ↔ Retailer)

All endpoints prefixed /api/dms/*
All collections prefixed dms_*
Tenant scope: `tnt-dms-oil` (dedicated tenant so existing tenancy wrapper isolates it).
"""
from __future__ import annotations

import uuid
import io
import base64
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Body, Query, UploadFile, File, Request, Form
from pydantic import BaseModel


DMS_TENANT_ID = "tnt-dms-oil"

# ────────────────────────────── helpers ──────────────────────────────
def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _nid(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:10]}"


def _clean(d: Dict[str, Any]) -> Dict[str, Any]:
    if d and "_id" in d:
        d.pop("_id", None)
    return d


def _round(v: Any, n: int = 2) -> float:
    try:
        return round(float(v), n)
    except Exception:
        return 0.0


# ──────────────────────── Invoice helpers (Vyapar-style) ────────────────────────
def _num_to_words_inr(amount: Any) -> str:
    """Indian numbering-system amount in words. Returns e.g.
    'Rupees One Lakh Twenty Three Thousand Four Hundred Fifty & Fifty Paise Only'."""
    try:
        amount = float(amount or 0)
    except Exception:
        amount = 0.0
    rupees = int(amount)
    paise = int(round((amount - rupees) * 100))
    if paise == 100:
        rupees += 1
        paise = 0

    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two(n: int) -> str:
        if n < 20:
            return ones[n]
        return (tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")).strip()

    def three(n: int) -> str:
        # 0..999
        h = n // 100
        r = n % 100
        out = ""
        if h:
            out += ones[h] + " Hundred"
            if r:
                out += " "
        if r:
            out += two(r)
        return out

    if rupees == 0:
        words = "Zero"
    else:
        crore = rupees // 10000000
        rem = rupees % 10000000
        lakh = rem // 100000
        rem = rem % 100000
        thousand = rem // 1000
        rem = rem % 1000
        hundred = rem
        parts: List[str] = []
        if crore:
            parts.append(three(crore) + " Crore")
        if lakh:
            parts.append(three(lakh) + " Lakh")
        if thousand:
            parts.append(three(thousand) + " Thousand")
        if hundred:
            parts.append(three(hundred))
        words = " ".join([p for p in parts if p]).strip()

    result = f"Rupees {words}"
    if paise:
        result += f" & {two(paise)} Paise"
    result += " Only"
    return result


def _make_upi_qr_dataurl(upi_id: str, name: str = "", amount: Any = None) -> str:
    """Generate a UPI payment QR as a base64 PNG data URL. Empty string on failure/no upi."""
    upi_id = (upi_id or "").strip()
    if not upi_id:
        return ""
    try:
        import qrcode  # local import to avoid hard dependency at import time
        params = [f"pa={upi_id}"]
        if name:
            params.append(f"pn={name}")
        params.append("cu=INR")
        try:
            if amount and float(amount) > 0:
                params.append(f"am={float(amount):.2f}")
        except Exception:
            pass
        payload = "upi://pay?" + "&".join(params)
        img = qrcode.make(payload)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/png;base64,{b64}"
    except Exception:
        return ""


def _gst_breakup(gst_total: float, seller_state_code: str = "", buyer_state_code: str = "") -> Dict[str, Any]:
    """Split total GST into SGST+CGST (intra-state) or IGST (inter-state)."""
    gst_total = _round(gst_total)
    ss = (seller_state_code or "").strip()
    bs = (buyer_state_code or "").strip()
    interstate = bool(ss and bs and ss != bs)
    if interstate:
        return {"is_interstate": True, "igst": gst_total, "sgst": 0.0, "cgst": 0.0}
    half = _round(gst_total / 2.0)
    return {"is_interstate": False, "igst": 0.0, "sgst": half, "cgst": _round(gst_total - half)}


# ────────────────────────────── router builder ──────────────────────────────
def build_dms_router(db, get_current_user):
    router = APIRouter(prefix="/dms", tags=["dms"])

    # =========================================================================
    # role guards
    # =========================================================================
    def _guard(*allowed):
        async def _dep(user: dict = Depends(get_current_user)) -> dict:
            role = user.get("role")
            if role in allowed or role == "super_admin":
                return user
            raise HTTPException(status_code=403, detail=f"Requires role in {allowed}")
        return _dep

    owner_only = _guard("owner")  # super_admin implicit
    owner_or_accountant = _guard("owner", "owner_accountant")
    distributor_only = _guard("distributor")
    retailer_only = _guard("retailer")
    salesperson_only = _guard("salesperson")
    team_leader_only = _guard("team_leader")
    regional_manager_only = _guard("regional_manager")
    dist_or_dacct = _guard("distributor", "distributor_accountant")

    # Field-staff guard — ONLY field roles (salesperson / team leader / regional
    # manager) may punch in/out and be GPS-tracked. Owner is office-only, and
    # distributors / retailers do not punch at all.
    def _field_user_guard():
        FIELD_ROLES = ("salesperson", "team_leader", "regional_manager")
        async def _dep(user: dict = Depends(get_current_user)) -> dict:
            if user.get("role") not in FIELD_ROLES:
                raise HTTPException(status_code=403, detail="Punch in/out is for field staff only")
            return user
        return _dep
    field_user_only = _field_user_guard()

    # =========================================================================
    # Notifications (simple in-app)
    # =========================================================================
    async def notify(recipient_id: str, kind: str, title: str, body: str, link: Optional[str] = None):
        await db.dms_notifications.insert_one({
            "id": _nid("ntf"),
            "recipient_id": recipient_id,
            "kind": kind,
            "title": title,
            "body": body,
            "link": link,
            "read": False,
            "created_at": _now(),
        })

    @router.get("/notifications")
    async def get_notifications(user: dict = Depends(get_current_user)):
        docs = await db.dms_notifications.find(
            {"recipient_id": user["id"]}, {"_id": 0}
        ).sort("created_at", -1).to_list(50)
        unread = await db.dms_notifications.count_documents({"recipient_id": user["id"], "read": False})
        return {"data": docs, "unread": unread}

    @router.post("/notifications/{nid}/read")
    async def mark_read(nid: str, user: dict = Depends(get_current_user)):
        await db.dms_notifications.update_one(
            {"id": nid, "recipient_id": user["id"]}, {"$set": {"read": True}}
        )
        return {"ok": True}

    @router.post("/notifications/read-all")
    async def mark_all_read(user: dict = Depends(get_current_user)):
        await db.dms_notifications.update_many(
            {"recipient_id": user["id"], "read": False}, {"$set": {"read": True}}
        )
        return {"ok": True}

    # =========================================================================
    # Short sequential document numbers (INV-0001, DC-0001, ...)
    # =========================================================================
    async def _next_no(counter_key: str, prefix: str, width: int = 4) -> str:
        await db.dms_counters.update_one({"id": counter_key}, {"$inc": {"seq": 1}}, upsert=True)
        doc = await db.dms_counters.find_one({"id": counter_key}, {"_id": 0, "seq": 1})
        n = int((doc or {}).get("seq", 1))
        return f"{prefix}-{n:0{width}d}"


    # =========================================================================
    # CATEGORIES (product types)
    # =========================================================================
    @router.get("/categories")
    async def list_categories(user: dict = Depends(get_current_user)):
        docs = await db.dms_categories.find({}, {"_id": 0}).sort("name", 1).to_list(500)
        return {"data": docs, "count": len(docs)}

    @router.post("/categories")
    async def create_category(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        if not body.get("name"):
            raise HTTPException(status_code=400, detail="name required")
        doc = {
            "id": _nid("cat"),
            "name": body["name"],
            "description": body.get("description", ""),
            "created_at": _now(),
        }
        await db.dms_categories.insert_one(doc)
        return _clean(doc)

    @router.put("/categories/{cid}")
    async def update_category(cid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        upd = {k: v for k, v in body.items() if k in {"name", "description"}}
        r = await db.dms_categories.update_one({"id": cid}, {"$set": upd})
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Category not found")
        return {"ok": True}

    @router.delete("/categories/{cid}")
    async def delete_category(cid: str, user: dict = Depends(owner_only)):
        # block if products exist
        cnt = await db.dms_products.count_documents({"category_id": cid})
        if cnt > 0:
            raise HTTPException(status_code=400, detail=f"Category has {cnt} products; delete them first")
        await db.dms_categories.delete_one({"id": cid})
        return {"ok": True}

    # =========================================================================
    # PRODUCTS + PRICE BATCHES
    # =========================================================================
    async def _resolve_visible_products_for_distributor(distributor_id: str) -> List[Dict[str, Any]]:
        """Return products that owner has NOT hidden for this distributor."""
        hidden = set()
        async for v in db.dms_dist_visibility.find({"distributor_id": distributor_id, "visible": False}, {"_id": 0}):
            hidden.add(v["product_id"])
        products = await db.dms_products.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(1000)
        return [p for p in products if p["id"] not in hidden]

    @router.get("/products")
    async def list_products(user: dict = Depends(get_current_user)):
        """Owner sees ALL; distributor sees visible ones only."""
        role = user.get("role")
        if role == "distributor":
            dist_id = user.get("distributor_id")
            if not dist_id:
                return {"data": [], "count": 0}
            docs = await _resolve_visible_products_for_distributor(dist_id)
        else:
            docs = await db.dms_products.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
        # Enrich with category name
        cats = {c["id"]: c["name"] async for c in db.dms_categories.find({}, {"_id": 0, "id": 1, "name": 1})}
        for p in docs:
            p["category_name"] = cats.get(p.get("category_id"), "")
        return {"data": docs, "count": len(docs)}

    @router.post("/products")
    async def create_product(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        required = ["name", "category_id", "sku_code", "box_qty", "unit_price"]
        for k in required:
            if body.get(k) in (None, ""):
                raise HTTPException(status_code=400, detail=f"{k} required")
        # sku unique
        existing = await db.dms_products.find_one({"sku_code": body["sku_code"]})
        if existing:
            raise HTTPException(status_code=400, detail="SKU code already exists")
        pid = _nid("prd")
        unit_price = _round(body["unit_price"])
        box_qty = int(body["box_qty"])
        doc = {
            "id": pid,
            "name": body["name"],
            "category_id": body["category_id"],
            "sku_code": body["sku_code"],
            "description": body.get("description", ""),
            "box_qty": box_qty,       # bottles per box
            "unit_price": unit_price, # per box
            "previous_price": None,
            "hsn": body.get("hsn", ""),
            "gst_pct": _round(body.get("gst_pct", 18)),
            "coupons_per_box": int(body.get("coupons_per_box", 100)),
            "points_value": _round(body.get("points_value", 10)),
            "active": True,
            "created_at": _now(),
        }
        await db.dms_products.insert_one(doc)
        # initial price batch
        await db.dms_price_batches.insert_one({
            "id": _nid("pb"),
            "product_id": pid,
            "price": unit_price,
            "from_date": _now(),
            "to_date": None,
            "created_at": _now(),
        })
        return _clean(doc)

    @router.put("/products/{pid}")
    async def update_product(pid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        current = await db.dms_products.find_one({"id": pid}, {"_id": 0})
        if not current:
            raise HTTPException(status_code=404, detail="Product not found")
        upd: Dict[str, Any] = {}
        for k in ["name", "description", "box_qty", "hsn", "gst_pct", "active", "coupons_per_box", "points_value"]:
            if k in body:
                upd[k] = body[k]
        if "unit_price" in body:
            new_price = _round(body["unit_price"])
            if new_price != current.get("unit_price"):
                # close current batch, open new batch
                await db.dms_price_batches.update_one(
                    {"product_id": pid, "to_date": None},
                    {"$set": {"to_date": _now()}}
                )
                await db.dms_price_batches.insert_one({
                    "id": _nid("pb"),
                    "product_id": pid,
                    "price": new_price,
                    "from_date": _now(),
                    "to_date": None,
                    "created_at": _now(),
                })
                upd["previous_price"] = current.get("unit_price")
                upd["unit_price"] = new_price
        upd["updated_at"] = _now()
        await db.dms_products.update_one({"id": pid}, {"$set": upd})
        doc = await db.dms_products.find_one({"id": pid}, {"_id": 0})
        return _clean(doc or {})

    @router.get("/products/{pid}/price-history")
    async def price_history(pid: str, user: dict = Depends(get_current_user)):
        docs = await db.dms_price_batches.find({"product_id": pid}, {"_id": 0}).sort("from_date", -1).to_list(200)
        return {"data": docs}

    # =========================================================================
    # OWNER INVENTORY
    # =========================================================================
    async def _get_owner_stock(product_id: str) -> int:
        doc = await db.dms_owner_inventory.find_one({"product_id": product_id}, {"_id": 0})
        return int(doc.get("qty_boxes", 0)) if doc else 0

    async def _adjust_owner_stock(product_id: str, delta_boxes: int, reason: str, ref: str = ""):
        cur = await db.dms_owner_inventory.find_one({"product_id": product_id})
        if cur:
            new_qty = int(cur.get("qty_boxes", 0)) + delta_boxes
            await db.dms_owner_inventory.update_one(
                {"product_id": product_id},
                {"$set": {"qty_boxes": max(new_qty, 0), "updated_at": _now()}},
            )
        else:
            await db.dms_owner_inventory.insert_one({
                "id": _nid("oinv"),
                "product_id": product_id,
                "qty_boxes": max(delta_boxes, 0),
                "updated_at": _now(),
            })
        # ledger row
        await db.dms_stock_ledger.insert_one({
            "id": _nid("sl"),
            "scope": "owner",
            "product_id": product_id,
            "delta_boxes": delta_boxes,
            "reason": reason,
            "reference": ref,
            "at": _now(),
        })

    @router.get("/owner/inventory")
    async def owner_inventory(user: dict = Depends(owner_or_accountant)):
        rows = await db.dms_owner_inventory.find({}, {"_id": 0}).to_list(1000)
        # enrich with product info
        pids = [r["product_id"] for r in rows]
        prods = {p["id"]: p async for p in db.dms_products.find({"id": {"$in": pids}}, {"_id": 0})}
        for r in rows:
            p = prods.get(r["product_id"], {})
            r["product_name"] = p.get("name", "")
            r["sku_code"] = p.get("sku_code", "")
            r["box_qty"] = p.get("box_qty", 0)
            r["unit_price"] = p.get("unit_price", 0)
            r["value"] = _round((r.get("qty_boxes", 0) or 0) * (p.get("unit_price", 0) or 0))
        rows.sort(key=lambda x: x.get("product_name", ""))
        return {"data": rows, "total_value": _round(sum(r.get("value", 0) for r in rows))}

    @router.post("/owner/inventory/adjust")
    async def owner_inv_adjust(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        pid = body.get("product_id")
        delta = int(body.get("delta_boxes", 0))
        if not pid or delta == 0:
            raise HTTPException(status_code=400, detail="product_id + non-zero delta_boxes required")
        await _adjust_owner_stock(pid, delta, body.get("reason", "manual_adjust"), body.get("reference", ""))
        return {"ok": True, "new_qty": await _get_owner_stock(pid)}

    # =========================================================================
    # DISTRIBUTORS + KYC + user
    # =========================================================================
    async def _create_dms_user(email: str, password: str, name: str, role: str, extra: Dict[str, Any] = None):
        import bcrypt
        existing = await db.users.find_one({"email": email.lower()})
        if existing:
            raise HTTPException(status_code=400, detail=f"User email {email} already exists")
        uid = _nid("usr")
        doc = {
            "id": uid,
            "email": email.lower(),
            "name": name,
            "role": role,
            "password_hash": bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(),
            "active": True,
            "created_at": _now(),
        }
        if extra:
            doc.update(extra)
        await db.users.insert_one(doc)
        return uid

    @router.get("/distributors")
    async def list_distributors(user: dict = Depends(get_current_user)):
        role = user.get("role")
        query: Dict[str, Any] = {}
        # salesperson: only assigned; team_leader: only their assigned; distributor: only themselves
        if role == "distributor":
            query = {"id": user.get("distributor_id")}
        elif role == "salesperson":
            # only assigned distributors
            assigns = await db.dms_sp_assignments.find({"salesperson_id": user["id"]}, {"_id": 0}).to_list(500)
            ids = [a["distributor_id"] for a in assigns]
            query = {"id": {"$in": ids}} if ids else {"id": "__none__"}
        elif role == "team_leader":
            assigns = await db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0}).to_list(500)
            ids = [a["distributor_id"] for a in assigns]
            query = {"id": {"$in": ids}} if ids else {"id": "__none__"}
        docs = await db.dms_distributors.find(query, {"_id": 0}).sort("name", 1).to_list(500)
        return {"data": docs, "count": len(docs)}

    @router.post("/distributors")
    async def create_distributor(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        # FULL-PROCESS ONBOARDING: a distributor login is created ONLY after all
        # required details + KYC + at least one document are provided.
        required_labels = {
            "name": "Business Name", "email": "Login Email", "password": "Login Password",
            "phone": "Phone", "address": "Address", "region": "Region",
            "gstin": "GSTIN", "pan": "PAN", "shop_license": "Shop / Trade License",
            "bank_name": "Bank Name", "bank_account": "Bank Account", "bank_ifsc": "Bank IFSC",
        }
        missing = [label for key, label in required_labels.items()
                   if not str(body.get(key) or "").strip()]
        if not body.get("documents"):
            missing.append("At least one uploaded Document")
        if missing:
            raise HTTPException(
                status_code=400,
                detail="Complete the full onboarding before creating the login. Missing: " + ", ".join(missing),
            )
        did = _nid("dist")
        # create user for distributor
        uid = await _create_dms_user(
            email=body["email"],
            password=body["password"],
            name=body["name"],
            role="distributor",
            extra={"distributor_id": did, "phone": body["phone"]},
        )
        # optional accountant
        accountant_uid = None
        if body.get("accountant_email"):
            accountant_uid = await _create_dms_user(
                email=body["accountant_email"],
                password=body.get("accountant_password") or "Demo@2026",
                name=body.get("accountant_name") or f"{body['name']} Accountant",
                role="distributor_accountant",
                extra={"distributor_id": did},
            )
        dist = {
            "id": did,
            "name": body["name"],
            "email": body["email"].lower(),
            "phone": body["phone"],
            "address": body["address"],
            "region": body.get("region", ""),
            "user_id": uid,
            "accountant_user_id": accountant_uid,
            # geo — for Owner live map
            "location_link": body.get("location_link", ""),
            "gps_lat": body.get("gps_lat"),
            "gps_lng": body.get("gps_lng"),
            # KYC
            "kyc": {
                "gstin": body.get("gstin", ""),
                "pan": body.get("pan", ""),
                "aadhaar": body.get("aadhaar", ""),
                "shop_license": body.get("shop_license", ""),
                "bank_name": body.get("bank_name", ""),
                "bank_account": body.get("bank_account", ""),
                "bank_ifsc": body.get("bank_ifsc", ""),
                "notes": body.get("kyc_notes", ""),
            },
            # Bank + UPI (used on retailer/direct-sale invoice "Pay To")
            "bank": body.get("bank") or {
                "gstin": body.get("gstin", ""),
                "bank_name": body.get("bank_name", ""),
                "bank_account": body.get("bank_account", ""),
                "bank_ifsc": body.get("bank_ifsc", ""),
                "bank_branch": body.get("bank_branch", ""),
                "upi_id": body.get("upi_id", ""),
                "upi_name": body.get("upi_name", ""),
                "qr_url": body.get("qr_url", ""),
            },
            "state": body.get("state", ""),
            "state_code": body.get("state_code", ""),
            "credit_limit": _round(body.get("credit_limit", 0)),
            "documents": body.get("documents", []),
            "active": True,
            "created_at": _now(),
        }
        await db.dms_distributors.insert_one(dist)
        return _clean(dist)

    @router.get("/distributors/{did}")
    async def get_distributor(did: str, user: dict = Depends(get_current_user)):
        doc = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Distributor not found")
        return doc

    @router.put("/distributors/{did}")
    async def update_distributor(did: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        upd: Dict[str, Any] = {}
        for k in ["name", "phone", "address", "region", "credit_limit", "active",
                  "location_link", "gps_lat", "gps_lng", "documents",
                  "bank", "state", "state_code"]:
            if k in body:
                upd[k] = body[k]
        if "kyc" in body:
            upd["kyc"] = body["kyc"]
        upd["updated_at"] = _now()
        r = await db.dms_distributors.update_one({"id": did}, {"$set": upd})
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Distributor not found")
        return {"ok": True}

    @router.delete("/distributors/{did}")
    async def delete_distributor(did: str, user: dict = Depends(owner_only)):
        dist = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
        if not dist:
            raise HTTPException(status_code=404, detail="Distributor not found")
        # block delete if there are primary orders referencing this distributor
        order_count = await db.dms_primary_orders.count_documents({"distributor_id": did})
        if order_count:
            raise HTTPException(status_code=400,
                detail=f"Cannot delete — {order_count} primary order(s) exist for this distributor. Deactivate instead.")
        await db.dms_distributors.delete_one({"id": did})
        await db.dms_dist_visibility.delete_many({"distributor_id": did})
        # remove linked login user(s)
        uid = dist.get("user_id")
        if uid:
            await db.users.delete_one({"id": uid})
        await db.users.delete_many({"distributor_id": did, "role": "distributor"})
        return {"ok": True, "deleted": did}

    # ── product visibility per distributor (owner control) ──
    @router.get("/distributors/{did}/visibility")
    async def get_dist_visibility(did: str, user: dict = Depends(owner_only)):
        products = await db.dms_products.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
        vis_map = {v["product_id"]: v.get("visible", True)
                   async for v in db.dms_dist_visibility.find({"distributor_id": did}, {"_id": 0})}
        out = []
        for p in products:
            out.append({
                "product_id": p["id"],
                "product_name": p["name"],
                "sku_code": p["sku_code"],
                "visible": vis_map.get(p["id"], True),
            })
        return {"data": out}

    @router.put("/distributors/{did}/visibility")
    async def set_dist_visibility(did: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        pid = body.get("product_id")
        visible = bool(body.get("visible", True))
        if not pid:
            raise HTTPException(status_code=400, detail="product_id required")
        await db.dms_dist_visibility.update_one(
            {"distributor_id": did, "product_id": pid},
            {"$set": {"distributor_id": did, "product_id": pid, "visible": visible, "updated_at": _now()}},
            upsert=True,
        )
        return {"ok": True}

    # =========================================================================
    # DISTRIBUTOR — browse & place primary order
    # =========================================================================
    @router.get("/distributor/browse")
    async def distributor_browse(user: dict = Depends(distributor_only)):
        did = user.get("distributor_id")
        if not did:
            raise HTTPException(status_code=400, detail="Not linked to a distributor")
        products = await _resolve_visible_products_for_distributor(did)
        cats = {c["id"]: c["name"] async for c in db.dms_categories.find({}, {"_id": 0, "id": 1, "name": 1})}
        # attach available owner stock, category name
        for p in products:
            p["category_name"] = cats.get(p.get("category_id"), "")
            p["owner_stock_boxes"] = await _get_owner_stock(p["id"])
        return {"data": products}

    # =========================================================================
    # PRIMARY ORDERS
    #   Lifecycle: pending → partially_fulfilled → ready_to_go → received
    # =========================================================================
    @router.post("/primary-orders")
    async def place_primary_order(body: Dict[str, Any] = Body(...), user: dict = Depends(distributor_only)):
        did = user.get("distributor_id")
        if not did:
            raise HTTPException(status_code=400, detail="Not linked to a distributor")
        items = body.get("items") or []
        if not items:
            raise HTTPException(status_code=400, detail="items[] required")
        dist = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
        if not dist:
            raise HTTPException(status_code=404, detail="Distributor not found")

        # normalise items — always price at CURRENT price
        order_items = []
        subtotal = 0.0
        gst_total = 0.0
        # global GST% from settings (default 0 until owner configures)
        settings_doc = await db.dms_settings.find_one({"id": "global"}, {"_id": 0}) or {}
        global_gst = float(settings_doc.get("gst_pct") or 0)
        for it in items:
            pid = it.get("product_id")
            qty_boxes = int(it.get("qty_boxes", 0))
            if not pid or qty_boxes <= 0:
                continue
            p = await db.dms_products.find_one({"id": pid}, {"_id": 0})
            if not p:
                raise HTTPException(status_code=400, detail=f"Product {pid} not found")
            unit_price = _round(p["unit_price"])
            line_sub = _round(unit_price * qty_boxes)
            line_gst = _round(line_sub * (global_gst / 100.0))
            subtotal += line_sub
            gst_total += line_gst
            order_items.append({
                "product_id": pid,
                "product_name": p["name"],
                "sku_code": p["sku_code"],
                "box_qty": p["box_qty"],
                "unit_price": unit_price,           # price applied
                "previous_price": p.get("previous_price"),
                "gst_pct": global_gst,
                "qty_boxes_ordered": qty_boxes,
                "qty_boxes_fulfilled": 0,
                "line_subtotal": line_sub,
                "line_gst": line_gst,
                "line_total": _round(line_sub + line_gst),
            })
        if not order_items:
            raise HTTPException(status_code=400, detail="No valid items")

        total = _round(subtotal + gst_total)
        order = {
            "id": _nid("po"),
            "order_no": f"PO-{datetime.now().strftime('%y%m%d%H%M%S')}",
            "distributor_id": did,
            "distributor_name": dist["name"],
            "items": order_items,
            "subtotal": _round(subtotal),
            "gst_total": _round(gst_total),
            "total": total,
            "status": "pending",
            "fulfillment_pct": 0,
            "notes": body.get("notes", ""),
            "created_at": _now(),
            "created_by": user["id"],
            "ready_at": None,
            "received_at": None,
            "ebill_id": None,
        }
        await db.dms_primary_orders.insert_one(order)

        # notify all owners + owner_accountants
        async for u in db.users.find({"role": {"$in": ["owner", "owner_accountant"]}}, {"_id": 0, "id": 1}):
            await notify(
                u["id"], "primary_order",
                f"New order from {dist['name']}",
                f"{order['order_no']} — {len(order_items)} items — ₹{total:,.0f}",
                f"/dms/owner/primary-orders/{order['id']}",
            )
        return _clean(order)

    @router.get("/primary-orders")
    async def list_primary_orders(status: Optional[str] = None, user: dict = Depends(get_current_user)):
        role = user.get("role")
        q: Dict[str, Any] = {}
        if status:
            q["status"] = status
        if role == "distributor":
            q["distributor_id"] = user.get("distributor_id")
        elif role in ("distributor_accountant",):
            q["distributor_id"] = user.get("distributor_id")
        docs = await db.dms_primary_orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
        return {"data": docs, "count": len(docs)}

    @router.get("/primary-orders/{oid}")
    async def get_primary_order(oid: str, user: dict = Depends(get_current_user)):
        doc = await db.dms_primary_orders.find_one({"id": oid}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Order not found")
        role = user.get("role")
        if role in ("distributor", "distributor_accountant") and doc.get("distributor_id") != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        # attach ebill + attachments
        if doc.get("ebill_id"):
            eb = await db.dms_ebills.find_one({"id": doc["ebill_id"]}, {"_id": 0})
            doc["ebill"] = eb
        atts = await db.dms_attachments.find({"reference_id": oid}, {"_id": 0}).to_list(50)
        doc["attachments"] = atts
        return doc

    @router.post("/primary-orders/{oid}/fulfill-line")
    async def fulfill_primary_line(oid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        """Set qty_boxes_fulfilled for a specific product line."""
        pid = body.get("product_id")
        qty = int(body.get("qty_boxes_fulfilled", 0))
        if not pid:
            raise HTTPException(status_code=400, detail="product_id required")
        order = await db.dms_primary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if order["status"] in ("ready_to_go", "received"):
            raise HTTPException(status_code=400, detail="Order already sent")
        # update line
        line = next((it for it in order["items"] if it["product_id"] == pid), None)
        if not line:
            raise HTTPException(status_code=400, detail="Line not in order")
        qty = max(0, min(qty, line["qty_boxes_ordered"]))
        # Stop-sale-on-negative check: current owner stock must cover this fulfillment
        # (compare qty against owner_stock; other fulfilled lines' stock is not yet decremented,
        #  so we only need this line's requested amount ≤ current owner stock).
        if qty > 0 and await _stop_sale_enabled():
            avail = await _get_owner_stock(pid)
            if qty > avail:
                raise HTTPException(status_code=400, detail=f"Insufficient owner stock: available {avail} boxes, requested {qty}. Enable stock or reduce fulfilled quantity.")
        line["qty_boxes_fulfilled"] = qty
        # recompute fulfillment_pct
        ord_total = sum(it["qty_boxes_ordered"] for it in order["items"])
        ful_total = sum(it["qty_boxes_fulfilled"] for it in order["items"])
        pct = int(round((ful_total / ord_total) * 100)) if ord_total > 0 else 0
        new_status = order["status"]
        if pct == 0:
            new_status = "pending"
        elif pct < 100:
            new_status = "partially_fulfilled"
        else:
            new_status = "fulfilled"
        await db.dms_primary_orders.update_one(
            {"id": oid},
            {"$set": {"items": order["items"], "fulfillment_pct": pct, "status": new_status, "updated_at": _now()}},
        )
        return {"ok": True, "fulfillment_pct": pct, "status": new_status}

    @router.post("/primary-orders/{oid}/ready")
    async def mark_ready_to_go(oid: str, user: dict = Depends(owner_only)):
        order = await db.dms_primary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if order["status"] not in ("partially_fulfilled", "fulfilled", "pending"):
            raise HTTPException(status_code=400, detail=f"Cannot ready an order in status {order['status']}")
        # decrement owner inventory for fulfilled quantities
        # (owner stock may already be low — we just record the movement)
        for it in order["items"]:
            if it["qty_boxes_fulfilled"] > 0:
                await _adjust_owner_stock(
                    it["product_id"], -it["qty_boxes_fulfilled"], "primary_dispatch", order["order_no"]
                )
        # generate e-bill
        # total based on fulfilled qty × price
        sub = 0.0
        gst = 0.0
        billed_items = []
        for it in order["items"]:
            if it["qty_boxes_fulfilled"] <= 0:
                continue
            line_sub = _round(it["unit_price"] * it["qty_boxes_fulfilled"])
            line_gst = _round(line_sub * (it["gst_pct"] / 100.0))
            sub += line_sub
            gst += line_gst
            billed_items.append({
                **it,
                "billed_qty_boxes": it["qty_boxes_fulfilled"],
                "line_subtotal": line_sub,
                "line_gst": line_gst,
                "line_total": _round(line_sub + line_gst),
            })
        total = _round(sub + gst)
        ebill = {
            "id": _nid("eb"),
            "ebill_no": f"EB-{datetime.now().strftime('%y%m%d%H%M%S')}",
            "order_id": oid,
            "order_no": order["order_no"],
            "distributor_id": order["distributor_id"],
            "distributor_name": order["distributor_name"],
            "items": billed_items,
            "subtotal": _round(sub),
            "gst_total": _round(gst),
            "total": total,
            "status": "issued",
            "created_at": _now(),
        }
        await db.dms_ebills.insert_one(ebill)
        # primary ledger: debit distributor (they owe)
        await db.dms_primary_ledger.insert_one({
            "id": _nid("le"),
            "distributor_id": order["distributor_id"],
            "kind": "invoice",
            "reference_id": ebill["id"],
            "reference_no": ebill["ebill_no"],
            "amount": total,
            "description": f"Invoice for order {order['order_no']}",
            "at": _now(),
        })
        # update order
        await db.dms_primary_orders.update_one(
            {"id": oid},
            {"$set": {"status": "ready_to_go", "ebill_id": ebill["id"], "ready_at": _now(), "updated_at": _now()}},
        )
        # NOTE: Coupon auto-assignment REMOVED — new GO OIL coupon engine does not
        # tie coupons to primary orders (coupons are inserted randomly in bottles
        # by production; distributor mapping is derived only during scan).
        # notify distributor
        dist_user = await db.users.find_one({"distributor_id": order["distributor_id"], "role": "distributor"}, {"_id": 0, "id": 1})
        if dist_user:
            await notify(
                dist_user["id"], "order_ready",
                f"Order {order['order_no']} ready to go",
                f"e-Bill {ebill['ebill_no']} • Total ₹{total:,.0f}",
                f"/dms/distributor/my-orders/{oid}",
            )
        return {"ok": True, "ebill_id": ebill["id"], "status": "ready_to_go"}

    @router.post("/primary-orders/{oid}/receive")
    async def mark_received(oid: str, user: dict = Depends(distributor_only)):
        order = await db.dms_primary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if order.get("distributor_id") != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Not your order")
        if order["status"] != "ready_to_go":
            raise HTTPException(status_code=400, detail=f"Cannot receive an order in status {order['status']}")
        # move stock into distributor inventory
        for it in order["items"]:
            q = int(it.get("qty_boxes_fulfilled", 0))
            if q <= 0:
                continue
            existing = await db.dms_distributor_inventory.find_one(
                {"distributor_id": order["distributor_id"], "product_id": it["product_id"]}
            )
            if existing:
                await db.dms_distributor_inventory.update_one(
                    {"id": existing["id"]},
                    {"$set": {
                        "qty_boxes": int(existing.get("qty_boxes", 0)) + q,
                        "cost_price": it["unit_price"],  # purchase cost
                        "updated_at": _now(),
                    }},
                )
            else:
                await db.dms_distributor_inventory.insert_one({
                    "id": _nid("dinv"),
                    "distributor_id": order["distributor_id"],
                    "product_id": it["product_id"],
                    "qty_boxes": q,
                    "cost_price": it["unit_price"],
                    "updated_at": _now(),
                })
            await db.dms_stock_ledger.insert_one({
                "id": _nid("sl"),
                "scope": "distributor",
                "distributor_id": order["distributor_id"],
                "product_id": it["product_id"],
                "delta_boxes": q,
                "reason": "primary_receive",
                "reference": order["order_no"],
                "at": _now(),
            })
        await db.dms_primary_orders.update_one(
            {"id": oid}, {"$set": {"status": "received", "received_at": _now(), "updated_at": _now()}}
        )
        # notify owner
        async for u in db.users.find({"role": {"$in": ["owner", "owner_accountant"]}}, {"_id": 0, "id": 1}):
            await notify(
                u["id"], "order_received",
                f"{order['distributor_name']} received {order['order_no']}",
                f"Received on {datetime.now().strftime('%d-%b-%y')}",
                f"/dms/owner/primary-orders/{oid}",
            )
        return {"ok": True, "status": "received"}

    # =========================================================================
    # E-BILL ATTACHMENTS (owner accountant uploads invoice image / doc URL)
    # =========================================================================
    @router.post("/attachments")
    async def add_attachment(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        """Simple attachment: {reference_id, kind, name, url}."""
        if not body.get("reference_id") or not body.get("url"):
            raise HTTPException(status_code=400, detail="reference_id + url required")
        doc = {
            "id": _nid("att"),
            "reference_id": body["reference_id"],
            "kind": body.get("kind", "invoice"),
            "name": body.get("name", "Attachment"),
            "url": body["url"],
            "uploaded_by": user["id"],
            "uploader_role": user["role"],
            "created_at": _now(),
        }
        await db.dms_attachments.insert_one(doc)
        return _clean(doc)

    @router.get("/attachments")
    async def list_attachments(reference_id: str, user: dict = Depends(get_current_user)):
        docs = await db.dms_attachments.find({"reference_id": reference_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
        return {"data": docs}

    # =========================================================================
    # MY BANK / PAYMENT DETAILS — self-service for distributor & retailer
    # =========================================================================
    async def _my_party(user: dict):
        role = user.get("role")
        if role in ("distributor", "distributor_accountant"):
            did = user.get("distributor_id")
            if not did:
                raise HTTPException(status_code=400, detail="No distributor linked to this account")
            doc = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
            return ("distributor", db.dms_distributors, did, doc)
        if role == "retailer":
            rid = user.get("retailer_id")
            if not rid:
                raise HTTPException(status_code=400, detail="No retailer linked to this account")
            doc = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
            return ("retailer", db.dms_retailers, rid, doc)
        raise HTTPException(status_code=403, detail="Only distributor or retailer can manage own bank details")

    @router.get("/my/bank")
    async def get_my_bank(user: dict = Depends(get_current_user)):
        party_type, _coll, pid, doc = await _my_party(user)
        if not doc:
            raise HTTPException(status_code=404, detail="Profile not found")
        return {
            "party_type": party_type, "id": pid, "name": doc.get("name"),
            "bank": doc.get("bank") or {},
        }

    @router.put("/my/bank")
    async def update_my_bank(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        party_type, coll, pid, doc = await _my_party(user)
        if not doc:
            raise HTTPException(status_code=404, detail="Profile not found")
        src = body.get("bank") if isinstance(body.get("bank"), dict) else body
        bank = {
            "gstin": str(src.get("gstin") or "").strip(),
            "bank_name": str(src.get("bank_name") or "").strip(),
            "bank_account": str(src.get("bank_account") or "").strip(),
            "bank_ifsc": str(src.get("bank_ifsc") or "").strip(),
            "bank_branch": str(src.get("bank_branch") or "").strip(),
            "upi_id": str(src.get("upi_id") or "").strip(),
            "upi_name": str(src.get("upi_name") or "").strip(),
            "qr_url": src.get("qr_url") or "",
        }
        await coll.update_one({"id": pid}, {"$set": {"bank": bank, "updated_at": _now()}})
        return {"ok": True, "party_type": party_type, "bank": bank}

    # =========================================================================
    # OWNER — reset demo data to a clean production state
    # =========================================================================
    @router.post("/owner/reset-demo-data")
    async def reset_demo_data(user: dict = Depends(get_current_user)):
        if user.get("role") not in ("owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Only the Owner can reset data")
        # Business collections wiped clean. Login users, global settings and the
        # tenant record are preserved so nobody gets locked out.
        BUSINESS = [
            "dms_distributors", "dms_retailers", "dms_ret_mode", "dms_products", "dms_categories",
            "dms_price_batches", "dms_price_circulars", "dms_owner_inventory", "dms_retailer_prices",
            "dms_primary_orders", "dms_secondary_orders", "dms_secondary_sales", "dms_bills",
            "dms_ebills", "dms_sp_assignments", "dms_tl_assignments", "dms_rm_assignments",
            "dms_terms", "dms_godowns", "dms_godown_inventory", "dms_stock_transfers",
            "dms_punch", "dms_sp_pings", "dms_punch_reopen",
            "dms_primary_ledger", "dms_secondary_ledger",
            "dms_bank_accounts", "dms_bank_transactions", "dms_cash_register",
            "dms_cheques", "dms_loan_accounts", "dms_loan_transactions",
            "dms_expenses", "dms_documents", "dms_attachments",
        ]
        removed = {}
        for c in BUSINESS:
            try:
                r = await db[c].delete_many({})
                if r.deleted_count:
                    removed[c] = r.deleted_count
            except Exception:
                pass
        # Unlink distributor/retailer references + clear any live GPS on users
        await db.users.update_many({}, {"$unset": {
            "distributor_id": "", "retailer_id": "", "last_gps": "", "last_active_at": "",
        }})
        return {"ok": True, "removed": removed, "total": sum(removed.values())}


    # =========================================================================
    # PRIMARY LEDGER (owner ↔ distributor)
    # =========================================================================
    @router.get("/ledger/primary")
    async def primary_ledger(distributor_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        role = user.get("role")
        q: Dict[str, Any] = {}
        # distributor scope
        if role in ("distributor", "distributor_accountant"):
            q["distributor_id"] = user.get("distributor_id")
        elif distributor_id:
            q["distributor_id"] = distributor_id
        entries = await db.dms_primary_ledger.find(q, {"_id": 0}).sort("at", -1).to_list(1000)
        # summary per distributor
        summary: Dict[str, Dict[str, Any]] = {}
        for e in entries:
            did = e["distributor_id"]
            s = summary.setdefault(did, {"distributor_id": did, "billed": 0.0, "paid": 0.0, "outstanding": 0.0})
            if e["kind"] in ("invoice", "debit_note"):
                s["billed"] += e["amount"]
                s["outstanding"] += e["amount"]
            elif e["kind"] in ("payment", "coupon_credit", "credit_note"):
                s["paid"] += e["amount"]
                s["outstanding"] -= e["amount"]
        # enrich with dist names
        dids = list(summary.keys())
        dnames = {d["id"]: d["name"] async for d in db.dms_distributors.find({"id": {"$in": dids}}, {"_id": 0, "id": 1, "name": 1})}
        for s in summary.values():
            s["distributor_name"] = dnames.get(s["distributor_id"], "")
            for k in ("billed", "paid", "outstanding"):
                s[k] = _round(s[k])
        return {"entries": entries, "summary": sorted(summary.values(), key=lambda x: -x["outstanding"])}

    @router.post("/ledger/primary/payment")
    async def record_primary_payment(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_accountant)):
        did = body.get("distributor_id")
        amt = _round(body.get("amount", 0))
        if not did or amt <= 0:
            raise HTTPException(status_code=400, detail="distributor_id + amount>0 required")
        entry = {
            "id": _nid("le"),
            "distributor_id": did,
            "kind": "payment",
            "reference_no": body.get("reference_no", f"PMT-{datetime.now().strftime('%y%m%d%H%M%S')}"),
            "amount": amt,
            "method": body.get("method", "bank_transfer"),
            "description": body.get("description", "Payment received"),
            "at": _now(),
            "recorded_by": user["id"],
        }
        await db.dms_primary_ledger.insert_one(entry)
        return _clean(entry)

    # =========================================================================
    # DASHBOARDS
    # =========================================================================
    @router.get("/dashboard/owner")
    async def owner_dashboard(user: dict = Depends(owner_or_accountant)):
        n_dist = await db.dms_distributors.count_documents({"active": True})
        n_products = await db.dms_products.count_documents({"active": True})
        n_pending = await db.dms_primary_orders.count_documents({"status": {"$in": ["pending", "partially_fulfilled"]}})
        n_ready = await db.dms_primary_orders.count_documents({"status": "ready_to_go"})
        # revenue mtd = sum of e-bill total
        mtd_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        revenue = 0.0
        async for eb in db.dms_ebills.find({"created_at": {"$gte": mtd_start}}, {"_id": 0, "total": 1}):
            revenue += eb.get("total", 0)
        # outstanding = sum(invoices) - sum(payments)
        billed = 0.0
        paid = 0.0
        async for e in db.dms_primary_ledger.find({}, {"_id": 0, "kind": 1, "amount": 1}):
            if e["kind"] == "invoice":
                billed += e["amount"]
            elif e["kind"] in ("payment", "coupon_credit"):
                paid += e["amount"]
        # owner inventory value
        inv_value = 0.0
        async for r in db.dms_owner_inventory.find({}, {"_id": 0}):
            p = await db.dms_products.find_one({"id": r["product_id"]}, {"_id": 0, "unit_price": 1})
            if p:
                inv_value += (r.get("qty_boxes", 0) or 0) * (p.get("unit_price", 0) or 0)
        return {
            "kpis": {
                "distributors": n_dist,
                "products": n_products,
                "pending_orders": n_pending,
                "ready_to_go": n_ready,
                "revenue_mtd": _round(revenue),
                "outstanding_receivable": _round(billed - paid),
                "inventory_value": _round(inv_value),
            }
        }

    @router.get("/dashboard/distributor")
    async def distributor_dashboard(user: dict = Depends(dist_or_dacct)):
        did = user.get("distributor_id")
        if not did:
            # Distributor user with no linked profile yet → clean empty dashboard
            return {
                "kpis": {
                    "stock_boxes": 0, "stock_value": 0, "payable_to_owner": 0,
                    "pending_primary_orders": 0, "ready_to_receive": 0,
                    "receivable_from_retailers": 0, "sales_mtd": 0, "revenue_mtd": 0,
                }
            }
        # current stock (value at cost)
        stock_qty = 0
        stock_value = 0.0
        async for r in db.dms_distributor_inventory.find({"distributor_id": did}, {"_id": 0}):
            q = r.get("qty_boxes", 0) or 0
            cp = r.get("cost_price", 0) or 0
            stock_qty += q
            stock_value += q * cp
        # payable to owner = primary outstanding
        billed = 0.0
        paid = 0.0
        async for e in db.dms_primary_ledger.find({"distributor_id": did}, {"_id": 0}):
            if e["kind"] == "invoice":
                billed += e["amount"]
            elif e["kind"] in ("payment", "coupon_credit"):
                paid += e["amount"]
        payable = _round(billed - paid)
        # pending orders
        pend = await db.dms_primary_orders.count_documents({"distributor_id": did, "status": {"$in": ["pending", "partially_fulfilled"]}})
        ready = await db.dms_primary_orders.count_documents({"distributor_id": did, "status": "ready_to_go"})
        return {
            "kpis": {
                "stock_boxes": stock_qty,
                "stock_value": _round(stock_value),
                "payable_to_owner": payable,
                "pending_primary_orders": pend,
                "ready_to_receive": ready,
                # secondary sales — iteration 2
                "receivable_from_retailers": 0,
                "sales_mtd": 0,
                "revenue_mtd": 0,
            }
        }

    # =========================================================================
    # ME endpoint — enriched with distributor info if applicable
    # =========================================================================
    @router.get("/me")
    async def me(user: dict = Depends(get_current_user)):
        out = dict(user)
        if user.get("distributor_id"):
            d = await db.dms_distributors.find_one({"id": user["distributor_id"]}, {"_id": 0})
            out["distributor"] = d
        if user.get("retailer_id"):
            r = await db.dms_retailers.find_one({"id": user["retailer_id"]}, {"_id": 0})
            out["retailer"] = r
        return out

    # =========================================================================
    # ITERATION 2 — SECONDARY SALES (Distributor ↔ Retailer)
    # =========================================================================

    async def _resolve_visible_products_for_retailer(distributor_id: str, retailer_id: str) -> List[Dict[str, Any]]:
        """Return products distributor has NOT hidden for this retailer AND has stock for."""
        hidden = set()
        async for v in db.dms_ret_visibility.find({"distributor_id": distributor_id, "retailer_id": retailer_id, "visible": False}, {"_id": 0}):
            hidden.add(v["product_id"])
        # only include products distributor has stock in
        prods = []
        async for inv in db.dms_distributor_inventory.find({"distributor_id": distributor_id}, {"_id": 0}):
            if inv["product_id"] in hidden:
                continue
            if int(inv.get("qty_boxes", 0)) <= 0:
                continue
            p = await db.dms_products.find_one({"id": inv["product_id"]}, {"_id": 0})
            if not p or not p.get("active"):
                continue
            # distributor's SP = cost + margin (we allow override in retailer_price)
            sp_map = await db.dms_retailer_prices.find_one({"distributor_id": distributor_id, "product_id": p["id"]}, {"_id": 0})
            selling_price = _round(sp_map["selling_price"]) if sp_map else _round(inv.get("cost_price", p["unit_price"]) * 1.15)
            prods.append({
                **p,
                "distributor_stock_boxes": int(inv["qty_boxes"]),
                "selling_price": selling_price,
                "cost_price": inv.get("cost_price", p["unit_price"]),
            })
        return prods

    async def _get_retailer_selling_mode(distributor_id: str, retailer_id: str) -> str:
        doc = await db.dms_ret_mode.find_one({"distributor_id": distributor_id, "retailer_id": retailer_id}, {"_id": 0})
        return doc.get("mode", "box") if doc else "box"

    # ── retailer prices (distributor's selling price to retailers, configurable by owner/TL) ──
    @router.get("/distributors/{did}/retailer-prices")
    async def get_retailer_prices(did: str, user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role not in ("owner", "super_admin", "team_leader", "distributor") or (role == "distributor" and user.get("distributor_id") != did):
            raise HTTPException(status_code=403, detail="Forbidden")
        products = await db.dms_products.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(1000)
        price_map = {p["product_id"]: p["selling_price"] async for p in db.dms_retailer_prices.find({"distributor_id": did}, {"_id": 0})}
        # get purchase price from distributor inventory
        inv_map = {i["product_id"]: i.get("cost_price", 0) async for i in db.dms_distributor_inventory.find({"distributor_id": did}, {"_id": 0})}
        out = []
        for p in products:
            cp = inv_map.get(p["id"], p["unit_price"])
            out.append({
                "product_id": p["id"],
                "product_name": p["name"],
                "sku_code": p["sku_code"],
                "cost_price": _round(cp),
                "selling_price": price_map.get(p["id"], _round(cp * 1.15)),
            })
        return {"data": out}

    @router.put("/distributors/{did}/retailer-prices")
    async def set_retailer_price(did: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        # Owner or Team Leader can set; distributor cannot set their own selling price (per doc)
        if role not in ("owner", "super_admin", "team_leader"):
            raise HTTPException(status_code=403, detail="Only owner or team leader can set selling prices")
        pid = body.get("product_id"); sp = _round(body.get("selling_price", 0))
        if not pid or sp <= 0:
            raise HTTPException(status_code=400, detail="product_id + selling_price>0 required")
        await db.dms_retailer_prices.update_one(
            {"distributor_id": did, "product_id": pid},
            {"$set": {"distributor_id": did, "product_id": pid, "selling_price": sp, "updated_at": _now(), "updated_by": user["id"]}},
            upsert=True,
        )
        return {"ok": True}

    # ── retailers ──
    async def _get_dist_id_for_user(user: dict) -> Optional[str]:
        role = user.get("role")
        if role == "distributor":
            return user.get("distributor_id")
        if role == "distributor_accountant":
            return user.get("distributor_id")
        if role == "salesperson":
            # returns first assigned distributor
            a = await db.dms_sp_assignments.find_one({"salesperson_id": user["id"]}, {"_id": 0})
            return a.get("distributor_id") if a else None
        return None

    @router.get("/retailers")
    async def list_retailers(distributor_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        role = user.get("role")
        q: Dict[str, Any] = {"active": True}
        if role == "retailer":
            q["id"] = user.get("retailer_id")
        elif role in ("distributor", "distributor_accountant"):
            q["distributor_id"] = user.get("distributor_id")
        elif role == "salesperson":
            # retailers under my assigned distributors
            assigns = await db.dms_sp_assignments.find({"salesperson_id": user["id"]}, {"_id": 0}).to_list(500)
            dids = [a["distributor_id"] for a in assigns]
            q["distributor_id"] = {"$in": dids} if dids else "__none__"
        elif distributor_id:
            q["distributor_id"] = distributor_id
        docs = await db.dms_retailers.find(q, {"_id": 0}).sort("name", 1).to_list(1000)
        # enrich with login access status (from linked user account)
        uids = [d.get("user_id") for d in docs if d.get("user_id")]
        umap = {}
        if uids:
            async for u in db.users.find({"id": {"$in": uids}}, {"_id": 0, "id": 1, "login_enabled": 1}):
                umap[u["id"]] = u
        for d in docs:
            has_login = bool(d.get("user_id") or d.get("email"))
            d["has_login"] = has_login
            u = umap.get(d.get("user_id") or "")
            # default enabled unless explicitly disabled on the user account
            d["login_enabled"] = (u.get("login_enabled") is not False) if u else (d.get("login_enabled") is not False)
        return {"data": docs, "count": len(docs)}

    @router.post("/retailers")
    async def create_retailer(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role not in ("distributor", "salesperson", "owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        did = body.get("distributor_id") or await _get_dist_id_for_user(user)
        if not did:
            raise HTTPException(status_code=400, detail="distributor_id required")
        required = ["name", "phone", "address"]
        for k in required:
            if not body.get(k):
                raise HTTPException(status_code=400, detail=f"{k} required")
        # FULL-PROCESS ONBOARDING: if a login (email) is being created for this
        # retailer, all details + KYC + at least one document are mandatory first.
        if str(body.get("email") or "").strip():
            login_labels = {
                "region": "Region", "gstin": "GSTIN", "shop_license": "Shop License",
            }
            missing = [label for key, label in login_labels.items()
                       if not str(body.get(key) or "").strip()]
            if not str(body.get("password") or "").strip():
                missing.append("Login Password")
            if not body.get("documents"):
                missing.append("At least one uploaded Document")
            if missing:
                raise HTTPException(
                    status_code=400,
                    detail="Complete the full onboarding before creating the retailer login. Missing: " + ", ".join(missing),
                )
        rid = _nid("ret")
        # user login (optional)
        ruser_id = None
        if body.get("email"):
            try:
                ruser_id = await _create_dms_user(
                    email=body["email"],
                    password=body.get("password") or "Demo@2026",
                    name=body["name"],
                    role="retailer",
                    extra={"retailer_id": rid, "distributor_id": did, "phone": body["phone"]},
                )
            except HTTPException:
                # email already exists — link to existing user if it's a retailer
                existing = await db.users.find_one({"email": body["email"].lower()})
                if existing and existing.get("role") == "retailer":
                    ruser_id = existing["id"]
                    await db.users.update_one({"id": existing["id"]}, {"$set": {"retailer_id": rid, "distributor_id": did}})
                else:
                    raise
        doc = {
            "id": rid,
            "name": body["name"],
            "phone": body["phone"],
            "email": (body.get("email") or "").lower(),
            "address": body["address"],
            "region": body.get("region", ""),
            "gps_lat": body.get("gps_lat"),
            "gps_lng": body.get("gps_lng"),
            "location_link": body.get("location_link", ""),
            "distributor_id": did,
            "onboarded_by": user["id"],
            "onboarded_by_role": user["role"],
            "user_id": ruser_id,
            "kyc": {
                "gstin": body.get("gstin", ""),
                "shop_license": body.get("shop_license", ""),
                "notes": body.get("kyc_notes", ""),
            },
            # Bank + UPI (retailer's own — visible to owner; used when retailer bills a customer)
            "bank": body.get("bank") or {
                "bank_name": body.get("bank_name", ""),
                "bank_account": body.get("bank_account", ""),
                "bank_ifsc": body.get("bank_ifsc", ""),
                "bank_branch": body.get("bank_branch", ""),
                "upi_id": body.get("upi_id", ""),
                "upi_name": body.get("upi_name", ""),
                "qr_url": body.get("qr_url", ""),
            },
            "state": body.get("state", ""),
            "state_code": body.get("state_code", ""),
            "documents": body.get("documents", []),
            "credit_limit": _round(body.get("credit_limit", 0)),
            "active": True,
            "created_at": _now(),
        }
        await db.dms_retailers.insert_one(doc)
        return _clean(doc)

    @router.get("/retailers/{rid}")
    async def get_retailer(rid: str, user: dict = Depends(get_current_user)):
        doc = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Retailer not found")
        return doc

    @router.put("/retailers/{rid}")
    async def update_retailer(rid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        upd: Dict[str, Any] = {}
        for k in ["name", "phone", "address", "region", "gps_lat", "gps_lng", "location_link", "credit_limit", "active", "documents", "bank", "state", "state_code"]:
            if k in body:
                upd[k] = body[k]
        if "kyc" in body:
            upd["kyc"] = body["kyc"]
        upd["updated_at"] = _now()
        r = await db.dms_retailers.update_one({"id": rid}, {"$set": upd})
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Retailer not found")
        return {"ok": True}

    @router.delete("/retailers/{rid}")
    async def delete_retailer(rid: str, user: dict = Depends(get_current_user)):
        if user.get("role") not in ("owner", "owner_accountant", "distributor",
                                    "team_leader", "salesperson", "super_admin"):
            raise HTTPException(status_code=403, detail="Not allowed to delete retailers")
        ret = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not ret:
            raise HTTPException(status_code=404, detail="Retailer not found")
        order_count = await db.dms_secondary_orders.count_documents({"retailer_id": rid})
        if order_count:
            raise HTTPException(status_code=400,
                detail=f"Cannot delete — {order_count} order(s) exist for this retailer. Deactivate instead.")
        await db.dms_retailers.delete_one({"id": rid})
        await db.dms_ret_visibility.delete_many({"retailer_id": rid})
        uid = ret.get("user_id")
        if uid:
            await db.users.delete_one({"id": uid})
        await db.users.delete_many({"retailer_id": rid, "role": "retailer"})
        return {"ok": True, "deleted": rid}

    @router.put("/retailers/{rid}/login-access")
    async def set_retailer_login_access(rid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        """Owner toggles a retailer's login access ON/OFF (Item 3)."""
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        enabled = bool(body.get("enabled", True))
        await db.dms_retailers.update_one({"id": rid}, {"$set": {"login_enabled": enabled, "updated_at": _now()}})
        # mirror on the linked retailer user account (login lookup reads this)
        if retailer.get("user_id"):
            await db.users.update_one({"id": retailer["user_id"]}, {"$set": {"login_enabled": enabled}})
        elif retailer.get("email"):
            await db.users.update_one({"email": retailer["email"].lower(), "role": "retailer"}, {"$set": {"login_enabled": enabled}})
        return {"ok": True, "login_enabled": enabled}

    # ── retailer visibility (distributor controls) ──
    @router.get("/retailers/{rid}/visibility")
    async def get_ret_visibility(rid: str, user: dict = Depends(get_current_user)):
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        did = retailer["distributor_id"]
        # only distributor/dist_acct/owner can view
        if user["role"] not in ("owner", "super_admin", "team_leader") and user.get("distributor_id") != did:
            raise HTTPException(status_code=403, detail="Forbidden")
        products = await db.dms_products.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(1000)
        vis_map = {v["product_id"]: v.get("visible", True) async for v in db.dms_ret_visibility.find({"distributor_id": did, "retailer_id": rid}, {"_id": 0})}
        out = []
        for p in products:
            out.append({
                "product_id": p["id"],
                "product_name": p["name"],
                "sku_code": p["sku_code"],
                "visible": vis_map.get(p["id"], True),
            })
        return {"data": out}

    @router.put("/retailers/{rid}/visibility")
    async def set_ret_visibility(rid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        did = retailer["distributor_id"]
        if user["role"] not in ("owner", "super_admin") and user.get("distributor_id") != did:
            raise HTTPException(status_code=403, detail="Forbidden")
        pid = body.get("product_id"); visible = bool(body.get("visible", True))
        if not pid:
            raise HTTPException(status_code=400, detail="product_id required")
        await db.dms_ret_visibility.update_one(
            {"distributor_id": did, "retailer_id": rid, "product_id": pid},
            {"$set": {"distributor_id": did, "retailer_id": rid, "product_id": pid, "visible": visible, "updated_at": _now()}},
            upsert=True,
        )
        return {"ok": True}

    # ── retailer selling mode ──
    @router.get("/retailers/{rid}/selling-mode")
    async def get_ret_mode(rid: str, user: dict = Depends(get_current_user)):
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        did = retailer["distributor_id"]
        mode = await _get_retailer_selling_mode(did, rid)
        return {"mode": mode}

    @router.put("/retailers/{rid}/selling-mode")
    async def set_ret_mode(rid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        did = retailer["distributor_id"]
        if user["role"] not in ("owner", "super_admin") and user.get("distributor_id") != did:
            raise HTTPException(status_code=403, detail="Forbidden")
        mode = body.get("mode", "box")
        if mode not in ("box", "box_pcs"):
            raise HTTPException(status_code=400, detail="mode must be 'box' or 'box_pcs'")
        await db.dms_ret_mode.update_one(
            {"distributor_id": did, "retailer_id": rid},
            {"$set": {"distributor_id": did, "retailer_id": rid, "mode": mode, "updated_at": _now()}},
            upsert=True,
        )
        return {"ok": True, "mode": mode}

    # ── retailer browse ──
    @router.get("/retailer/browse")
    async def retailer_browse(retailer_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        # retailer sees their own; salesperson may pass retailer_id explicitly
        role = user.get("role")
        rid = retailer_id or user.get("retailer_id")
        if not rid:
            # Retailer with no linked profile yet → clean empty browse
            if role == "retailer":
                return {"data": [], "mode": None, "retailer": None, "pending": []}
            raise HTTPException(status_code=400, detail="retailer_id required")
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        # access check
        if role == "retailer" and user.get("retailer_id") != rid:
            raise HTTPException(status_code=403, detail="Forbidden")
        did = retailer["distributor_id"]
        products = await _resolve_visible_products_for_retailer(did, rid)
        cats = {c["id"]: c["name"] async for c in db.dms_categories.find({}, {"_id": 0, "id": 1, "name": 1})}
        for p in products:
            p["category_name"] = cats.get(p.get("category_id"), "")
        mode = await _get_retailer_selling_mode(did, rid)
        # pending qty for this retailer
        pending = []
        async for pd in db.dms_retailer_pending.find({"retailer_id": rid, "distributor_id": did}, {"_id": 0}):
            if int(pd.get("pending_qty_boxes", 0)) > 0 or int(pd.get("pending_qty_pcs", 0)) > 0:
                pending.append(pd)
        return {"data": products, "mode": mode, "retailer": retailer, "pending": pending}

    # ── secondary orders ──
    @router.post("/secondary-orders")
    async def place_secondary_order(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role not in ("retailer", "salesperson", "distributor", "owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        rid = body.get("retailer_id") or user.get("retailer_id")
        if not rid:
            raise HTTPException(status_code=400, detail="retailer_id required")
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        if role == "retailer" and user.get("retailer_id") != rid:
            raise HTTPException(status_code=403, detail="Forbidden")
        did = retailer["distributor_id"]
        items = body.get("items") or []
        include_pending = bool(body.get("include_pending", False))
        if not items and not include_pending:
            raise HTTPException(status_code=400, detail="items[] required")
        mode = await _get_retailer_selling_mode(did, rid)
        # merge with pending if include_pending
        pending_add: Dict[str, Dict[str, Any]] = {}
        if include_pending:
            async for pd in db.dms_retailer_pending.find({"retailer_id": rid, "distributor_id": did}, {"_id": 0}):
                if int(pd.get("pending_qty_boxes", 0)) > 0 or int(pd.get("pending_qty_pcs", 0)) > 0:
                    pending_add[pd["product_id"]] = pd

        order_items = []
        subtotal = 0.0
        gst_total = 0.0
        # global GST% from settings
        _s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0}) or {}
        global_gst = float(_s.get("gst_pct") or 0)
        for it in items:
            pid = it.get("product_id")
            qty_boxes = int(it.get("qty_boxes", 0))
            qty_pcs = int(it.get("qty_pcs", 0)) if mode == "box_pcs" else 0
            if not pid or (qty_boxes == 0 and qty_pcs == 0):
                continue
            p = await db.dms_products.find_one({"id": pid}, {"_id": 0})
            if not p:
                continue
            sp_map = await db.dms_retailer_prices.find_one({"distributor_id": did, "product_id": pid}, {"_id": 0})
            box_price = _round(sp_map["selling_price"]) if sp_map else _round(p["unit_price"] * 1.15)
            pcs_price = _round(box_price / max(p["box_qty"], 1))
            # add pending
            pend = pending_add.pop(pid, None)
            if pend:
                qty_boxes += int(pend.get("pending_qty_boxes", 0))
                qty_pcs += int(pend.get("pending_qty_pcs", 0))
            line_sub = _round(box_price * qty_boxes + pcs_price * qty_pcs)
            line_gst = _round(line_sub * (global_gst / 100.0))
            subtotal += line_sub; gst_total += line_gst
            order_items.append({
                "product_id": pid,
                "product_name": p["name"],
                "sku_code": p["sku_code"],
                "box_qty": p["box_qty"],
                "box_price": box_price,
                "pcs_price": pcs_price,
                "gst_pct": global_gst,
                "qty_boxes_ordered": qty_boxes,
                "qty_pcs_ordered": qty_pcs,
                "qty_boxes_dispatched": 0,
                "qty_pcs_dispatched": 0,
                "line_subtotal": line_sub,
                "line_gst": line_gst,
                "line_total": _round(line_sub + line_gst),
                "carried_pending": bool(pend),
            })
        # any remaining pending items also included?
        for pid, pend in pending_add.items():
            p = await db.dms_products.find_one({"id": pid}, {"_id": 0})
            if not p:
                continue
            sp_map = await db.dms_retailer_prices.find_one({"distributor_id": did, "product_id": pid}, {"_id": 0})
            box_price = _round(sp_map["selling_price"]) if sp_map else _round(p["unit_price"] * 1.15)
            pcs_price = _round(box_price / max(p["box_qty"], 1))
            qb = int(pend.get("pending_qty_boxes", 0)); qp = int(pend.get("pending_qty_pcs", 0))
            line_sub = _round(box_price * qb + pcs_price * qp)
            line_gst = _round(line_sub * (global_gst / 100.0))
            subtotal += line_sub; gst_total += line_gst
            order_items.append({
                "product_id": pid, "product_name": p["name"], "sku_code": p["sku_code"],
                "box_qty": p["box_qty"], "box_price": box_price, "pcs_price": pcs_price,
                "gst_pct": global_gst, "qty_boxes_ordered": qb, "qty_pcs_ordered": qp,
                "qty_boxes_dispatched": 0, "qty_pcs_dispatched": 0,
                "line_subtotal": line_sub, "line_gst": line_gst, "line_total": _round(line_sub + line_gst),
                "carried_pending": True,
            })

        if not order_items:
            raise HTTPException(status_code=400, detail="No valid items")
        total = _round(subtotal + gst_total)
        order = {
            "id": _nid("so"),
            "order_no": f"SO-{datetime.now().strftime('%y%m%d%H%M%S')}",
            "retailer_id": rid,
            "retailer_name": retailer["name"],
            "distributor_id": did,
            "mode": mode,
            "items": order_items,
            "subtotal": _round(subtotal),
            "gst_total": _round(gst_total),
            "total": total,
            "status": "pending",
            "fulfillment_pct": 0,
            "notes": body.get("notes", ""),
            "placed_by": user["id"],
            "placed_by_role": user["role"],
            "created_at": _now(),
        }
        await db.dms_secondary_orders.insert_one(order)
        # if pending was included, mark those pending records consumed
        if include_pending:
            await db.dms_retailer_pending.update_many(
                {"retailer_id": rid, "distributor_id": did},
                {"$set": {"pending_qty_boxes": 0, "pending_qty_pcs": 0, "consumed_at": _now(), "consumed_by_order": order["id"]}},
            )
        # notify distributor + dist accountant
        async for u in db.users.find({"distributor_id": did, "role": {"$in": ["distributor", "distributor_accountant"]}}, {"_id": 0, "id": 1}):
            await notify(u["id"], "secondary_order", f"New order from {retailer['name']}",
                         f"{order['order_no']} — {len(order_items)} items — \u20b9{total:,.0f}",
                         f"/dms/distributor/retail-orders/{order['id']}")
        return _clean(order)

    @router.get("/secondary-orders")
    async def list_secondary_orders(status: Optional[str] = None, user: dict = Depends(get_current_user)):
        role = user.get("role")
        q: Dict[str, Any] = {}
        if status:
            q["status"] = status
        if role == "retailer":
            q["retailer_id"] = user.get("retailer_id")
        elif role in ("distributor", "distributor_accountant"):
            q["distributor_id"] = user.get("distributor_id")
        elif role == "salesperson":
            # BUG FIX (Phase 1): SP should see orders they placed AND orders under their assigned distributors
            assigns = await db.dms_sp_assignments.find({"salesperson_id": user["id"]}, {"_id": 0}).to_list(500)
            dids = [a["distributor_id"] for a in assigns]
            or_clauses = [{"placed_by": user["id"]}]
            if dids:
                or_clauses.append({"distributor_id": {"$in": dids}})
            q["$or"] = or_clauses
        docs = await db.dms_secondary_orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
        # Enrich with placed_by user name and distributor name for UI convenience
        placed_ids = list({d.get("placed_by") for d in docs if d.get("placed_by")})
        placed_map = {}
        if placed_ids:
            async for u in db.users.find({"id": {"$in": placed_ids}}, {"_id": 0, "id": 1, "name": 1, "role": 1}):
                placed_map[u["id"]] = {"name": u.get("name"), "role": u.get("role")}
        dist_ids = list({d.get("distributor_id") for d in docs if d.get("distributor_id")})
        dist_map = {}
        if dist_ids:
            async for dd in db.dms_distributors.find({"id": {"$in": dist_ids}}, {"_id": 0, "id": 1, "name": 1}):
                dist_map[dd["id"]] = dd.get("name")
        for d in docs:
            pb = placed_map.get(d.get("placed_by"))
            d["placed_by_name"] = pb.get("name") if pb else None
            d["distributor_name"] = d.get("distributor_name") or dist_map.get(d.get("distributor_id"))
        return {"data": docs, "count": len(docs)}

    @router.get("/secondary-orders/{oid}")
    async def get_secondary_order(oid: str, user: dict = Depends(get_current_user)):
        doc = await db.dms_secondary_orders.find_one({"id": oid}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Order not found")
        role = user.get("role")
        if role == "retailer" and doc["retailer_id"] != user.get("retailer_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        if role in ("distributor", "distributor_accountant") and doc["distributor_id"] != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        atts = await db.dms_attachments.find({"reference_id": oid}, {"_id": 0}).to_list(50)
        doc["attachments"] = atts
        # enrich retailer & distributor names
        r = await db.dms_retailers.find_one({"id": doc["retailer_id"]}, {"_id": 0, "name": 1, "phone": 1, "address": 1})
        doc["retailer"] = r
        d = await db.dms_distributors.find_one({"id": doc["distributor_id"]}, {"_id": 0, "name": 1, "phone": 1, "address": 1, "kyc": 1})
        doc["distributor"] = d
        # Phase 1: enrich placed_by user info for TL Order Monitoring detail
        if doc.get("placed_by"):
            pu = await db.users.find_one({"id": doc["placed_by"]}, {"_id": 0, "id": 1, "name": 1, "role": 1, "phone": 1})
            if pu:
                doc["placed_by_user"] = pu
                doc["placed_by_name"] = pu.get("name")
        return doc

    @router.post("/secondary-orders/{oid}/invoice")
    async def generate_invoice_secondary(oid: str, body: Dict[str, Any] = Body(default={}), user: dict = Depends(get_current_user)):
        """Step 1 of the retailer-order flow: generate the Invoice (short number INV-0001).
        Body (optional): {items: [{product_id, qty_boxes, qty_pcs}]} — quantities to invoice.
        If omitted, the full ordered quantity is invoiced. No stock movement here."""
        role = user.get("role")
        if role not in ("distributor", "distributor_accountant", "owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        order = await db.dms_secondary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if role in ("distributor", "distributor_accountant") and user.get("distributor_id") != order["distributor_id"]:
            raise HTTPException(status_code=403, detail="Forbidden")
        if order.get("status") != "pending":
            raise HTTPException(status_code=400, detail=f"Invoice can only be generated for a pending order (current: {order.get('status')})")
        did = order["distributor_id"]; rid = order["retailer_id"]
        req_items = body.get("items") or []
        item_map = {it["product_id"]: it for it in req_items}
        billed_items = []
        subtotal = 0.0; gst_total = 0.0
        for it in order["items"]:
            if req_items:
                di = item_map.get(it["product_id"], {})
                inv_boxes = int(di.get("qty_boxes", 0))
                inv_pcs = int(di.get("qty_pcs", 0)) if order["mode"] == "box_pcs" else 0
            else:
                inv_boxes = int(it["qty_boxes_ordered"])
                inv_pcs = int(it["qty_pcs_ordered"]) if order["mode"] == "box_pcs" else 0
            inv_boxes = min(inv_boxes, it["qty_boxes_ordered"])
            inv_pcs = min(inv_pcs, it["qty_pcs_ordered"])
            it["qty_boxes_invoiced"] = inv_boxes
            it["qty_pcs_invoiced"] = inv_pcs
            line_sub = _round(it["box_price"] * inv_boxes + it["pcs_price"] * inv_pcs)
            line_gst = _round(line_sub * (it["gst_pct"] / 100.0))
            subtotal += line_sub; gst_total += line_gst
            billed_items.append({
                **it,
                "invoiced_qty_boxes": inv_boxes, "invoiced_qty_pcs": inv_pcs,
                "line_subtotal": line_sub, "line_gst": line_gst, "line_total": _round(line_sub + line_gst),
            })
            # pending qty (ordered - invoiced)
            pending_boxes = it["qty_boxes_ordered"] - inv_boxes
            pending_pcs = it["qty_pcs_ordered"] - inv_pcs
            if pending_boxes > 0 or pending_pcs > 0:
                await db.dms_retailer_pending.update_one(
                    {"retailer_id": rid, "distributor_id": did, "product_id": it["product_id"]},
                    {"$set": {
                        "retailer_id": rid, "distributor_id": did, "product_id": it["product_id"],
                        "pending_qty_boxes": pending_boxes, "pending_qty_pcs": pending_pcs,
                        "product_name": it["product_name"], "sku_code": it["sku_code"],
                        "updated_at": _now(),
                    }},
                    upsert=True,
                )
        total = _round(subtotal + gst_total)
        invoice_no = await _next_no("retailer_invoice", "INV", 4)
        bill = {
            "id": _nid("rb"), "bill_no": invoice_no,
            "order_id": oid, "order_no": order["order_no"],
            "retailer_id": rid, "distributor_id": did,
            "items": billed_items, "subtotal": _round(subtotal), "gst_total": _round(gst_total), "total": total,
            "status": "issued", "created_at": _now(),
        }
        await db.dms_retailer_bills.insert_one(bill)
        await db.dms_retailer_ledger.insert_one({
            "id": _nid("rle"),
            "distributor_id": did, "retailer_id": rid,
            "kind": "invoice", "reference_id": bill["id"], "reference_no": invoice_no,
            "amount": total, "description": f"Invoice {invoice_no} for {order['order_no']}", "at": _now(),
        })
        await db.dms_secondary_orders.update_one(
            {"id": oid},
            {"$set": {"items": order["items"], "status": "invoiced",
                      "bill_id": bill["id"], "invoice_no": invoice_no, "invoiced_at": _now(), "updated_at": _now()}},
        )
        r_user = await db.users.find_one({"retailer_id": rid, "role": "retailer"}, {"_id": 0, "id": 1})
        if r_user:
            await notify(r_user["id"], "order_invoiced", f"Invoice {invoice_no} generated",
                         f"Order {order['order_no']} \u2022 \u20b9{total:,.0f}",
                         f"/dms/retailer/my-orders/{oid}")
        return {"ok": True, "bill_id": bill["id"], "invoice_no": invoice_no, "total": total, "status": "invoiced"}

    @router.post("/secondary-orders/{oid}/dispatch")
    async def dispatch_secondary(oid: str, body: Dict[str, Any] = Body(default={}), user: dict = Depends(get_current_user)):
        """Step 2: Dispatch an invoiced order. Decrements distributor stock and
        auto-generates a Delivery Challan (short number DC-0001)."""
        role = user.get("role")
        if role not in ("distributor", "distributor_accountant", "owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        order = await db.dms_secondary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if role in ("distributor", "distributor_accountant") and user.get("distributor_id") != order["distributor_id"]:
            raise HTTPException(status_code=403, detail="Forbidden")
        if order["status"] in ("dispatched", "completed"):
            raise HTTPException(status_code=400, detail="Already dispatched")
        if order["status"] != "invoiced":
            raise HTTPException(status_code=400, detail="Generate the Invoice before dispatching this order")
        did = order["distributor_id"]; rid = order["retailer_id"]
        # Stop-sale-on-negative pre-check for distributor stock (using invoiced quantities)
        if await _stop_sale_enabled():
            for it in order["items"]:
                db_qty = int(it.get("qty_boxes_invoiced", 0))
                dp_qty = int(it.get("qty_pcs_invoiced", 0)) if order["mode"] == "box_pcs" else 0
                if db_qty <= 0 and dp_qty <= 0:
                    continue
                inv = await db.dms_distributor_inventory.find_one({"distributor_id": did, "product_id": it["product_id"]}, {"_id": 0, "qty_boxes": 1})
                avail = int((inv or {}).get("qty_boxes", 0) or 0)
                need_boxes = db_qty + (1 if (dp_qty > 0 and (dp_qty % max(it["box_qty"], 1)) > 0) else 0) + (dp_qty // max(it["box_qty"], 1))
                if need_boxes > avail:
                    raise HTTPException(status_code=400, detail=f"Insufficient distributor stock for {it.get('product_name', it['product_id'])}: available {avail} boxes, need {need_boxes}. Reduce dispatch qty or receive more stock.")
        # apply — dispatch the invoiced quantities and decrement stock
        challan_items = []
        for it in order["items"]:
            db_qty = int(it.get("qty_boxes_invoiced", 0))
            dp_qty = int(it.get("qty_pcs_invoiced", 0)) if order["mode"] == "box_pcs" else 0
            it["qty_boxes_dispatched"] = db_qty
            it["qty_pcs_dispatched"] = dp_qty
            total_pcs = db_qty * it["box_qty"] + dp_qty
            if total_pcs > 0:
                inv = await db.dms_distributor_inventory.find_one({"distributor_id": did, "product_id": it["product_id"]})
                if inv:
                    remaining_pcs = dp_qty % max(it["box_qty"], 1)
                    new_qty = max(0, int(inv.get("qty_boxes", 0)) - db_qty - (1 if remaining_pcs > 0 else 0))
                    await db.dms_distributor_inventory.update_one({"id": inv["id"]}, {"$set": {"qty_boxes": new_qty, "updated_at": _now()}})
                    await db.dms_stock_ledger.insert_one({
                        "id": _nid("sl"), "scope": "distributor", "distributor_id": did,
                        "product_id": it["product_id"], "delta_boxes": -(db_qty + (1 if remaining_pcs > 0 else 0)),
                        "reason": "secondary_dispatch", "reference": order["order_no"], "at": _now(),
                    })
            challan_items.append({
                "product_id": it["product_id"], "product_name": it["product_name"], "sku_code": it["sku_code"],
                "qty_boxes": db_qty, "qty_pcs": dp_qty,
            })
        # fulfillment
        ord_total_pcs = sum(it["qty_boxes_ordered"] * it["box_qty"] + it["qty_pcs_ordered"] for it in order["items"])
        disp_total_pcs = sum(it["qty_boxes_dispatched"] * it["box_qty"] + it["qty_pcs_dispatched"] for it in order["items"])
        pct = int(round((disp_total_pcs / ord_total_pcs) * 100)) if ord_total_pcs > 0 else 0
        # auto-generate Delivery Challan
        challan_no = await _next_no("delivery_challan", "DC", 4)
        challan = {
            "id": _nid("dc"), "challan_no": challan_no,
            "order_id": oid, "order_no": order["order_no"],
            "invoice_id": order.get("bill_id"), "invoice_no": order.get("invoice_no"),
            "retailer_id": rid, "distributor_id": did,
            "items": challan_items, "created_at": _now(), "created_by": user["id"],
        }
        await db.dms_delivery_challans.insert_one(challan)
        await db.dms_secondary_orders.update_one(
            {"id": oid},
            {"$set": {"items": order["items"], "fulfillment_pct": pct, "status": "dispatched",
                      "challan_id": challan["id"], "challan_no": challan_no,
                      "dispatched_at": _now(), "updated_at": _now()}},
        )
        r_user = await db.users.find_one({"retailer_id": rid, "role": "retailer"}, {"_id": 0, "id": 1})
        if r_user:
            await notify(r_user["id"], "order_dispatched", f"Order {order['order_no']} dispatched",
                         f"Challan {challan_no} \u2022 Invoice {order.get('invoice_no', '')}",
                         f"/dms/retailer/my-orders/{oid}")
        return {"ok": True, "bill_id": order.get("bill_id"), "invoice_no": order.get("invoice_no"),
                "challan_id": challan["id"], "challan_no": challan_no, "fulfillment_pct": pct, "status": "dispatched"}

    @router.get("/secondary-orders/{oid}/challan")
    async def get_order_challan(oid: str, user: dict = Depends(get_current_user)):
        c = await db.dms_delivery_challans.find_one({"order_id": oid}, {"_id": 0}, sort=[("created_at", -1)])
        if not c:
            raise HTTPException(status_code=404, detail="No delivery challan for this order")
        return c

    @router.get("/print/challan/{challan_id}")
    async def print_challan(challan_id: str, user: dict = Depends(get_current_user)):
        c = await db.dms_delivery_challans.find_one({"id": challan_id}, {"_id": 0})
        if not c:
            raise HTTPException(status_code=404, detail="Challan not found")
        r = await db.dms_retailers.find_one({"id": c.get("retailer_id")}, {"_id": 0})
        d = await db.dms_distributors.find_one({"id": c.get("distributor_id")}, {"_id": 0})
        s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0}) or {}
        c["retailer"] = r; c["distributor"] = d
        c["company_name"] = s.get("company_name") or "GO OIL Lubricants"
        return c

    # ── Cancel / Edit secondary order (Phase 1 additions) ──
    @router.post("/secondary-orders/{oid}/cancel")
    async def cancel_secondary_order(oid: str, body: Dict[str, Any] = Body(default={}), user: dict = Depends(get_current_user)):
        """Cancel a pending secondary (retailer) order. Allowed: TL (assigned distributor), SP (own placed_by), Owner, Super Admin, Distributor (own)."""
        role = user.get("role")
        order = await db.dms_secondary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if order.get("status") in ("dispatched", "delivered", "completed", "cancelled"):
            raise HTTPException(status_code=400, detail=f"Cannot cancel — order is {order.get('status')}")
        # Phase 2A: FY lock check
        await _check_fy_lock(order.get("created_at"), "order")

        # RBAC
        allowed = False
        if role in ("owner", "super_admin"):
            allowed = True
        elif role in ("distributor", "distributor_accountant") and user.get("distributor_id") == order.get("distributor_id"):
            allowed = True
        elif role == "team_leader":
            tl_dists = [a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0, "distributor_id": 1})]
            if order.get("distributor_id") in tl_dists:
                allowed = True
        elif role == "salesperson":
            if order.get("placed_by") == user["id"]:
                allowed = True
        if not allowed:
            raise HTTPException(status_code=403, detail="Forbidden")

        reason = str(body.get("reason") or "").strip() or "Cancelled"
        await db.dms_secondary_orders.update_one(
            {"id": oid},
            {"$set": {
                "status": "cancelled",
                "cancelled_at": _now(),
                "cancelled_by": user["id"],
                "cancelled_by_role": role,
                "cancel_reason": reason,
                "updated_at": _now(),
            }},
        )
        # notify distributor
        async for u in db.users.find({"distributor_id": order["distributor_id"], "role": {"$in": ["distributor", "distributor_accountant"]}}, {"_id": 0, "id": 1}):
            await notify(u["id"], "order_cancelled", f"Order {order['order_no']} cancelled",
                         f"Reason: {reason}", f"/dms/distributor/retail-orders/{oid}")
        return {"ok": True, "status": "cancelled"}

    @router.put("/secondary-orders/{oid}")
    async def edit_secondary_order(oid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        """Edit a pending secondary order's items. Allowed: SP who placed, TL (assigned dist), Owner/SuperAdmin, Distributor."""
        role = user.get("role")
        order = await db.dms_secondary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if order.get("status") != "pending":
            raise HTTPException(status_code=400, detail=f"Can only edit pending orders (current: {order.get('status')})")
        # Phase 2A: FY lock check
        await _check_fy_lock(order.get("created_at"), "order")

        # RBAC (same as cancel)
        allowed = False
        if role in ("owner", "super_admin"):
            allowed = True
        elif role in ("distributor", "distributor_accountant") and user.get("distributor_id") == order.get("distributor_id"):
            allowed = True
        elif role == "team_leader":
            tl_dists = [a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0, "distributor_id": 1})]
            if order.get("distributor_id") in tl_dists:
                allowed = True
        elif role == "salesperson" and order.get("placed_by") == user["id"]:
            allowed = True
        if not allowed:
            raise HTTPException(status_code=403, detail="Forbidden")

        did = order["distributor_id"]
        rid = order["retailer_id"]
        mode = order.get("mode", "box")
        items = body.get("items") or []
        if not items:
            raise HTTPException(status_code=400, detail="items[] required")
        _s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0}) or {}
        global_gst = float(_s.get("gst_pct") or 0)
        new_items = []
        subtotal = 0.0
        gst_total = 0.0
        for it in items:
            pid = it.get("product_id")
            qty_boxes = int(it.get("qty_boxes", 0))
            qty_pcs = int(it.get("qty_pcs", 0)) if mode == "box_pcs" else 0
            # Phase 1: also honor qty_pcs even in box mode if explicitly sent (user asked Box + Nos always available)
            if mode != "box_pcs" and it.get("qty_pcs"):
                qty_pcs = int(it.get("qty_pcs", 0))
            if not pid or (qty_boxes == 0 and qty_pcs == 0):
                continue
            p = await db.dms_products.find_one({"id": pid}, {"_id": 0})
            if not p:
                continue
            sp_map = await db.dms_retailer_prices.find_one({"distributor_id": did, "product_id": pid}, {"_id": 0})
            box_price = _round(sp_map["selling_price"]) if sp_map else _round(p["unit_price"] * 1.15)
            pcs_price = _round(box_price / max(p["box_qty"], 1))
            line_sub = _round(box_price * qty_boxes + pcs_price * qty_pcs)
            line_gst = _round(line_sub * (global_gst / 100.0))
            subtotal += line_sub; gst_total += line_gst
            new_items.append({
                "product_id": pid, "product_name": p["name"], "sku_code": p["sku_code"],
                "box_qty": p["box_qty"], "box_price": box_price, "pcs_price": pcs_price,
                "gst_pct": global_gst, "qty_boxes_ordered": qty_boxes, "qty_pcs_ordered": qty_pcs,
                "qty_boxes_dispatched": 0, "qty_pcs_dispatched": 0,
                "line_subtotal": line_sub, "line_gst": line_gst, "line_total": _round(line_sub + line_gst),
                "carried_pending": False,
            })
        if not new_items:
            raise HTTPException(status_code=400, detail="No valid items")
        total = _round(subtotal + gst_total)
        await db.dms_secondary_orders.update_one(
            {"id": oid},
            {"$set": {
                "items": new_items,
                "subtotal": _round(subtotal), "gst_total": _round(gst_total), "total": total,
                "notes": body.get("notes", order.get("notes", "")),
                "edited_at": _now(),
                "edited_by": user["id"],
                "updated_at": _now(),
            }},
        )
        return {"ok": True, "id": oid, "total": total, "item_count": len(new_items)}

    # ── secondary ledger ──
    @router.get("/ledger/secondary")
    async def secondary_ledger(retailer_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        role = user.get("role")
        q: Dict[str, Any] = {}
        if role == "retailer":
            q["retailer_id"] = user.get("retailer_id")
        elif role in ("distributor", "distributor_accountant"):
            q["distributor_id"] = user.get("distributor_id")
        if retailer_id and role not in ("retailer",):
            q["retailer_id"] = retailer_id
        entries = await db.dms_retailer_ledger.find(q, {"_id": 0}).sort("at", -1).to_list(2000)
        # summary per retailer
        summary: Dict[str, Dict[str, Any]] = {}
        for e in entries:
            rid = e["retailer_id"]
            s = summary.setdefault(rid, {"retailer_id": rid, "billed": 0.0, "paid": 0.0, "outstanding": 0.0})
            if e["kind"] in ("invoice", "debit_note"):
                s["billed"] += e["amount"]; s["outstanding"] += e["amount"]
            elif e["kind"] in ("payment", "coupon_credit", "credit_note"):
                s["paid"] += e["amount"]; s["outstanding"] -= e["amount"]
        rids = list(summary.keys())
        rnames = {r["id"]: r["name"] async for r in db.dms_retailers.find({"id": {"$in": rids}}, {"_id": 0, "id": 1, "name": 1})}
        for s in summary.values():
            s["retailer_name"] = rnames.get(s["retailer_id"], "")
            for k in ("billed", "paid", "outstanding"):
                s[k] = _round(s[k])
        return {"entries": entries, "summary": sorted(summary.values(), key=lambda x: -x["outstanding"])}

    @router.post("/ledger/secondary/payment")
    async def record_secondary_payment(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        # Phase 1: allow salesperson to record cash payment collected from retailer
        if role not in ("distributor", "distributor_accountant", "owner", "super_admin", "salesperson"):
            raise HTTPException(status_code=403, detail="Forbidden")
        rid = body.get("retailer_id")
        amt = _round(body.get("amount", 0))
        if not rid or amt <= 0:
            raise HTTPException(status_code=400, detail="retailer_id + amount>0 required")
        retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        # If SP: retailer must be under one of SP's assigned distributors
        if role == "salesperson":
            assigns = await db.dms_sp_assignments.find({"salesperson_id": user["id"]}, {"_id": 0, "distributor_id": 1}).to_list(500)
            dids = [a["distributor_id"] for a in assigns]
            if retailer.get("distributor_id") not in dids:
                raise HTTPException(status_code=403, detail="Retailer not in your assigned distributors")
        entry = {
            "id": _nid("rle"),
            "distributor_id": retailer["distributor_id"], "retailer_id": rid,
            "kind": "payment",
            "reference_no": body.get("reference_no", f"PMT-{datetime.now().strftime('%y%m%d%H%M%S')}"),
            "amount": amt, "method": body.get("method", "cash" if role == "salesperson" else "bank_transfer"),
            "description": body.get("description", "Payment received"),
            "at": _now(), "recorded_by": user["id"], "recorded_by_role": role,
        }
        await db.dms_retailer_ledger.insert_one(entry)
        return _clean(entry)

    # =========================================================================
    # SALES TEAM (Salesperson + Team Leader + Regional Manager)
    # =========================================================================

    # ── assignments ──
    @router.get("/assignments/tl-distributors")
    async def list_tl_dist_assignments(team_leader_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        q: Dict[str, Any] = {}
        if team_leader_id:
            q["team_leader_id"] = team_leader_id
        elif user["role"] == "team_leader":
            q["team_leader_id"] = user["id"]
        docs = await db.dms_tl_assignments.find(q, {"_id": 0}).to_list(500)
        return {"data": docs}

    @router.post("/assignments/tl-distributors")
    async def assign_tl_dist(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        if user["role"] not in ("owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Only owner can assign distributors to team leaders")
        tid = body.get("team_leader_id"); did = body.get("distributor_id")
        if not tid or not did:
            raise HTTPException(status_code=400, detail="team_leader_id + distributor_id required")
        await db.dms_tl_assignments.update_one(
            {"team_leader_id": tid, "distributor_id": did},
            {"$set": {"team_leader_id": tid, "distributor_id": did, "assigned_by": user["id"], "at": _now()}},
            upsert=True,
        )
        return {"ok": True}

    @router.delete("/assignments/tl-distributors")
    async def unassign_tl_dist(team_leader_id: str, distributor_id: str, user: dict = Depends(get_current_user)):
        if user["role"] not in ("owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        await db.dms_tl_assignments.delete_one({"team_leader_id": team_leader_id, "distributor_id": distributor_id})
        return {"ok": True}

    @router.get("/assignments/sp-distributors")
    async def list_sp_assignments(salesperson_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        q: Dict[str, Any] = {}
        if salesperson_id:
            q["salesperson_id"] = salesperson_id
        elif user["role"] == "salesperson":
            q["salesperson_id"] = user["id"]
        docs = await db.dms_sp_assignments.find(q, {"_id": 0}).to_list(500)
        return {"data": docs}

    @router.post("/assignments/sp-distributors")
    async def assign_sp_dist(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        if user["role"] not in ("owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Only owner can assign salespersons to distributors")
        spid = body.get("salesperson_id"); did = body.get("distributor_id")
        if not spid or not did:
            raise HTTPException(status_code=400, detail="salesperson_id + distributor_id required")
        await db.dms_sp_assignments.update_one(
            {"salesperson_id": spid, "distributor_id": did},
            {"$set": {"salesperson_id": spid, "distributor_id": did, "assigned_by": user["id"], "at": _now()}},
            upsert=True,
        )
        return {"ok": True}

    @router.delete("/assignments/sp-distributors")
    async def unassign_sp_dist(salesperson_id: str, distributor_id: str, user: dict = Depends(get_current_user)):
        if user["role"] not in ("owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        await db.dms_sp_assignments.delete_one({"salesperson_id": salesperson_id, "distributor_id": distributor_id})
        return {"ok": True}

    # ── consolidated hierarchy tree (owner view) ──
    # ── bulk move retailers to another distributor ──
    @router.post("/owner/retailers/bulk-assign-distributor")
    async def bulk_assign_retailers(body: Dict[str, Any] = Body(...),
                                    user: dict = Depends(owner_or_accountant)):
        retailer_ids = body.get("retailer_ids") or []
        distributor_id = body.get("distributor_id")
        if not retailer_ids or not distributor_id:
            raise HTTPException(status_code=400, detail="retailer_ids[] and distributor_id required")
        dist = await db.dms_distributors.find_one({"id": distributor_id}, {"_id": 0, "id": 1})
        if not dist:
            raise HTTPException(status_code=404, detail="Distributor not found")
        res = await db.dms_retailers.update_many(
            {"id": {"$in": retailer_ids}},
            {"$set": {"distributor_id": distributor_id, "updated_at": _now()}})
        await db.users.update_many(
            {"retailer_id": {"$in": retailer_ids}, "role": "retailer"},
            {"$set": {"distributor_id": distributor_id}})
        return {"ok": True, "moved": res.modified_count, "distributor_id": distributor_id}

    @router.get("/owner/hierarchy")
    async def owner_hierarchy(user: dict = Depends(owner_or_accountant)):
        """Full org tree: Regional Managers → Team Leaders → Distributors →
        (Salespersons + Retailers). Also returns flat lists of all users/entities
        so the UI can offer assignment dropdowns."""
        users = await db.users.find(
            {"tenant_id": DMS_TENANT_ID},
            {"_id": 0, "id": 1, "name": 1, "role": 1, "email": 1}).to_list(2000)
        by_role: Dict[str, List[Dict[str, Any]]] = {}
        umap: Dict[str, Dict[str, Any]] = {}
        for u in users:
            umap[u["id"]] = u
            by_role.setdefault(u.get("role"), []).append(u)

        distributors = await db.dms_distributors.find(
            {}, {"_id": 0, "id": 1, "name": 1, "region": 1}).to_list(1000)
        retailers = await db.dms_retailers.find(
            {}, {"_id": 0, "id": 1, "name": 1, "distributor_id": 1, "region": 1}).to_list(5000)

        sp_asg = await db.dms_sp_assignments.find({}, {"_id": 0}).to_list(5000)
        tl_asg = await db.dms_tl_assignments.find({}, {"_id": 0}).to_list(5000)
        rm_asg = await db.dms_rm_assignments.find({}, {"_id": 0}).to_list(5000)

        # index
        ret_by_dist: Dict[str, List[Dict[str, Any]]] = {}
        for r in retailers:
            ret_by_dist.setdefault(r.get("distributor_id"), []).append(
                {"id": r["id"], "name": r.get("name"), "region": r.get("region")})
        sp_by_dist: Dict[str, List[str]] = {}
        for a in sp_asg:
            sp_by_dist.setdefault(a["distributor_id"], []).append(a["salesperson_id"])
        dist_by_tl: Dict[str, List[str]] = {}
        for a in tl_asg:
            dist_by_tl.setdefault(a["team_leader_id"], []).append(a["distributor_id"])
        tl_by_rm: Dict[str, List[str]] = {}
        for a in rm_asg:
            tl_by_rm.setdefault(a["regional_manager_id"], []).append(a["team_leader_id"])

        dmap = {d["id"]: d for d in distributors}

        def _sp_list(did):
            return [{"id": sid, "name": (umap.get(sid) or {}).get("name", sid)}
                    for sid in sp_by_dist.get(did, [])]

        def _dist_node(did):
            d = dmap.get(did, {"id": did, "name": did})
            return {"id": did, "name": d.get("name"), "region": d.get("region"),
                    "salespersons": _sp_list(did),
                    "retailers": ret_by_dist.get(did, [])}

        assigned_dist_ids = set()
        for lst in dist_by_tl.values():
            assigned_dist_ids.update(lst)

        tree = []
        for rm in by_role.get("regional_manager", []):
            tl_nodes = []
            for tlid in tl_by_rm.get(rm["id"], []):
                tl = umap.get(tlid, {"id": tlid, "name": tlid})
                tl_nodes.append({"id": tlid, "name": tl.get("name"),
                                 "distributors": [_dist_node(x) for x in dist_by_tl.get(tlid, [])]})
            tree.append({"id": rm["id"], "name": rm.get("name"), "team_leaders": tl_nodes})

        # team leaders not under any RM
        assigned_tl_ids = {t for lst in tl_by_rm.values() for t in lst}
        loose_tls = []
        for tl in by_role.get("team_leader", []):
            if tl["id"] not in assigned_tl_ids:
                loose_tls.append({"id": tl["id"], "name": tl.get("name"),
                                  "distributors": [_dist_node(x) for x in dist_by_tl.get(tl["id"], [])]})

        # distributors not under any TL
        loose_dists = [_dist_node(d["id"]) for d in distributors if d["id"] not in assigned_dist_ids]

        return {
            "tree": tree,
            "unassigned_team_leaders": loose_tls,
            "unassigned_distributors": loose_dists,
            "all": {
                "regional_managers": by_role.get("regional_manager", []),
                "team_leaders": by_role.get("team_leader", []),
                "salespersons": by_role.get("salesperson", []),
                "distributors": [{"id": d["id"], "name": d.get("name")} for d in distributors],
            },
        }

    async def list_rm_tl_assignments(regional_manager_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        q: Dict[str, Any] = {}
        if regional_manager_id:
            q["regional_manager_id"] = regional_manager_id
        elif user["role"] == "regional_manager":
            q["regional_manager_id"] = user["id"]
        docs = await db.dms_rm_assignments.find(q, {"_id": 0}).to_list(500)
        return {"data": docs}

    @router.post("/assignments/rm-tls")
    async def assign_rm_tl(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        if user["role"] not in ("owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        rmid = body.get("regional_manager_id"); tlid = body.get("team_leader_id")
        if not rmid or not tlid:
            raise HTTPException(status_code=400, detail="regional_manager_id + team_leader_id required")
        await db.dms_rm_assignments.update_one(
            {"regional_manager_id": rmid, "team_leader_id": tlid},
            {"$set": {"regional_manager_id": rmid, "team_leader_id": tlid, "assigned_by": user["id"], "at": _now()}},
            upsert=True,
        )
        return {"ok": True}

    @router.delete("/assignments/rm-tls")
    async def unassign_rm_tl(regional_manager_id: str, team_leader_id: str, user: dict = Depends(get_current_user)):
        if user["role"] not in ("owner", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        await db.dms_rm_assignments.delete_one({"regional_manager_id": regional_manager_id, "team_leader_id": team_leader_id})
        return {"ok": True}

    # ── list users by role (used by assignment UIs) ──
    @router.get("/users")
    async def list_dms_users(role: Optional[str] = None, user: dict = Depends(get_current_user)):
        if user["role"] not in ("owner", "super_admin", "team_leader", "regional_manager"):
            raise HTTPException(status_code=403, detail="Forbidden")
        q: Dict[str, Any] = {}
        if role:
            q["role"] = role
        docs = await db.users.find(q, {"_id": 0, "password_hash": 0}).to_list(1000)
        return {"data": docs}

    # ── punch in / out ──
    @router.post("/punch/in")
    async def punch_in(body: Dict[str, Any] = Body(...), user: dict = Depends(field_user_only)):
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        existing = await db.dms_punch.find_one({"salesperson_id": user["id"], "date": today, "out_at": None})
        if existing:
            return {"ok": True, "already": True, "punch": _clean(existing)}
        # If already punched out today, block re-punch unless Owner has granted a reopen
        closed = await db.dms_punch.find_one({"salesperson_id": user["id"], "date": today, "out_at": {"$ne": None}})
        if closed:
            grant = await db.dms_punch_reopen.find_one({"salesperson_id": user["id"], "date": today, "consumed": False})
            if not grant:
                raise HTTPException(status_code=400, detail="You have already punched out today. Please ask the Owner to allow Punch In again.")
            await db.dms_punch_reopen.update_one({"id": grant["id"]}, {"$set": {"consumed": True, "consumed_at": _now()}})
        doc = {
            "id": _nid("pn"), "salesperson_id": user["id"], "date": today,
            "in_at": _now(), "out_at": None,
            "gps_in": {"lat": body.get("lat"), "lng": body.get("lng")},
            "gps_out": None,
        }
        await db.dms_punch.insert_one(doc)
        return {"ok": True, "punch": _clean(doc)}

    @router.post("/punch/out")
    async def punch_out(body: Dict[str, Any] = Body(...), user: dict = Depends(field_user_only)):
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        p = await db.dms_punch.find_one({"salesperson_id": user["id"], "date": today, "out_at": None})
        if not p:
            raise HTTPException(status_code=400, detail="Not punched in today")
        await db.dms_punch.update_one({"id": p["id"]}, {"$set": {
            "out_at": _now(),
            "gps_out": {"lat": body.get("lat"), "lng": body.get("lng")},
        }})
        p2 = await db.dms_punch.find_one({"id": p["id"]}, {"_id": 0})
        return {"ok": True, "punch": p2}

    @router.get("/punch/today")
    async def punch_today(user: dict = Depends(get_current_user)):
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        punches = await db.dms_punch.find({"salesperson_id": user["id"], "date": today}, {"_id": 0}).sort("in_at", -1).to_list(50)
        latest = punches[0] if punches else None
        open_exists = any(not p.get("out_at") for p in punches)
        grant = await db.dms_punch_reopen.find_one({"salesperson_id": user["id"], "date": today, "consumed": False})
        can_punch_in = (len(punches) == 0) or (not open_exists and grant is not None)
        return {
            "punch": latest,
            "punched_in": open_exists,
            "can_punch_in": can_punch_in,
            "reopen_granted": grant is not None,
        }

    @router.get("/punch/history")
    async def punch_history(salesperson_id: Optional[str] = None, user: dict = Depends(get_current_user)):
        target = salesperson_id
        if not target and user["role"] == "salesperson":
            target = user["id"]
        docs = await db.dms_punch.find({"salesperson_id": target}, {"_id": 0}).sort("in_at", -1).to_list(60)
        return {"data": docs}

    # ── Owner: allow Punch-In again for a salesperson (no request workflow) ──
    @router.post("/owner/punch/reopen/{sp_id}")
    async def owner_reopen_punch(sp_id: str, user: dict = Depends(owner_only)):
        sp = await db.users.find_one({"id": sp_id, "role": "salesperson"}, {"_id": 0, "password_hash": 0})
        if not sp:
            raise HTTPException(status_code=404, detail="Salesperson not found")
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        # remove any stale unconsumed grant then create a fresh one
        await db.dms_punch_reopen.delete_many({"salesperson_id": sp_id, "date": today, "consumed": False})
        doc = {
            "id": _nid("rop"), "salesperson_id": sp_id, "date": today,
            "granted_by": user["id"], "granted_at": _now(), "consumed": False,
        }
        await db.dms_punch_reopen.insert_one(doc)
        await notify(sp_id, "punch_reopened", "Punch In re-enabled",
                     "The Owner has allowed you to Punch In again for today.", "/dms")
        return {"ok": True}

    # ── Unified Attendance (role-aware) ──
    async def _attendance_rows(sp_ids: List[str], days: int = 30):
        """Build attendance rows for given user ids (salespersons/TLs) over last N days."""
        rows = []
        if not sp_ids:
            return rows
        umap = {}
        async for u in db.users.find({"id": {"$in": sp_ids}}, {"_id": 0, "password_hash": 0}):
            umap[u["id"]] = u
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        punches = await db.dms_punch.find({"salesperson_id": {"$in": sp_ids}}, {"_id": 0}).sort("in_at", -1).to_list(3000)
        for p in punches:
            u = umap.get(p["salesperson_id"], {})
            grant = None
            if p["date"] == today and p.get("out_at"):
                grant = await db.dms_punch_reopen.find_one({"salesperson_id": p["salesperson_id"], "date": today, "consumed": False})
            rows.append({
                "punch_id": p.get("id"),
                "user_id": p["salesperson_id"],
                "name": u.get("name", "—"),
                "role": u.get("role"),
                "date": p.get("date"),
                "in_at": p.get("in_at"),
                "out_at": p.get("out_at"),
                "gps_in": p.get("gps_in"),
                "gps_out": p.get("gps_out"),
                "is_today": p["date"] == today,
                "reopen_granted": grant is not None,
                "can_reopen": (p["date"] == today and bool(p.get("out_at")) and grant is None),
            })
        return rows

    @router.get("/attendance")
    async def attendance(user: dict = Depends(get_current_user)):
        role = user.get("role")
        ids: List[str] = []
        if role == "salesperson":
            ids = [user["id"]]
        elif role == "team_leader":
            dids = [a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0, "distributor_id": 1})]
            sp_ids = list({a["salesperson_id"] async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "salesperson_id": 1})})
            ids = [user["id"]] + sp_ids
        elif role == "regional_manager":
            tlids = [a["team_leader_id"] async for a in db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0, "team_leader_id": 1})]
            dids = list({a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "distributor_id": 1})})
            sp_ids = list({a["salesperson_id"] async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "salesperson_id": 1})})
            ids = [user["id"]] + tlids + sp_ids
        elif role in ("owner", "super_admin", "owner_accountant"):
            # Owner sees ALL field staff attendance (everyone except owner logins)
            ids = [u["id"] async for u in db.users.find(
                {"role": {"$nin": ["owner", "super_admin", "owner_accountant"]}},
                {"_id": 0, "id": 1})]
        else:
            # distributor / retailer / distributor_accountant → own history only
            ids = [user["id"]]
        rows = await _attendance_rows(ids)
        return {"data": rows}

    # =========================================================================
    # DASHBOARDS — sales team roles
    # =========================================================================
    @router.get("/dashboard/salesperson")
    async def sp_dashboard(user: dict = Depends(salesperson_only)):
        assigns = await db.dms_sp_assignments.find({"salesperson_id": user["id"]}, {"_id": 0}).to_list(500)
        dids = [a["distributor_id"] for a in assigns]
        n_dists = len(dids)
        n_retailers = await db.dms_retailers.count_documents({"distributor_id": {"$in": dids}, "active": True}) if dids else 0
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        orders_today = await db.dms_secondary_orders.count_documents({
            "placed_by": user["id"],
            "created_at": {"$gte": today + "T00:00:00"},
        })
        punch = await db.dms_punch.find_one({"salesperson_id": user["id"], "date": today}, {"_id": 0})
        return {
            "kpis": {
                "assigned_distributors": n_dists,
                "assigned_retailers": n_retailers,
                "orders_today": orders_today,
                "punched_in": bool(punch and not punch.get("out_at")),
            },
            "today_punch": punch,
        }

    @router.get("/dashboard/team-leader")
    async def tl_dashboard(user: dict = Depends(team_leader_only)):
        """
        Strict-spec TL dashboard. Returns:
          today_sales, monthly_sales, total_orders, pending_orders, fulfillment_pct,
          assigned_distributors, assigned_salespersons, total_retailers, stock_alerts
        """
        my_dists = await db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0}).to_list(500)
        dids = [a["distributor_id"] for a in my_dists]

        today = _today()
        mtd_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

        today_sales = 0.0; monthly_sales = 0.0
        total_orders = 0; pending_orders = 0; delivered_orders = 0
        async for so in db.dms_secondary_orders.find({"distributor_id": {"$in": dids}}, {"_id": 0}):
            total_orders += 1
            if so.get("created_at", "").startswith(today):
                today_sales += so.get("total", 0)
            if so.get("created_at", "") >= mtd_start:
                monthly_sales += so.get("total", 0)
            st = so.get("status")
            if st == "pending":
                pending_orders += 1
            elif st in ("dispatched", "delivered"):
                delivered_orders += 1
        fulfillment_pct = _round((delivered_orders / total_orders * 100) if total_orders else 0, 1)

        n_sp = await db.dms_sp_assignments.count_documents({"distributor_id": {"$in": dids}}) if dids else 0
        n_ret = await db.dms_retailers.count_documents({"distributor_id": {"$in": dids}, "active": True}) if dids else 0

        # Stock alerts — distributor-wise low stock (<5 boxes) — count only
        stock_alerts = 0
        async for inv in db.dms_distributor_inventory.find({"distributor_id": {"$in": dids}}, {"_id": 0, "qty_boxes": 1}):
            if inv.get("qty_boxes", 0) < 5:
                stock_alerts += 1

        return {
            "kpis": {
                "today_sales": _round(today_sales),
                "monthly_sales": _round(monthly_sales),
                "total_orders": total_orders,
                "pending_orders": pending_orders,
                "fulfillment_pct": fulfillment_pct,
                "assigned_distributors": len(dids),
                "assigned_salespersons": n_sp,
                "total_retailers": n_ret,
                "stock_alerts": stock_alerts,
            }
        }

    # -------- TL: Distributor performance ---------
    @router.get("/tl/distributors")
    async def tl_distributor_performance(user: dict = Depends(team_leader_only)):
        assigns = await db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0}).to_list(500)
        dids = [a["distributor_id"] for a in assigns]
        today = _today()
        mtd_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        out = []
        for did in dids:
            d = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
            if not d: continue
            stock = 0
            async for inv in db.dms_distributor_inventory.find({"distributor_id": did}, {"_id": 0, "qty_boxes": 1}):
                stock += inv.get("qty_boxes", 0)
            # Outstanding payable = primary (owner→dist) invoices - payments
            po_billed = 0.0; po_paid = 0.0
            async for e in db.dms_primary_ledger.find({"distributor_id": did}, {"_id": 0}):
                if e.get("kind") == "invoice": po_billed += e.get("amount", 0)
                elif e.get("kind") in ("payment", "coupon_credit"): po_paid += e.get("amount", 0)
            outstanding_payable = po_billed - po_paid
            # Outstanding receivable = secondary (dist→retailer) invoices - payments
            so_billed = 0.0; so_paid = 0.0
            async for e in db.dms_retailer_ledger.find({"distributor_id": did}, {"_id": 0}):
                if e.get("kind") == "invoice": so_billed += e.get("amount", 0)
                elif e.get("kind") in ("payment", "coupon_credit"): so_paid += e.get("amount", 0)
            outstanding_receivable = so_billed - so_paid
            today_s = 0.0; monthly_s = 0.0; pending = 0
            async for so in db.dms_secondary_orders.find({"distributor_id": did}, {"_id": 0}):
                if so.get("created_at", "").startswith(today): today_s += so.get("total", 0)
                if so.get("created_at", "") >= mtd_start: monthly_s += so.get("total", 0)
                if so.get("status") == "pending": pending += 1
            out.append({
                "id": did, "name": d.get("name"), "region": d.get("region"),
                "available_stock": stock,
                "outstanding_payable_to_owner": _round(outstanding_payable),
                "outstanding_receivable_from_retailers": _round(outstanding_receivable),
                "today_sales": _round(today_s),
                "monthly_sales": _round(monthly_s),
                "revenue": _round(so_billed),
                "pending_orders": pending,
            })
        return {"data": out}

    # -------- TL: Salesperson list with live status ---------
    @router.get("/tl/salespersons")
    async def tl_salespersons(user: dict = Depends(team_leader_only)):
        assigns = await db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0}).to_list(500)
        dids = [a["distributor_id"] for a in assigns]
        sp_ids = set()
        async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0}):
            sp_ids.add(a["salesperson_id"])
        today = _today()
        out = []
        async for u in db.users.find({"role": "salesperson", "id": {"$in": list(sp_ids)}}, {"_id": 0, "password_hash": 0}):
            last_at = (u.get("last_gps") or {}).get("at") or u.get("last_active_at")
            online = False
            try:
                if last_at:
                    dt = datetime.fromisoformat(last_at.replace("Z", "+00:00"))
                    online = (datetime.now(timezone.utc) - dt).total_seconds() < 300
            except Exception:
                pass
            punch = await db.dms_punch.find_one({"salesperson_id": u["id"], "date": today}, {"_id": 0})
            # today's orders + new retailers
            orders_today = await db.dms_secondary_orders.count_documents({"placed_by": u["id"], "created_at": {"$gte": today + "T00:00:00"}})
            new_ret_today = await db.dms_retailers.count_documents({"onboarded_by": u["id"], "created_at": {"$gte": today + "T00:00:00"}})
            # unique retailer visits today (from proximity to any retailer)
            pings = await db.dms_sp_pings.find({"salesperson_id": u["id"], "date": today}, {"_id": 0}).to_list(2000)
            visits = 0
            if pings:
                async for r in db.dms_retailers.find({"distributor_id": {"$in": dids}, "gps_lat": {"$ne": None}}, {"_id": 0}):
                    for p in pings:
                        if _haversine_km(r["gps_lat"], r["gps_lng"], p["lat"], p["lng"]) < 0.20:
                            visits += 1
                            break
            out.append({
                "id": u["id"], "name": u.get("name"), "phone": u.get("phone"),
                "online": online, "last_seen_at": last_at,
                "live_location": u.get("last_gps"),
                "punch_in": (punch or {}).get("in_at"),
                "punch_out": (punch or {}).get("out_at"),
                "today_visits": visits,
                "orders_today": orders_today,
                "new_retailers_today": new_ret_today,
            })
        return {"data": out}

    # -------- TL: Order monitoring ---------
    @router.get("/tl/orders")
    async def tl_orders(
        status: Optional[str] = None,
        distributor_id: Optional[str] = None,
        salesperson_id: Optional[str] = None,
        retailer_id: Optional[str] = None,
        user: dict = Depends(team_leader_only),
    ):
        assigns = await db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0}).to_list(500)
        my_dids = [a["distributor_id"] for a in assigns]
        q: Dict[str, Any] = {"distributor_id": {"$in": my_dids}}
        if status: q["status"] = status
        if distributor_id and distributor_id in my_dids: q["distributor_id"] = distributor_id
        if salesperson_id: q["placed_by"] = salesperson_id
        if retailer_id: q["retailer_id"] = retailer_id
        docs = await db.dms_secondary_orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
        # enrich
        for o in docs:
            r = await db.dms_retailers.find_one({"id": o.get("retailer_id")}, {"_id": 0, "name": 1})
            d = await db.dms_distributors.find_one({"id": o.get("distributor_id")}, {"_id": 0, "name": 1})
            o["retailer_name"] = (r or {}).get("name")
            o["distributor_name"] = (d or {}).get("name")
            # "Placed By" — salesperson (or whoever placed the order)
            if o.get("placed_by"):
                pb = await db.users.find_one({"id": o.get("placed_by")}, {"_id": 0, "name": 1, "role": 1})
                o["placed_by_name"] = (pb or {}).get("name")
                o["placed_by_role"] = (pb or {}).get("role")
            else:
                o["placed_by_name"] = o.get("placed_by_name")
        return {"data": docs, "count": len(docs)}

    # -------- TL: Retailers ---------
    @router.get("/tl/retailers")
    async def tl_retailers(user: dict = Depends(team_leader_only)):
        assigns = await db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0}).to_list(500)
        dids = [a["distributor_id"] for a in assigns]
        out = []
        async for r in db.dms_retailers.find({"distributor_id": {"$in": dids}, "active": True}, {"_id": 0}):
            # outstanding
            billed = 0.0; paid = 0.0
            async for e in db.dms_retailer_ledger.find({"retailer_id": r["id"]}, {"_id": 0}):
                if e.get("kind") == "invoice": billed += e.get("amount", 0)
                elif e.get("kind") in ("payment", "coupon_credit"): paid += e.get("amount", 0)
            # last order
            last = await db.dms_secondary_orders.find_one({"retailer_id": r["id"]}, {"_id": 0}, sort=[("created_at", -1)])
            total_purchases = 0.0
            async for so in db.dms_secondary_orders.find({"retailer_id": r["id"]}, {"_id": 0, "total": 1}):
                total_purchases += so.get("total", 0)
            out.append({
                "id": r["id"], "name": r.get("name"), "phone": r.get("phone"),
                "address": r.get("address"), "region": r.get("region"),
                "gps_lat": r.get("gps_lat"), "gps_lng": r.get("gps_lng"),
                "location_link": r.get("location_link"),
                "outstanding": _round(billed - paid),
                "last_order_at": (last or {}).get("created_at"),
                "total_purchases": _round(total_purchases),
            })
        return {"data": out}

    # -------- TL: Attendance (own) ---------
    @router.get("/tl/attendance")
    async def tl_attendance(user: dict = Depends(team_leader_only)):
        docs = await db.dms_punch.find({"salesperson_id": user["id"]}, {"_id": 0}).sort("in_at", -1).to_list(60)
        return {"data": docs}

    @router.post("/tl/punch/in")
    async def tl_punch_in(body: Dict[str, Any] = Body(...), user: dict = Depends(team_leader_only)):
        today = _today()
        existing = await db.dms_punch.find_one({"salesperson_id": user["id"], "date": today, "out_at": None})
        if existing:
            return {"ok": True, "already": True}
        doc = {"id": _nid("pn"), "salesperson_id": user["id"], "date": today,
               "in_at": _now(), "out_at": None,
               "gps_in": {"lat": body.get("lat"), "lng": body.get("lng")}, "gps_out": None}
        await db.dms_punch.insert_one(doc)
        return {"ok": True, "punch": _clean(doc)}

    @router.post("/tl/punch/out")
    async def tl_punch_out(body: Dict[str, Any] = Body(...), user: dict = Depends(team_leader_only)):
        today = _today()
        p = await db.dms_punch.find_one({"salesperson_id": user["id"], "date": today, "out_at": None})
        if not p:
            raise HTTPException(status_code=400, detail="Not punched in today")
        await db.dms_punch.update_one({"id": p["id"]}, {"$set": {
            "out_at": _now(), "gps_out": {"lat": body.get("lat"), "lng": body.get("lng")},
        }})
        return {"ok": True}

    # =========================================================================
    # OWNER — TL Performance Dashboard + Distributor Sales Drilldown
    # =========================================================================
    @router.get("/owner/tl-performance")
    async def owner_tl_performance(user: dict = Depends(owner_only)):
        today = _today()
        mtd_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        rows = []
        async for tl in db.users.find({"role": "team_leader"}, {"_id": 0, "password_hash": 0}):
            assigns = await db.dms_tl_assignments.find({"team_leader_id": tl["id"]}, {"_id": 0}).to_list(500)
            dids = [a["distributor_id"] for a in assigns]
            n_sp = await db.dms_sp_assignments.count_documents({"distributor_id": {"$in": dids}}) if dids else 0
            total = 0.0; today_s = 0.0; monthly = 0.0; n_orders = 0; pending = 0
            # date-wise last 7 days
            by_date: Dict[str, float] = {}
            async for so in db.dms_secondary_orders.find({"distributor_id": {"$in": dids}}, {"_id": 0}):
                amt = so.get("total", 0)
                total += amt
                n_orders += 1
                d = so.get("created_at", "")[:10]
                by_date[d] = by_date.get(d, 0) + amt
                if d == today: today_s += amt
                if so.get("created_at", "") >= mtd_start: monthly += amt
                if so.get("status") == "pending": pending += 1
            # last 7 days series
            series = []
            for i in range(6, -1, -1):
                d = (datetime.now(timezone.utc) - timedelta(days=i)).strftime("%Y-%m-%d")
                series.append({"date": d, "sales": _round(by_date.get(d, 0))})
            rows.append({
                "team_leader_id": tl["id"],
                "name": tl.get("name"),
                "email": tl.get("email"),
                "assigned_distributors": len(dids),
                "assigned_salespersons": n_sp,
                "total_sales": _round(total),
                "today_sales": _round(today_s),
                "monthly_sales": _round(monthly),
                "total_orders": n_orders,
                "pending_orders": pending,
                "series_7d": series,
            })
        # ranking by total_sales desc
        rows.sort(key=lambda r: r["total_sales"], reverse=True)
        return {"data": rows}

    @router.get("/owner/distributor-sales/{did}")
    async def owner_distributor_sales(did: str, user: dict = Depends(owner_only)):
        """Complete drilldown: which retailers, which products, quantities, prices."""
        dist = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
        if not dist:
            raise HTTPException(status_code=404, detail="Distributor not found")
        # all secondary orders for this distributor
        orders = await db.dms_secondary_orders.find({"distributor_id": did}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        # aggregate by retailer & by product
        by_retailer: Dict[str, Dict[str, Any]] = {}
        by_product: Dict[str, Dict[str, Any]] = {}
        for o in orders:
            rid = o.get("retailer_id")
            if rid not in by_retailer:
                r = await db.dms_retailers.find_one({"id": rid}, {"_id": 0, "name": 1, "address": 1})
                by_retailer[rid] = {"retailer_id": rid, "retailer_name": (r or {}).get("name"), "address": (r or {}).get("address"),
                                    "orders": 0, "total_revenue": 0.0}
            by_retailer[rid]["orders"] += 1
            by_retailer[rid]["total_revenue"] += o.get("total", 0)
            for it in (o.get("items") or []):
                pid = it.get("product_id")
                if pid not in by_product:
                    p = await db.dms_products.find_one({"id": pid}, {"_id": 0, "name": 1, "sku_code": 1})
                    by_product[pid] = {"product_id": pid, "product_name": (p or {}).get("name"),
                                       "sku_code": (p or {}).get("sku_code"),
                                       "qty_boxes": 0, "qty_pcs": 0, "revenue": 0.0,
                                       "prices_seen": set()}
                by_product[pid]["qty_boxes"] += int(it.get("qty_boxes", 0) or 0)
                by_product[pid]["qty_pcs"] += int(it.get("qty_pcs", 0) or 0)
                by_product[pid]["revenue"] += float(it.get("line_total", 0) or 0)
                if it.get("box_price"): by_product[pid]["prices_seen"].add(float(it["box_price"]))
        # normalise prices set → list
        for p in by_product.values():
            p["prices_seen"] = sorted(list(p["prices_seen"]))
            p["revenue"] = _round(p["revenue"])
        for r in by_retailer.values():
            r["total_revenue"] = _round(r["total_revenue"])
        # slim orders (last 30)
        recent = []
        for o in orders[:30]:
            r = await db.dms_retailers.find_one({"id": o.get("retailer_id")}, {"_id": 0, "name": 1})
            recent.append({
                "id": o["id"], "at": o.get("created_at"),
                "retailer": (r or {}).get("name"),
                "total": _round(o.get("total", 0)),
                "status": o.get("status"),
                "items_count": len(o.get("items") or []),
            })
        return {
            "distributor": {"id": dist["id"], "name": dist.get("name"), "region": dist.get("region"),
                             "gps_lat": dist.get("gps_lat"), "gps_lng": dist.get("gps_lng")},
            "by_retailer": list(by_retailer.values()),
            "by_product": list(by_product.values()),
            "recent_orders": recent,
            "totals": {
                "orders": len(orders),
                "revenue": _round(sum(o.get("total", 0) for o in orders)),
                "retailers_active": len(by_retailer),
                "products_sold": len(by_product),
            },
        }

    @router.get("/dashboard/team-leader-legacy-sales-mtd")
    async def _tl_sales_mtd_legacy(user: dict = Depends(team_leader_only)):
        # kept only so any old client that still calls this doesn't 404
        return {"kpis": {"sales_mtd": 0}}


    @router.get("/dashboard/regional-manager")
    async def rm_dashboard(user: dict = Depends(regional_manager_only)):
        """
        Strict-spec RM dashboard: total_tls, total_distributors, total_retailers,
        total_salespersons, today_sales, monthly_sales, outstanding, revenue, fulfillment_pct.
        """
        my_tls = await db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0}).to_list(500)
        tlids = [a["team_leader_id"] for a in my_tls]
        # distributors under those TLs
        dids = list({a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "distributor_id": 1})})
        sp_ids = list({a["salesperson_id"] async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "salesperson_id": 1})})
        n_ret = await db.dms_retailers.count_documents({"distributor_id": {"$in": dids}, "active": True}) if dids else 0
        today = _today()
        mtd_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        today_s = 0.0; monthly = 0.0; total_rev = 0.0; total_n = 0; delivered = 0
        async for so in db.dms_secondary_orders.find({"distributor_id": {"$in": dids}}, {"_id": 0}):
            total_n += 1
            total_rev += so.get("total", 0)
            if so.get("created_at", "").startswith(today): today_s += so.get("total", 0)
            if so.get("created_at", "") >= mtd_start: monthly += so.get("total", 0)
            if so.get("status") in ("dispatched", "delivered"): delivered += 1
        # outstanding (retailer ledger) across those distributors
        billed = 0.0; paid = 0.0
        async for e in db.dms_retailer_ledger.find({"distributor_id": {"$in": dids}}, {"_id": 0}):
            if e.get("kind") == "invoice": billed += e.get("amount", 0)
            elif e.get("kind") in ("payment", "coupon_credit"): paid += e.get("amount", 0)
        return {
            "kpis": {
                "team_leaders": len(tlids),
                "distributors": len(dids),
                "retailers": n_ret,
                "salespersons": len(sp_ids),
                "today_sales": _round(today_s),
                "monthly_sales": _round(monthly),
                "outstanding": _round(billed - paid),
                "revenue": _round(total_rev),
                "fulfillment_pct": _round((delivered / total_n * 100) if total_n else 0, 1),
            }
        }

    # --------- RM: TL monitoring ---------
    @router.get("/rm/team-leaders")
    async def rm_team_leaders(user: dict = Depends(regional_manager_only)):
        my_tls = await db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0}).to_list(500)
        tlids = [a["team_leader_id"] for a in my_tls]
        rows = []
        for tlid in tlids:
            tl = await db.users.find_one({"id": tlid}, {"_id": 0, "password_hash": 0})
            if not tl: continue
            dids = [a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": tlid}, {"_id": 0, "distributor_id": 1})]
            active_sp = 0
            for spid in [a["salesperson_id"] async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "salesperson_id": 1})]:
                u = await db.users.find_one({"id": spid}, {"_id": 0, "last_active_at": 1})
                la = (u or {}).get("last_active_at")
                if la:
                    try:
                        dt = datetime.fromisoformat(la.replace("Z", "+00:00"))
                        if (datetime.now(timezone.utc) - dt).total_seconds() < 300: active_sp += 1
                    except Exception: pass
            sales = 0.0; pending = 0
            async for so in db.dms_secondary_orders.find({"distributor_id": {"$in": dids}}, {"_id": 0}):
                sales += so.get("total", 0)
                if so.get("status") == "pending": pending += 1
            rows.append({
                "id": tlid, "name": tl.get("name"), "email": tl.get("email"),
                "sales": _round(sales),
                "active_salespersons": active_sp,
                "active_distributors": len(dids),
                "pending_orders": pending,
                "revenue": _round(sales),
            })
        rows.sort(key=lambda r: r["sales"], reverse=True)
        return {"data": rows}

    # --------- RM: Distributor monitoring (read only) ---------
    @router.get("/rm/distributors")
    async def rm_distributors(user: dict = Depends(regional_manager_only)):
        my_tls = await db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0}).to_list(500)
        tlids = [a["team_leader_id"] for a in my_tls]
        dids = list({a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "distributor_id": 1})})
        rows = []
        for did in dids:
            d = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
            if not d: continue
            stock = 0
            async for inv in db.dms_distributor_inventory.find({"distributor_id": did}, {"_id": 0, "qty_boxes": 1}):
                stock += inv.get("qty_boxes", 0)
            po_billed = 0.0; po_paid = 0.0
            async for e in db.dms_primary_ledger.find({"distributor_id": did}, {"_id": 0}):
                if e.get("kind") == "invoice": po_billed += e.get("amount", 0)
                elif e.get("kind") in ("payment", "coupon_credit"): po_paid += e.get("amount", 0)
            revenue = 0.0; orders = 0
            async for so in db.dms_secondary_orders.find({"distributor_id": did}, {"_id": 0}):
                revenue += so.get("total", 0); orders += 1
            n_ret = await db.dms_retailers.count_documents({"distributor_id": did, "active": True})
            rows.append({
                "id": did, "name": d.get("name"), "region": d.get("region"),
                "stock": stock, "pending_payments": _round(po_billed - po_paid),
                "revenue": _round(revenue), "retailers": n_ret, "orders": orders,
            })
        return {"data": rows}

    # --------- RM: Salesperson monitoring (read only) ---------
    @router.get("/rm/salespersons")
    async def rm_salespersons(user: dict = Depends(regional_manager_only)):
        my_tls = await db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0}).to_list(500)
        tlids = [a["team_leader_id"] for a in my_tls]
        dids = list({a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "distributor_id": 1})})
        sp_ids = list({a["salesperson_id"] async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "salesperson_id": 1})})
        today = _today()
        rows = []
        async for u in db.users.find({"role": "salesperson", "id": {"$in": sp_ids}}, {"_id": 0, "password_hash": 0}):
            la = (u.get("last_gps") or {}).get("at") or u.get("last_active_at")
            online = False
            try:
                if la:
                    dt = datetime.fromisoformat(la.replace("Z", "+00:00"))
                    online = (datetime.now(timezone.utc) - dt).total_seconds() < 300
            except Exception: pass
            punch = await db.dms_punch.find_one({"salesperson_id": u["id"], "date": today}, {"_id": 0})
            visits = 0
            pings = await db.dms_sp_pings.find({"salesperson_id": u["id"], "date": today}, {"_id": 0}).to_list(2000)
            if pings:
                async for r in db.dms_retailers.find({"distributor_id": {"$in": dids}, "gps_lat": {"$ne": None}}, {"_id": 0}):
                    for p in pings:
                        if _haversine_km(r["gps_lat"], r["gps_lng"], p["lat"], p["lng"]) < 0.20:
                            visits += 1; break
            orders_today = await db.dms_secondary_orders.count_documents({"placed_by": u["id"], "created_at": {"$gte": today + "T00:00:00"}})
            new_ret = await db.dms_retailers.count_documents({"onboarded_by": u["id"], "created_at": {"$gte": today + "T00:00:00"}})
            rows.append({
                "id": u["id"], "name": u.get("name"), "phone": u.get("phone"),
                "online": online, "live_location": u.get("last_gps"),
                "punch_in": (punch or {}).get("in_at"), "punch_out": (punch or {}).get("out_at"),
                "orders_today": orders_today, "today_visits": visits, "new_retailers_today": new_ret,
            })
        return {"data": rows}

    # --------- RM: Region performance (dist-wise, TL-wise, SP-wise sales) ---------
    @router.get("/rm/region-performance")
    async def rm_region_performance(user: dict = Depends(regional_manager_only)):
        my_tls = await db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0}).to_list(500)
        tlids = [a["team_leader_id"] for a in my_tls]
        tl_dists = {a["team_leader_id"]: [] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0})}
        # accumulate
        async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0}):
            tl_dists.setdefault(a["team_leader_id"], []).append(a["distributor_id"])
        all_dids = list({d for lst in tl_dists.values() for d in lst})

        by_dist = {}; by_tl = {}; by_sp = {}
        async for so in db.dms_secondary_orders.find({"distributor_id": {"$in": all_dids}}, {"_id": 0}):
            amt = so.get("total", 0)
            by_dist[so["distributor_id"]] = by_dist.get(so["distributor_id"], 0) + amt
            placed_by = so.get("placed_by")
            if placed_by: by_sp[placed_by] = by_sp.get(placed_by, 0) + amt

        # attribute to TLs via TL→dist mapping
        for tlid, dlist in tl_dists.items():
            by_tl[tlid] = sum(by_dist.get(d, 0) for d in dlist)

        # enrich with names
        async def _name(uid):
            u = await db.users.find_one({"id": uid}, {"_id": 0, "name": 1})
            return (u or {}).get("name", uid)
        async def _dname(did):
            d = await db.dms_distributors.find_one({"id": did}, {"_id": 0, "name": 1})
            return (d or {}).get("name", did)

        out_dist = [{"id": did, "name": await _dname(did), "sales": _round(v)} for did, v in by_dist.items()]
        out_tl = [{"id": tid, "name": await _name(tid), "sales": _round(v)} for tid, v in by_tl.items()]
        out_sp = [{"id": sid, "name": await _name(sid), "sales": _round(v)} for sid, v in by_sp.items()]
        for lst in (out_dist, out_tl, out_sp): lst.sort(key=lambda r: r["sales"], reverse=True)
        return {"by_distributor": out_dist, "by_team_leader": out_tl, "by_salesperson": out_sp}

    # --------- RM: My Retailers (all retailers under RM's TLs & SPs) ---------
    @router.get("/rm/retailers")
    async def rm_retailers(user: dict = Depends(regional_manager_only)):
        tlids = [a["team_leader_id"] async for a in db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0, "team_leader_id": 1})]
        dids = list({a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "distributor_id": 1})})
        dist_map = {}
        async for d in db.dms_distributors.find({"id": {"$in": dids}}, {"_id": 0, "id": 1, "name": 1}):
            dist_map[d["id"]] = d.get("name")
        rows = []
        async for r in db.dms_retailers.find({"distributor_id": {"$in": dids}, "active": True}, {"_id": 0}):
            billed = 0.0; paid = 0.0
            async for e in db.dms_retailer_ledger.find({"retailer_id": r["id"]}, {"_id": 0}):
                if e.get("kind") == "invoice": billed += e.get("amount", 0)
                elif e.get("kind") in ("payment", "coupon_credit"): paid += e.get("amount", 0)
            last = await db.dms_secondary_orders.find_one({"retailer_id": r["id"]}, {"_id": 0}, sort=[("created_at", -1)])
            onboard = await db.users.find_one({"id": r.get("onboarded_by")}, {"_id": 0, "name": 1}) if r.get("onboarded_by") else None
            rows.append({
                "id": r["id"], "name": r.get("name"), "phone": r.get("phone"),
                "address": r.get("address"), "region": r.get("region"),
                "distributor_id": r.get("distributor_id"),
                "distributor_name": dist_map.get(r.get("distributor_id")),
                "onboarded_by_name": (onboard or {}).get("name"),
                "gps_lat": r.get("gps_lat"), "gps_lng": r.get("gps_lng"),
                "outstanding": _round(billed - paid),
                "last_order_at": (last or {}).get("created_at"),
            })
        rows.sort(key=lambda x: (x.get("distributor_name") or "", x.get("name") or ""))
        return {"data": rows}


    @router.get("/dashboard/retailer")
    async def retailer_dashboard(user: dict = Depends(retailer_only)):
        rid = user.get("retailer_id")
        if not rid:
            # Retailer with no linked profile yet → clean empty dashboard
            return {"kpis": {"total_orders": 0, "in_transit": 0,
                             "outstanding": 0, "pending_items": 0}}
        billed = 0.0; paid = 0.0
        async for e in db.dms_retailer_ledger.find({"retailer_id": rid}, {"_id": 0}):
            if e["kind"] == "invoice":
                billed += e["amount"]
            elif e["kind"] in ("payment", "coupon_credit"):
                paid += e["amount"]
        pending = 0
        async for pd in db.dms_retailer_pending.find({"retailer_id": rid}, {"_id": 0}):
            pending += int(pd.get("pending_qty_boxes", 0)) + int(pd.get("pending_qty_pcs", 0))
        n_orders = await db.dms_secondary_orders.count_documents({"retailer_id": rid})
        n_dispatched = await db.dms_secondary_orders.count_documents({"retailer_id": rid, "status": "dispatched"})
        return {
            "kpis": {
                "total_orders": n_orders,
                "in_transit": n_dispatched,
                "outstanding": _round(billed - paid),
                "pending_items": pending,
            }
        }

    @router.get("/dashboard/super-admin")
    async def superadmin_dashboard(user: dict = Depends(_guard())):
        # Owner + super_admin both allowed (roles unified per user request)
        if user["role"] not in ("super_admin", "owner"):
            raise HTTPException(status_code=403, detail="Owner only")
        n_owners = await db.users.count_documents({"role": "owner"})
        n_tl = await db.users.count_documents({"role": "team_leader"})
        n_sp = await db.users.count_documents({"role": "salesperson"})
        n_dist = await db.dms_distributors.count_documents({"active": True})
        n_ret = await db.dms_retailers.count_documents({"active": True})
        n_po = await db.dms_primary_orders.count_documents({})
        n_so = await db.dms_secondary_orders.count_documents({})
        return {
            "kpis": {
                "owners": n_owners, "team_leaders": n_tl, "salespersons": n_sp,
                "distributors": n_dist, "retailers": n_ret,
                "primary_orders": n_po, "secondary_orders": n_so,
            }
        }

    # =========================================================================
    # LIVE TRACKING — Salesperson GPS pings (Phase 2 + 3)
    # =========================================================================
    def _haversine_km(lat1, lng1, lat2, lng2) -> float:
        """Great-circle distance between two lat/lng points in KM."""
        from math import radians, sin, cos, asin, sqrt
        try:
            lat1, lng1, lat2, lng2 = map(float, (lat1, lng1, lat2, lng2))
        except Exception:
            return 0.0
        R = 6371.0
        dLat = radians(lat2 - lat1)
        dLng = radians(lng2 - lng1)
        a = sin(dLat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dLng/2)**2
        return 2 * R * asin(sqrt(a))

    def _yyyy_mm_dd(iso_or_dt) -> str:
        if isinstance(iso_or_dt, str):
            return iso_or_dt[:10]
        return iso_or_dt.strftime("%Y-%m-%d")

    def _today() -> str:
        return datetime.now(timezone.utc).strftime("%Y-%m-%d")

    def _tl_can_see_sp(tl_id: str, sp_user: dict, db_ref) -> bool:
        # placeholder — narrower check happens inline; kept for readability
        return True

    async def _sp_visible_ids_for(user: dict) -> List[str]:
        """
        Return the list of salesperson user_ids the given user is allowed to see.
        owner / super_admin → all
        regional_manager → SPs assigned to TLs under this RM
        team_leader → SPs assigned to distributors under this TL
        """
        role = user.get("role")
        if role in ("owner", "super_admin"):
            ids = [u["id"] async for u in db.users.find({"role": "salesperson"}, {"_id": 0, "id": 1})]
            return ids
        if role == "team_leader":
            tl_dists = [a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0, "distributor_id": 1})]
            if not tl_dists:
                return []
            ids = set()
            async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": tl_dists}}, {"_id": 0, "salesperson_id": 1}):
                ids.add(a["salesperson_id"])
            return list(ids)
        if role == "regional_manager":
            tls = [a["team_leader_id"] async for a in db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0, "team_leader_id": 1})]
            dists = set()
            async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tls}}, {"_id": 0, "distributor_id": 1}):
                dists.add(a["distributor_id"])
            ids = set()
            async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": list(dists)}}, {"_id": 0, "salesperson_id": 1}):
                ids.add(a["salesperson_id"])
            return list(ids)
        if role == "salesperson":
            return [user["id"]]
        return []

    @router.post("/tracking/ping")
    async def tracking_ping(body: Dict[str, Any] = Body(...), user: dict = Depends(field_user_only)):
        """Any field user (non-owner) posts current GPS while punched in."""
        lat = body.get("lat"); lng = body.get("lng")
        if lat is None or lng is None:
            raise HTTPException(status_code=400, detail="lat and lng required")
        now = _now()
        doc = {
            "id": _nid("png"),
            "salesperson_id": user["id"],
            "lat": float(lat),
            "lng": float(lng),
            "accuracy": body.get("accuracy"),
            "speed": body.get("speed"),
            "date": _today(),
            "created_at": now,
        }
        await db.dms_sp_pings.insert_one(doc)
        # also stamp "last_active_at" on the user so live status reflects
        await db.users.update_one({"id": user["id"]}, {"$set": {
            "last_active_at": now,
            "last_gps": {"lat": doc["lat"], "lng": doc["lng"], "at": now},
        }})
        return {"ok": True}

    @router.get("/tracking/live")
    async def tracking_live(user: dict = Depends(get_current_user)):
        """
        Current live map data for Owner / TL / RM.
        Returns:
          - salespersons: [{id, name, phone, lat, lng, last_ping_at, online}]
          - distributors: [{id, name, lat, lng, address}]
          - retailers:    [{id, name, lat, lng, address, distributor_id}]
        """
        role = user.get("role")
        if role not in ("owner", "super_admin", "team_leader", "regional_manager"):
            raise HTTPException(status_code=403, detail="Forbidden")

        sp_ids = await _sp_visible_ids_for(user)
        sps = []
        async for u in db.users.find({"role": "salesperson", "id": {"$in": sp_ids}}, {"_id": 0, "password_hash": 0}):
            gps = u.get("last_gps") or {}
            last_at = gps.get("at") or u.get("last_active_at")
            online = False
            try:
                if last_at:
                    dt = datetime.fromisoformat(last_at.replace("Z", "+00:00"))
                    online = (datetime.now(timezone.utc) - dt).total_seconds() < 300
            except Exception:
                pass
            sps.append({
                "id": u["id"], "name": u.get("name"), "phone": u.get("phone"),
                "lat": gps.get("lat"), "lng": gps.get("lng"),
                "last_ping_at": last_at,
                "online": online,
            })

        # distributors — filter by role hierarchy
        dq: Dict[str, Any] = {"active": True}
        if role == "team_leader":
            dids = [a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": user["id"]}, {"_id": 0, "distributor_id": 1})]
            dq["id"] = {"$in": dids}
        elif role == "regional_manager":
            tls = [a["team_leader_id"] async for a in db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0, "team_leader_id": 1})]
            dids = [a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tls}}, {"_id": 0, "distributor_id": 1})]
            dq["id"] = {"$in": dids}
        dists = []
        async for d in db.dms_distributors.find(dq, {"_id": 0}):
            dists.append({
                "id": d["id"], "name": d.get("name"),
                "lat": d.get("gps_lat"), "lng": d.get("gps_lng"),
                "location_link": d.get("location_link"),
                "address": d.get("address"), "region": d.get("region"),
            })

        # retailers — under those distributors
        rq: Dict[str, Any] = {"active": True}
        if "id" in dq:
            rq["distributor_id"] = dq["id"]
        rets = []
        async for r in db.dms_retailers.find(rq, {"_id": 0}):
            rets.append({
                "id": r["id"], "name": r.get("name"),
                "lat": r.get("gps_lat"), "lng": r.get("gps_lng"),
                "location_link": r.get("location_link"),
                "address": r.get("address"),
                "distributor_id": r.get("distributor_id"),
            })

        # Phase 1: Team Leaders live positions (for RSM / Owner view — "ASM"=TL in this DMS)
        tls: List[Dict[str, Any]] = []
        tl_query_ids: List[str] = []
        if role == "regional_manager":
            tl_query_ids = [a["team_leader_id"] async for a in db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0, "team_leader_id": 1})]
        elif role in ("owner", "super_admin"):
            tl_query_ids = [u["id"] async for u in db.users.find({"role": "team_leader"}, {"_id": 0, "id": 1})]
        if tl_query_ids:
            async for u in db.users.find({"id": {"$in": tl_query_ids}, "role": "team_leader"}, {"_id": 0, "password_hash": 0}):
                gps = u.get("last_gps") or {}
                last_at = gps.get("at") or u.get("last_active_at")
                online = False
                try:
                    if last_at:
                        dt = datetime.fromisoformat(last_at.replace("Z", "+00:00"))
                        online = (datetime.now(timezone.utc) - dt).total_seconds() < 300
                except Exception:
                    pass
                tls.append({
                    "id": u["id"], "name": u.get("name"), "phone": u.get("phone"),
                    "lat": gps.get("lat"), "lng": gps.get("lng"),
                    "last_ping_at": last_at, "online": online,
                })

        # CONTINUATION v6 (Task 4): ALL punched-in field staff on the map.
        # Every non-owner role can punch-in + GPS-ping, so surface them all.
        today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        open_ids = [p["salesperson_id"] async for p in db.dms_punch.find(
            {"date": today_str, "out_at": None}, {"_id": 0, "salesperson_id": 1})]
        field_staff: List[Dict[str, Any]] = []
        if open_ids:
            # scope allowed ids for TL / RM (owner + super_admin see everyone)
            allowed: Optional[set] = None
            if role in ("team_leader", "regional_manager"):
                allowed = set(sp_ids)
                allowed.update(tl_query_ids)
                allowed.add(user["id"])
                scoped_dids = dq.get("id", {}).get("$in") if isinstance(dq.get("id"), dict) else None
                dfilter = {"id": {"$in": scoped_dids}} if scoped_dids is not None else {}
                async for d in db.dms_distributors.find(dfilter, {"_id": 0, "user_id": 1, "accountant_user_id": 1, "id": 1}):
                    if d.get("user_id"):
                        allowed.add(d["user_id"])
                    if d.get("accountant_user_id"):
                        allowed.add(d["accountant_user_id"])
                    rf = {"distributor_id": d["id"]}
                    async for r in db.dms_retailers.find(rf, {"_id": 0, "user_id": 1}):
                        if r.get("user_id"):
                            allowed.add(r["user_id"])
            role_labels = {
                "salesperson": "Salesperson", "team_leader": "Team Leader",
                "regional_manager": "Regional Manager", "distributor": "Distributor",
                "distributor_accountant": "Distributor Accountant", "retailer": "Retailer",
                "owner_accountant": "Owner Accountant",
            }
            async for u in db.users.find({"id": {"$in": open_ids}}, {"_id": 0, "password_hash": 0}):
                if u.get("role") == "owner":
                    continue
                if allowed is not None and u["id"] not in allowed:
                    continue
                gps = u.get("last_gps") or {}
                last_at = gps.get("at") or u.get("last_active_at")
                online = False
                try:
                    if last_at:
                        dt = datetime.fromisoformat(last_at.replace("Z", "+00:00"))
                        online = (datetime.now(timezone.utc) - dt).total_seconds() < 300
                except Exception:
                    pass
                field_staff.append({
                    "id": u["id"], "name": u.get("name"), "phone": u.get("phone"),
                    "role": u.get("role"), "role_label": role_labels.get(u.get("role"), u.get("role")),
                    "lat": gps.get("lat"), "lng": gps.get("lng"),
                    "last_ping_at": last_at, "online": online, "punched_in": True,
                })

        return {"salespersons": sps, "distributors": dists, "retailers": rets,
                "team_leaders": tls, "field_staff": field_staff}

    @router.get("/tracking/salesperson/{sid}")
    async def tracking_salesperson_detail(
        sid: str,
        date: Optional[str] = Query(None, description="YYYY-MM-DD; defaults to today"),
        user: dict = Depends(get_current_user),
    ):
        """
        Full detail for one salesperson on one date:
          - profile + current live location
          - punch in / out + working hours
          - full route (ordered pings)
          - total distance travelled
          - visited distributors / retailers (proximity < 200m to any ping)
        """
        # RBAC — same visibility rules as /tracking/live
        allowed = await _sp_visible_ids_for(user)
        if sid not in allowed:
            raise HTTPException(status_code=403, detail="Not permitted to view this salesperson")

        day = date or _today()
        sp = await db.users.find_one({"id": sid, "role": "salesperson"}, {"_id": 0, "password_hash": 0})
        if not sp:
            raise HTTPException(status_code=404, detail="Salesperson not found")

        # punch
        punch = await db.dms_punch.find_one({"salesperson_id": sid, "date": day}, {"_id": 0})
        working_hours = 0.0
        if punch and punch.get("in_at"):
            try:
                a = datetime.fromisoformat(punch["in_at"].replace("Z", "+00:00"))
                b = datetime.fromisoformat((punch.get("out_at") or _now()).replace("Z", "+00:00"))
                working_hours = max(0.0, (b - a).total_seconds() / 3600.0)
            except Exception:
                pass

        # pings for the day
        pings = await db.dms_sp_pings.find(
            {"salesperson_id": sid, "date": day}, {"_id": 0}
        ).sort("created_at", 1).to_list(5000)

        # distance
        distance_km = 0.0
        for i in range(1, len(pings)):
            distance_km += _haversine_km(pings[i-1]["lat"], pings[i-1]["lng"], pings[i]["lat"], pings[i]["lng"])

        # visited shops = distributors/retailers with a ping within 200m
        def _near(pt_lat, pt_lng) -> bool:
            for p in pings:
                if _haversine_km(pt_lat, pt_lng, p["lat"], p["lng"]) < 0.20:
                    return True
            return False

        visited_dists = []
        async for d in db.dms_distributors.find({"active": True, "gps_lat": {"$ne": None}}, {"_id": 0}):
            if d.get("gps_lat") is None or d.get("gps_lng") is None: continue
            if _near(d["gps_lat"], d["gps_lng"]):
                visited_dists.append({"id": d["id"], "name": d["name"], "lat": d.get("gps_lat"), "lng": d.get("gps_lng")})

        visited_rets = []
        async for r in db.dms_retailers.find({"active": True, "gps_lat": {"$ne": None}}, {"_id": 0}):
            if r.get("gps_lat") is None or r.get("gps_lng") is None: continue
            if _near(r["gps_lat"], r["gps_lng"]):
                visited_rets.append({"id": r["id"], "name": r["name"], "lat": r.get("gps_lat"), "lng": r.get("gps_lng")})

        return {
            "salesperson": {
                "id": sp["id"], "name": sp.get("name"), "phone": sp.get("phone"),
                "current_gps": sp.get("last_gps"),
                "last_active_at": sp.get("last_active_at"),
            },
            "date": day,
            "punch": punch,
            "working_hours": _round(working_hours, 2),
            "distance_km": _round(distance_km, 2),
            "route": pings,
            "visited": {"distributors": visited_dists, "retailers": visited_rets},
        }

    @router.get("/tracking/salesperson/{sid}/history")
    async def tracking_history(
        sid: str,
        days: int = Query(30, ge=1, le=365),
        user: dict = Depends(get_current_user),
    ):
        """Date-wise summary (last N days) for the salesperson."""
        allowed = await _sp_visible_ids_for(user)
        if sid not in allowed:
            raise HTTPException(status_code=403, detail="Not permitted")
        end = datetime.now(timezone.utc)
        start = end - timedelta(days=days)
        # group pings by date
        pipe = [
            {"$match": {"salesperson_id": sid, "created_at": {"$gte": start.isoformat()}}},
            {"$group": {"_id": "$date", "count": {"$sum": 1}}},
            {"$sort": {"_id": -1}},
        ]
        counts = {r["_id"]: r["count"] async for r in db.dms_sp_pings.aggregate(pipe)}
        punches = {p["date"]: p async for p in db.dms_punch.find({"salesperson_id": sid, "date": {"$gte": start.strftime("%Y-%m-%d")}}, {"_id": 0})}
        out = []
        for date_str, cnt in sorted(counts.items(), reverse=True):
            p = punches.get(date_str)
            hrs = 0.0
            if p and p.get("in_at"):
                try:
                    a = datetime.fromisoformat(p["in_at"].replace("Z", "+00:00"))
                    b = datetime.fromisoformat((p.get("out_at") or _now()).replace("Z", "+00:00"))
                    hrs = (b - a).total_seconds() / 3600.0
                except Exception:
                    pass
            out.append({
                "date": date_str, "pings": cnt,
                "in_at": (p or {}).get("in_at"), "out_at": (p or {}).get("out_at"),
                "working_hours": _round(hrs, 2),
            })
        return {"data": out}

    @router.get("/tracking/salesperson/{sid}/routes")
    async def tracking_multi_routes(
        sid: str,
        days: int = Query(7, ge=1, le=31),
        user: dict = Depends(get_current_user),
    ):
        """Per-day full routes for the last N days — for multi-day route comparison/playback."""
        allowed = await _sp_visible_ids_for(user)
        if sid not in allowed:
            raise HTTPException(status_code=403, detail="Not permitted")
        start = datetime.now(timezone.utc) - timedelta(days=days)
        # all pings in window, ordered
        pings = await db.dms_sp_pings.find(
            {"salesperson_id": sid, "created_at": {"$gte": start.isoformat()}},
            {"_id": 0, "lat": 1, "lng": 1, "created_at": 1, "date": 1},
        ).sort("created_at", 1).to_list(20000)
        punches = {p["date"]: p async for p in db.dms_punch.find(
            {"salesperson_id": sid, "date": {"$gte": start.strftime("%Y-%m-%d")}}, {"_id": 0})}
        by_date: Dict[str, list] = {}
        for p in pings:
            by_date.setdefault(p["date"], []).append(p)
        out = []
        for date_str in sorted(by_date.keys(), reverse=True):
            pts = by_date[date_str]
            dist = 0.0
            for i in range(1, len(pts)):
                dist += _haversine_km(pts[i-1]["lat"], pts[i-1]["lng"], pts[i]["lat"], pts[i]["lng"])
            pk = punches.get(date_str) or {}
            out.append({
                "date": date_str,
                "points": pts,
                "distance_km": _round(dist, 2),
                "in_at": pk.get("in_at"),
                "out_at": pk.get("out_at"),
            })
        sp = await db.users.find_one({"id": sid}, {"_id": 0, "name": 1, "phone": 1})
        return {"salesperson": sp or {"name": ""}, "days": days, "data": out}

    # =========================================================================
    # OWNER — Complete User Management + Impersonation (Phase 1)
    # =========================================================================
    # Roles the owner can create via the quick "New User" panel. Distributor &
    # Retailer are intentionally EXCLUDED — their logins must be created through
    # the full onboarding flow (Distributors / Retailers pages) which requires
    # complete details + KYC + documents.
    OWNER_MANAGEABLE_ROLES = [
        "owner_accountant", "distributor_accountant",
        "salesperson", "team_leader", "regional_manager",
    ]

    def _is_online(u: dict) -> bool:
        """A user is considered online if last activity ping / login was <5min ago."""
        last = u.get("last_login_at") or u.get("last_active_at")
        if not last:
            return False
        try:
            dt = datetime.fromisoformat(last.replace("Z", "+00:00"))
            return (datetime.now(timezone.utc) - dt).total_seconds() < 300
        except Exception:
            return False

    @router.get("/owner/users")
    async def owner_list_users(role: Optional[str] = None, user: dict = Depends(owner_only)):
        q: Dict[str, Any] = {"tenant_id": DMS_TENANT_ID}
        if role:
            q["role"] = role
        docs = await db.users.find(q, {"_id": 0, "password_hash": 0}).sort("role", 1).to_list(2000)
        for u in docs:
            u["online"] = _is_online(u)
        return {"data": docs, "count": len(docs)}

    @router.post("/owner/users")
    async def owner_create_user(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        import bcrypt
        for k in ("email", "password", "name", "role"):
            if not body.get(k):
                raise HTTPException(status_code=400, detail=f"{k} required")
        role = body["role"]
        if role not in OWNER_MANAGEABLE_ROLES:
            raise HTTPException(status_code=400, detail=f"Cannot create role={role} from owner panel")
        email = body["email"].lower().strip()
        if await db.users.find_one({"email": email}):
            raise HTTPException(status_code=400, detail=f"Email {email} already exists")
        uid = _nid("usr")
        doc = {
            "id": uid,
            "tenant_id": DMS_TENANT_ID,
            "email": email,
            "name": body["name"],
            "role": role,
            "phone": body.get("phone", ""),
            "password_hash": bcrypt.hashpw(body["password"].encode(), bcrypt.gensalt()).decode(),
            "active": True,
            "created_at": _now(),
            "avatar": "".join([w[0] for w in body["name"].split()[:2]]).upper(),
            "created_by": user["id"],
        }
        # optional linkage
        for k in ("distributor_id", "retailer_id"):
            if body.get(k):
                doc[k] = body[k]
        await db.users.insert_one(doc)
        doc.pop("password_hash", None)
        doc.pop("_id", None)
        return {"ok": True, "user": doc}

    @router.patch("/owner/users/{uid}")
    async def owner_update_user(uid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        target = await db.users.find_one({"id": uid, "tenant_id": DMS_TENANT_ID})
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        updatable = {"name", "phone", "active", "distributor_id", "retailer_id"}
        upd = {k: v for k, v in body.items() if k in updatable}
        # Owner may also change the login email (login id) of any user, incl. self.
        if "email" in body:
            new_email = str(body.get("email") or "").strip().lower()
            if not new_email or "@" not in new_email:
                raise HTTPException(status_code=400, detail="A valid email is required")
            clash = await db.users.find_one({"email": new_email, "id": {"$ne": uid}})
            if clash:
                raise HTTPException(status_code=400, detail=f"Email {new_email} is already in use")
            upd["email"] = new_email
        if not upd:
            raise HTTPException(status_code=400, detail="Nothing to update")
        upd["updated_at"] = _now()
        await db.users.update_one({"id": uid}, {"$set": upd})
        return {"ok": True}

    @router.delete("/owner/users/{uid}")
    async def owner_delete_user(uid: str, user: dict = Depends(owner_only)):
        if uid == user["id"]:
            raise HTTPException(status_code=400, detail="You cannot delete your own account")
        target = await db.users.find_one({"id": uid, "tenant_id": DMS_TENANT_ID}, {"_id": 0})
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        if target.get("role") == "owner":
            raise HTTPException(status_code=400, detail="Cannot delete the company owner account")
        await db.users.delete_one({"id": uid})
        # unlink from distributor/retailer profile if any (keep the business entity)
        await db.dms_distributors.update_many({"user_id": uid}, {"$set": {"user_id": None}})
        await db.dms_retailers.update_many({"user_id": uid}, {"$set": {"user_id": None}})
        return {"ok": True, "deleted": uid}

    @router.post("/owner/users/{uid}/reset-password")
    async def owner_reset_password(uid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        import bcrypt
        new_pw = (body.get("new_password") or "").strip()
        if len(new_pw) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        target = await db.users.find_one({"id": uid, "tenant_id": DMS_TENANT_ID})
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        await db.users.update_one(
            {"id": uid},
            {"$set": {"password_hash": bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()}},
        )
        return {"ok": True}

    @router.post("/owner/impersonate/{uid}")
    async def owner_impersonate(uid: str, user: dict = Depends(owner_only)):
        target = await db.users.find_one({"id": uid, "tenant_id": DMS_TENANT_ID}, {"_id": 0, "password_hash": 0})
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        if target.get("role") == "owner":
            raise HTTPException(status_code=400, detail="Cannot impersonate another owner")
        import jwt as _jwt, os as _os
        payload = {
            "sub": target["id"], "email": target["email"], "role": target["role"],
            "tenant_id": target.get("tenant_id"),
            "exp": datetime.now(timezone.utc) + timedelta(hours=2),
            "type": "access",
            "impersonated_by": user["id"],
        }
        tok = _jwt.encode(payload, _os.environ["JWT_SECRET"], algorithm="HS256")
        return {"token": tok, "user": target, "impersonated_by": {"id": user["id"], "name": user.get("name"), "email": user.get("email")}}


    # =========================================================================
    # PRODUCTS — Excel Import / Export (Owner)
    # =========================================================================
    @router.get("/owner/products/export")
    async def export_products(user: dict = Depends(owner_or_accountant)):
        from openpyxl import Workbook
        from io import BytesIO
        from fastapi.responses import Response
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["sku_code", "name", "category_name", "description", "box_qty", "hsn", "gst_pct", "unit_price", "coupons_per_box", "points_value", "active"]
        ws.append(headers)
        cats = {c["id"]: c["name"] async for c in db.dms_categories.find({}, {"_id": 0, "id": 1, "name": 1})}
        async for p in db.dms_products.find({}, {"_id": 0}):
            ws.append([
                p.get("sku_code"), p.get("name"),
                cats.get(p.get("category_id"), ""),
                p.get("description", ""),
                p.get("box_qty", 0), p.get("hsn", ""),
                p.get("gst_pct", 18),
                p.get("unit_price", 0),
                p.get("coupons_per_box", 100),
                p.get("points_value", 10),
                bool(p.get("active", True)),
            ])
        # widen columns
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws.column_dimensions[col_letter].width = 18
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'},
        )

    @router.post("/owner/products/import")
    async def import_products(file: UploadFile = File(...), user: dict = Depends(owner_or_accountant)):
        from openpyxl import load_workbook
        from io import BytesIO
        raw = await file.read()
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
            ws = wb.active
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read xlsx: {e}")
        # read headers
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise HTTPException(status_code=400, detail="Empty file")
        headers = [str(h or "").strip().lower() for h in header_row]
        col = {h: i for i, h in enumerate(headers)}
        required = ["sku_code", "name", "category_name", "box_qty", "unit_price"]
        for req in required:
            if req not in col:
                raise HTTPException(status_code=400, detail=f"Missing required column: {req}")

        # cache category id map
        cats = {c["name"].strip().lower(): c["id"] async for c in db.dms_categories.find({}, {"_id": 0, "name": 1, "id": 1})}

        created, updated, skipped, errors = 0, 0, 0, []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or v == "" for v in row):
                continue
            try:
                sku = (str(row[col["sku_code"]]) if row[col["sku_code"]] is not None else "").strip()
                name = (str(row[col["name"]]) if row[col["name"]] is not None else "").strip()
                cat_name = (str(row[col["category_name"]]) if row[col["category_name"]] is not None else "").strip()
                if not sku or not name or not cat_name:
                    skipped += 1; errors.append(f"Row {i}: missing sku/name/category"); continue
                cat_id = cats.get(cat_name.lower())
                if not cat_id:
                    # auto-create category
                    cid = _nid("cat")
                    await db.dms_categories.insert_one({"id": cid, "name": cat_name, "created_at": _now()})
                    cats[cat_name.lower()] = cid
                    cat_id = cid
                box_qty = int(row[col["box_qty"]] or 0)
                unit_price = float(row[col["unit_price"]] or 0)
                gst = float(row[col.get("gst_pct", -1)] if "gst_pct" in col and row[col["gst_pct"]] is not None else 18)
                hsn = str(row[col.get("hsn", -1)] or "") if "hsn" in col else ""
                desc = str(row[col.get("description", -1)] or "") if "description" in col else ""
                per_box = int(row[col["coupons_per_box"]] or 100) if "coupons_per_box" in col and row[col["coupons_per_box"]] is not None else 100
                points = float(row[col["points_value"]] or 10) if "points_value" in col and row[col["points_value"]] is not None else 10
                active = bool(row[col["active"]]) if "active" in col and row[col["active"]] is not None else True

                existing = await db.dms_products.find_one({"sku_code": sku}, {"_id": 0})
                if existing:
                    upd = {
                        "name": name, "category_id": cat_id, "description": desc,
                        "box_qty": box_qty, "hsn": hsn, "gst_pct": _round(gst),
                        "coupons_per_box": per_box, "points_value": _round(points),
                        "active": active, "updated_at": _now(),
                    }
                    if float(existing.get("unit_price", 0)) != unit_price:
                        # price change → close previous batch, open new (same as PUT path)
                        await db.dms_price_batches.update_one({"product_id": existing["id"], "to_date": None}, {"$set": {"to_date": _now()}})
                        await db.dms_price_batches.insert_one({
                            "id": _nid("pb"), "product_id": existing["id"], "price": _round(unit_price),
                            "from_date": _now(), "to_date": None, "created_at": _now(),
                        })
                        upd["previous_price"] = existing.get("unit_price")
                        upd["unit_price"] = _round(unit_price)
                    await db.dms_products.update_one({"id": existing["id"]}, {"$set": upd})
                    updated += 1
                else:
                    pid = _nid("prd")
                    await db.dms_products.insert_one({
                        "id": pid, "name": name, "category_id": cat_id, "sku_code": sku,
                        "description": desc, "box_qty": box_qty, "unit_price": _round(unit_price),
                        "previous_price": None, "hsn": hsn, "gst_pct": _round(gst),
                        "coupons_per_box": per_box, "points_value": _round(points),
                        "active": active, "created_at": _now(),
                    })
                    await db.dms_price_batches.insert_one({
                        "id": _nid("pb"), "product_id": pid, "price": _round(unit_price),
                        "from_date": _now(), "to_date": None, "created_at": _now(),
                    })
                    created += 1
            except Exception as e:
                skipped += 1
                errors.append(f"Row {i}: {e}")

        return {"ok": True, "created": created, "updated": updated, "skipped": skipped, "errors": errors[:20]}

    # =========================================================================
    # SUPER ADMIN — login-as (impersonation)
    # =========================================================================
    @router.get("/admin/users")
    async def admin_list_users(user: dict = Depends(get_current_user)):
        if user["role"] not in ("super_admin", "owner"):
            raise HTTPException(status_code=403, detail="Owner only")
        docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
        return {"data": docs}

    @router.post("/admin/impersonate/{uid}")
    async def impersonate(uid: str, user: dict = Depends(get_current_user)):
        if user["role"] not in ("super_admin", "owner"):
            raise HTTPException(status_code=403, detail="Owner only")
        target = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        # Generate a token for the target user (super_admin origin recorded)
        import jwt as _jwt, os as _os
        payload = {
            "sub": target["id"], "email": target["email"], "role": target["role"],
            "tenant_id": target.get("tenant_id"),
            "exp": datetime.now(timezone.utc) + timedelta(hours=2),
            "type": "access",
            "impersonated_by": user["id"],
        }
        tok = _jwt.encode(payload, _os.environ["JWT_SECRET"], algorithm="HS256")
        return {"token": tok, "user": target}

    # =========================================================================
    # PRINTABLE E-BILL / RETAILER BILL data  (Vyapar-style unified invoice)
    # =========================================================================
    def _company_seller_block(s: Dict[str, Any]) -> Dict[str, Any]:
        """GO OIL company as seller (from global settings)."""
        return {
            "name": s.get("company_name") or "GO OIL Lubricants",
            "gstin": s.get("company_gstin") or "",
            "address": s.get("company_address") or "",
            "state": s.get("company_state") or "",
            "state_code": s.get("company_state_code") or "",
            "phone": s.get("company_phone") or "",
            "email": s.get("company_email") or "",
            "logo_url": s.get("company_logo_url") or "",
            "bank_name": s.get("company_bank_name") or "",
            "bank_account": s.get("company_bank_account") or "",
            "bank_ifsc": s.get("company_bank_ifsc") or "",
            "bank_branch": s.get("company_bank_branch") or "",
            "upi_id": s.get("company_upi_id") or "",
            "upi_name": s.get("company_upi_name") or (s.get("company_name") or ""),
            "qr_url": "",
            "signatory": s.get("invoice_signatory") or (s.get("company_name") or ""),
        }

    def _distributor_seller_block(d: Dict[str, Any], s: Dict[str, Any]) -> Dict[str, Any]:
        """Distributor as seller (for retailer / direct-sale bills)."""
        d = d or {}
        kyc = d.get("kyc") or {}
        bank = d.get("bank") or {}
        return {
            "name": d.get("name") or "",
            "gstin": bank.get("gstin") or kyc.get("gstin") or "",
            "address": d.get("address") or "",
            "state": d.get("state") or d.get("region") or "",
            "state_code": d.get("state_code") or "",
            "phone": d.get("phone") or "",
            "email": d.get("email") or "",
            "logo_url": s.get("company_logo_url") or "",
            "bank_name": bank.get("bank_name") or kyc.get("bank_name") or "",
            "bank_account": bank.get("bank_account") or kyc.get("bank_account") or "",
            "bank_ifsc": bank.get("bank_ifsc") or kyc.get("bank_ifsc") or "",
            "bank_branch": bank.get("bank_branch") or "",
            "upi_id": bank.get("upi_id") or "",
            "upi_name": bank.get("upi_name") or d.get("name") or "",
            "qr_url": bank.get("qr_url") or "",
            "signatory": d.get("name") or "",
        }

    def _party_buyer_block(p: Dict[str, Any]) -> Dict[str, Any]:
        p = p or {}
        kyc = p.get("kyc") or {}
        return {
            "name": p.get("name") or "",
            "gstin": kyc.get("gstin") or "",
            "address": p.get("address") or "",
            "state": p.get("state") or p.get("region") or "",
            "state_code": p.get("state_code") or "",
            "phone": p.get("phone") or "",
        }

    def _assemble_invoice(*, doc_title: str, doc_no: str, date: str,
                          seller: Dict[str, Any], buyer: Dict[str, Any],
                          ship_to: Optional[Dict[str, Any]],
                          items: List[Dict[str, Any]], subtotal: float, gst_total: float,
                          total: float, settings: Dict[str, Any],
                          transport: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        inv_items = []
        for it in (items or []):
            qb = it.get("billed_qty_boxes")
            if qb is None:
                qb = it.get("dispatched_qty_boxes", it.get("qty_boxes_dispatched", it.get("qty_boxes_fulfilled", it.get("qty_boxes", 0))))
            qp = it.get("dispatched_qty_pcs", it.get("qty_pcs_dispatched", it.get("qty_pcs", 0)))
            try:
                qb = int(qb or 0); qp = int(qp or 0)
            except Exception:
                qb, qp = 0, 0
            qty_bits = []
            if qb:
                qty_bits.append(f"{qb} Box")
            if qp:
                qty_bits.append(f"{qp} Pcs")
            rate = it.get("box_price", it.get("unit_price", 0))
            inv_items.append({
                "name": it.get("product_name") or it.get("name") or "",
                "sku_code": it.get("sku_code") or "",
                "hsn": it.get("hsn") or it.get("hsn_sac") or "27101980",
                "qty_boxes": qb,
                "qty_pcs": qp,
                "qty_label": " + ".join(qty_bits) or "-",
                "rate": _round(rate),
                "taxable": _round(it.get("line_subtotal", 0)),
                "gst_pct": _round(it.get("gst_pct", 0)),
                "gst_amt": _round(it.get("line_gst", 0)),
                "amount": _round(it.get("line_total", 0)),
            })
        breakup = _gst_breakup(gst_total, seller.get("state_code", ""), buyer.get("state_code", ""))
        grand = _round(total)
        grand_rounded = float(round(grand))
        round_off = _round(grand_rounded - grand)
        upi_qr = seller.get("qr_url") or _make_upi_qr_dataurl(
            seller.get("upi_id", ""), seller.get("upi_name") or seller.get("name", ""), grand_rounded)
        return {
            "doc_title": doc_title,
            "doc_no": doc_no,
            "date": date,
            "seller": seller,
            "bill_to": buyer,
            "ship_to": ship_to or buyer,
            "transport": transport or {},
            "items": inv_items,
            "totals": {
                "subtotal": _round(subtotal),
                "gst_total": _round(gst_total),
                "sgst": breakup["sgst"],
                "cgst": breakup["cgst"],
                "igst": breakup["igst"],
                "is_interstate": breakup["is_interstate"],
                "round_off": round_off,
                "grand_total": grand_rounded,
            },
            "amount_in_words": _num_to_words_inr(grand_rounded),
            "terms": settings.get("invoice_terms") or "",
            "message": settings.get("invoice_message") or "",
            "signatory": seller.get("signatory") or seller.get("name", ""),
            "upi_qr": upi_qr,
            "acknowledgement_enabled": bool(settings.get("invoice_show_acknowledgement")),
        }

    @router.get("/print/ebill/{ebill_id}")
    async def print_ebill(ebill_id: str, user: dict = Depends(get_current_user)):
        eb = await db.dms_ebills.find_one({"id": ebill_id}, {"_id": 0})
        if not eb:
            raise HTTPException(status_code=404, detail="e-Bill not found")
        # RBAC
        role = user.get("role")
        if role in ("distributor", "distributor_accountant") and eb["distributor_id"] != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        dist = await db.dms_distributors.find_one({"id": eb["distributor_id"]}, {"_id": 0})
        eb["distributor"] = dist
        s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0}) or {}
        eb["invoice_terms"] = s.get("invoice_terms") or ""
        eb["invoice_message"] = s.get("invoice_message") or ""
        eb["company_name"] = s.get("company_name") or "GO OIL Lubricants"
        # Vyapar-style unified invoice: seller = GO OIL, buyer = distributor
        eb["invoice"] = _assemble_invoice(
            doc_title="TAX INVOICE",
            doc_no=eb.get("ebill_no") or eb.get("id"),
            date=(eb.get("created_at") or "")[:10],
            seller=_company_seller_block(s),
            buyer=_party_buyer_block(dist or {}),
            ship_to=_party_buyer_block(dist or {}),
            items=eb.get("items") or [],
            subtotal=eb.get("subtotal", 0),
            gst_total=eb.get("gst_total", 0),
            total=eb.get("total", 0),
            settings=s,
        )
        return eb

    @router.get("/print/retailer-bill/{bill_id}")
    async def print_retailer_bill(bill_id: str, user: dict = Depends(get_current_user)):
        b = await db.dms_retailer_bills.find_one({"id": bill_id}, {"_id": 0})
        if not b:
            raise HTTPException(status_code=404, detail="Bill not found")
        role = user.get("role")
        if role == "retailer" and b["retailer_id"] != user.get("retailer_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        if role in ("distributor", "distributor_accountant") and b["distributor_id"] != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        retailer = await db.dms_retailers.find_one({"id": b["retailer_id"]}, {"_id": 0})
        distributor = await db.dms_distributors.find_one({"id": b["distributor_id"]}, {"_id": 0})
        b["retailer"] = retailer
        b["distributor"] = distributor
        s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0}) or {}
        b["invoice_terms"] = s.get("invoice_terms") or ""
        b["invoice_message"] = s.get("invoice_message") or ""
        b["company_name"] = s.get("company_name") or "GO OIL Lubricants"
        # Vyapar-style unified invoice: seller = distributor, buyer = retailer OR walk-in customer
        cust = b.get("customer") or {}
        if cust and cust.get("name"):
            buyer_block = {
                "name": cust.get("name") or "", "gstin": cust.get("gstin") or "",
                "address": cust.get("address") or "", "state": "", "state_code": "",
                "phone": cust.get("phone") or "",
            }
        else:
            buyer_block = _party_buyer_block(retailer or {})
        b["invoice"] = _assemble_invoice(
            doc_title="TAX INVOICE",
            doc_no=b.get("bill_no") or b.get("id"),
            date=b.get("date") or (b.get("created_at") or "")[:10],
            seller=_distributor_seller_block(distributor or {}, s),
            buyer=buyer_block,
            ship_to=buyer_block,
            items=b.get("items") or [],
            subtotal=b.get("subtotal", 0),
            gst_total=b.get("gst_total", 0),
            total=b.get("total", 0),
            settings=s,
            transport=b.get("transport") or {},
        )
        return b

    # =========================================================================
    # SETTINGS  (global — GST % is configured here; default 0)
    # =========================================================================
    async def _get_settings() -> Dict[str, Any]:
        doc = await db.dms_settings.find_one({"id": "global"}, {"_id": 0})
        if not doc:
            doc = {"id": "global", "gst_pct": 0.0, "company_name": "GO OIL Lubricants", "updated_at": _now()}
            await db.dms_settings.insert_one(doc)
        return doc

    @router.get("/settings")
    async def get_settings(user: dict = Depends(get_current_user)):
        s = await _get_settings()
        # ensure default for stop_sale_on_negative
        if "stop_sale_on_negative" not in s:
            s["stop_sale_on_negative"] = True
        return _clean(s)

    @router.put("/settings")
    async def update_settings(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        upd: Dict[str, Any] = {}
        if "gst_pct" in body:
            try:
                gst = float(body["gst_pct"])
            except Exception:
                raise HTTPException(status_code=400, detail="gst_pct must be a number")
            if gst < 0 or gst > 100:
                raise HTTPException(status_code=400, detail="gst_pct must be between 0 and 100")
            upd["gst_pct"] = gst
        if "company_name" in body and isinstance(body["company_name"], str):
            upd["company_name"] = body["company_name"].strip() or "GO OIL Lubricants"
        # Phase 2A: invoice customization
        if "invoice_terms" in body:
            upd["invoice_terms"] = str(body.get("invoice_terms") or "").strip()
        if "invoice_message" in body:
            upd["invoice_message"] = str(body.get("invoice_message") or "").strip()
        # Phase 2A: Financial Year lock date (YYYY-MM-DD). Can only move forward.
        if "fy_lock_date" in body:
            lock = str(body.get("fy_lock_date") or "").strip()
            if lock:
                try:
                    datetime.strptime(lock, "%Y-%m-%d")
                except Exception:
                    raise HTTPException(status_code=400, detail="fy_lock_date must be YYYY-MM-DD")
                cur = await _get_settings()
                cur_lock = cur.get("fy_lock_date") or ""
                if cur_lock and lock < cur_lock:
                    raise HTTPException(status_code=400, detail=f"fy_lock_date can only move forward (current: {cur_lock})")
            upd["fy_lock_date"] = lock or None
        # Phase 2B: Stop Sale on Negative Stock toggle
        if "stop_sale_on_negative" in body:
            upd["stop_sale_on_negative"] = bool(body.get("stop_sale_on_negative"))
        # CONTINUATION v6: Company profile + invoice options (Vyapar-style invoice)
        _company_str_fields = [
            "company_gstin", "company_address", "company_state", "company_state_code",
            "company_phone", "company_email", "company_logo_url",
            "company_bank_name", "company_bank_account", "company_bank_ifsc", "company_bank_branch",
            "company_upi_id", "company_upi_name", "invoice_signatory",
        ]
        for k in _company_str_fields:
            if k in body:
                upd[k] = str(body.get(k) or "").strip()
        if "invoice_show_acknowledgement" in body:
            upd["invoice_show_acknowledgement"] = bool(body.get("invoice_show_acknowledgement"))
        upd["updated_at"] = _now()
        await db.dms_settings.update_one({"id": "global"}, {"$set": upd}, upsert=True)
        s = await _get_settings()
        return _clean(s)

    # -- Financial year lock helper -----------------------------------------
    async def _fy_lock_date() -> Optional[str]:
        s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0, "fy_lock_date": 1}) or {}
        return s.get("fy_lock_date")

    async def _check_fy_lock(entity_date_iso: Optional[str], what: str = "record") -> None:
        """Raise 400 if entity's date is on/before the FY lock date. Accepts ISO datetime or YYYY-MM-DD."""
        lock = await _fy_lock_date()
        if not lock or not entity_date_iso:
            return
        try:
            d = entity_date_iso[:10]
        except Exception:
            return
        if d <= lock:
            raise HTTPException(status_code=400, detail=f"Financial year locked (up to {lock}); {what} on {d} cannot be modified")

    @router.post("/finance/fy-close")
    async def fy_close(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        """Close (lock) the financial year up to a given date. Alias of PUT /settings with fy_lock_date."""
        lock = str(body.get("lock_date") or "").strip()
        if not lock:
            raise HTTPException(status_code=400, detail="lock_date required (YYYY-MM-DD)")
        try:
            datetime.strptime(lock, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="lock_date must be YYYY-MM-DD")
        cur = await _get_settings()
        cur_lock = cur.get("fy_lock_date") or ""
        if cur_lock and lock < cur_lock:
            raise HTTPException(status_code=400, detail=f"lock_date can only move forward (current: {cur_lock})")
        await db.dms_settings.update_one({"id": "global"}, {"$set": {"fy_lock_date": lock, "fy_locked_at": _now(), "fy_locked_by": user["id"], "updated_at": _now()}}, upsert=True)
        return {"ok": True, "fy_lock_date": lock}

    # =========================================================================
    # EXPENSES  (Phase 2A) — CRUD for all roles except retailer
    # =========================================================================
    def _exp_can_view(role: str) -> bool:
        return role in ("owner", "owner_accountant", "super_admin", "distributor", "distributor_accountant", "salesperson", "team_leader", "regional_manager")

    def _exp_all_visible_roles() -> tuple:
        return ("owner", "owner_accountant", "super_admin")

    @router.get("/expenses")
    async def list_expenses(user: dict = Depends(get_current_user), start: Optional[str] = None, end: Optional[str] = None, category: Optional[str] = None):
        role = user.get("role")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        q: Dict[str, Any] = {}
        if role in _exp_all_visible_roles():
            pass  # owner/accountant/super_admin see all
        elif role == "regional_manager":
            # RSM sees own expenses + expenses of salespersons reporting to them
            tlids = [a["team_leader_id"] async for a in db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0, "team_leader_id": 1})]
            dids = list({a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "distributor_id": 1})})
            sp_ids = list({a["salesperson_id"] async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "salesperson_id": 1})})
            q["created_by"] = {"$in": [user["id"]] + sp_ids}
        else:
            q["created_by"] = user["id"]
        if start:
            q["date"] = q.get("date", {}); q["date"]["$gte"] = start
        if end:
            q["date"] = q.get("date", {}); q["date"]["$lte"] = end
        if category:
            q["category"] = category
        docs = await db.dms_expenses.find(q, {"_id": 0}).sort("date", -1).to_list(1000)
        # summary
        total = sum(float(d.get("amount", 0)) for d in docs)
        # enrich creator names
        ids = list({d.get("created_by") for d in docs if d.get("created_by")})
        name_map = {}
        if ids:
            async for u in db.users.find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "name": 1, "role": 1}):
                name_map[u["id"]] = {"name": u.get("name"), "role": u.get("role")}
        for d in docs:
            u = name_map.get(d.get("created_by") or "")
            if u:
                d["created_by_name"] = u.get("name")
                d["created_by_role"] = d.get("created_by_role") or u.get("role")
        return {"data": docs, "count": len(docs), "total": _round(total)}

    @router.post("/expenses")
    async def create_expense(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        category = str(body.get("category") or "").strip()
        try:
            amount = float(body.get("amount") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="amount must be a number")
        if not category or amount <= 0:
            raise HTTPException(status_code=400, detail="category and amount>0 required")
        date = str(body.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d"))[:10]
        # Enforce FY lock on backdated new entries
        await _check_fy_lock(date, "expense")
        seq = await db.dms_expenses.count_documents({}) + 1
        # Salesperson expenses go through approval flow (RSM → Owner). No status/receipt on submit.
        if role == "salesperson":
            status = "submitted"
        else:
            status = body.get("status") or "Approved"
        doc = {
            "id": _nid("exp"),
            "expense_no": body.get("expense_no") or f"EXP-{80000 + seq}",
            "category": category,
            "amount": _round(amount),
            "date": date,
            "description": str(body.get("description") or "").strip(),
            "vendor": str(body.get("vendor") or "").strip(),
            "receipt_url": (str(body.get("receipt_url") or "").strip() or None) if role != "salesperson" else None,
            "status": status,
            "created_by": user["id"],
            "created_by_role": role,
            "created_at": _now(),
        }
        await db.dms_expenses.insert_one(doc)
        # Notify RSM(s) responsible for this salesperson
        if role == "salesperson":
            try:
                dids = [a["distributor_id"] async for a in db.dms_sp_assignments.find({"salesperson_id": user["id"]}, {"_id": 0, "distributor_id": 1})]
                tlids = list({a["team_leader_id"] async for a in db.dms_tl_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "team_leader_id": 1})})
                rm_ids = list({a["regional_manager_id"] async for a in db.dms_rm_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "regional_manager_id": 1})})
                for rid in rm_ids:
                    await notify(rid, "expense_submitted", "Expense pending review",
                                 f"{user.get('name','A salesperson')} submitted an expense of ₹{_round(amount)} for review.", "/dms")
            except Exception:
                pass
        return _clean(doc)

    @router.put("/expenses/{eid}")
    async def update_expense(eid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        doc = await db.dms_expenses.find_one({"id": eid}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Expense not found")
        # only creator OR owner/accountant can edit
        if role not in _exp_all_visible_roles() and doc.get("created_by") != user["id"]:
            raise HTTPException(status_code=403, detail="You can only edit your own expenses")
        # FY lock on existing date AND on new date
        await _check_fy_lock(doc.get("date"), "expense")
        upd: Dict[str, Any] = {}
        for f in ("category", "description", "vendor", "receipt_url", "status", "expense_no"):
            if f in body:
                upd[f] = str(body[f] or "").strip() or None
        if "amount" in body:
            try:
                amt = float(body["amount"])
                if amt <= 0:
                    raise ValueError()
                upd["amount"] = _round(amt)
            except Exception:
                raise HTTPException(status_code=400, detail="amount must be a positive number")
        if "date" in body:
            new_date = str(body["date"] or "")[:10]
            await _check_fy_lock(new_date, "expense")
            upd["date"] = new_date
        upd["updated_at"] = _now()
        upd["updated_by"] = user["id"]
        await db.dms_expenses.update_one({"id": eid}, {"$set": upd})
        doc.update(upd)
        return _clean(doc)

    @router.post("/expenses/{eid}/action")
    async def expense_action(eid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        """Approval workflow for salesperson expenses.
        RSM: approve (→ pending owner) or reject (→ rejected) a 'submitted' expense.
        Owner/Accountant: approve (→ approved) or reject (→ rejected) an 'rsm_approved' expense.
        """
        role = user.get("role")
        action = str(body.get("action") or "").lower()
        note = str(body.get("note") or "").strip()
        if action not in ("approve", "reject"):
            raise HTTPException(status_code=400, detail="action must be 'approve' or 'reject'")
        doc = await db.dms_expenses.find_one({"id": eid}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Expense not found")
        status = doc.get("status")
        upd: Dict[str, Any] = {"updated_at": _now(), "updated_by": user["id"]}
        if role == "regional_manager":
            if status != "submitted":
                raise HTTPException(status_code=400, detail="This expense is not pending RSM review")
            # verify this SP reports to this RSM
            tlids = [a["team_leader_id"] async for a in db.dms_rm_assignments.find({"regional_manager_id": user["id"]}, {"_id": 0, "team_leader_id": 1})]
            dids = list({a["distributor_id"] async for a in db.dms_tl_assignments.find({"team_leader_id": {"$in": tlids}}, {"_id": 0, "distributor_id": 1})})
            sp_ids = list({a["salesperson_id"] async for a in db.dms_sp_assignments.find({"distributor_id": {"$in": dids}}, {"_id": 0, "salesperson_id": 1})})
            if doc.get("created_by") not in sp_ids:
                raise HTTPException(status_code=403, detail="This salesperson does not report to you")
            if action == "approve":
                upd["status"] = "rsm_approved"
                upd["rsm_action"] = {"by": user["id"], "at": _now(), "note": note}
                # notify owners for final approval
                async for o in db.users.find({"role": "owner"}, {"_id": 0, "id": 1}):
                    await notify(o["id"], "expense_rsm_approved", "Expense pending final approval",
                                 f"An expense of ₹{doc.get('amount')} was approved by RSM and needs your approval.", "/dms")
            else:
                upd["status"] = "rejected"
                upd["rsm_action"] = {"by": user["id"], "at": _now(), "note": note, "rejected": True}
            await notify(doc.get("created_by"), "expense_update", f"Expense {upd['status'].replace('_',' ')}",
                         f"Your expense of ₹{doc.get('amount')} was {'approved by RSM' if action=='approve' else 'rejected by RSM'}.", "/dms")
        elif role in ("owner", "super_admin", "owner_accountant"):
            if status != "rsm_approved":
                raise HTTPException(status_code=400, detail="This expense is not pending Owner approval")
            if action == "approve":
                upd["status"] = "approved"
                upd["owner_action"] = {"by": user["id"], "at": _now(), "note": note}
            else:
                upd["status"] = "rejected"
                upd["owner_action"] = {"by": user["id"], "at": _now(), "note": note, "rejected": True}
            await notify(doc.get("created_by"), "expense_update", f"Expense {upd['status']}",
                         f"Your expense of ₹{doc.get('amount')} was {'approved' if action=='approve' else 'rejected'} by the Owner.", "/dms")
        else:
            raise HTTPException(status_code=403, detail="Not allowed to action expenses")
        await db.dms_expenses.update_one({"id": eid}, {"$set": upd})
        doc.update(upd)
        return _clean(doc)

    @router.delete("/expenses/{eid}")
    async def delete_expense(eid: str, user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role not in _exp_all_visible_roles():
            raise HTTPException(status_code=403, detail="Only Owner/Accountant can delete expenses")
        doc = await db.dms_expenses.find_one({"id": eid}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Expense not found")
        await _check_fy_lock(doc.get("date"), "expense")
        await db.dms_expenses.delete_one({"id": eid})
        return {"ok": True, "id": eid}

    @router.get("/expenses/categories")
    async def expense_categories(user: dict = Depends(get_current_user)):
        if user.get("role") == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        # distinct categories used
        seen = set()
        async for d in db.dms_expenses.find({}, {"_id": 0, "category": 1}):
            c = (d.get("category") or "").strip()
            if c:
                seen.add(c)
        # baseline defaults
        for c in ("Transport", "Fuel", "Warehouse Rent", "Salaries", "Marketing", "Utilities", "Repairs", "Travel", "Office Supplies", "Miscellaneous"):
            seen.add(c)
        return {"data": sorted(seen)}

    # =========================================================================
    # INVOICE/BILL NUMBER OVERRIDE (Phase 2A) — Owner/Accountant/Distributor
    # =========================================================================
    @router.put("/ebills/{ebid}/number")
    async def update_ebill_number(ebid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role not in ("owner", "owner_accountant", "super_admin"):
            raise HTTPException(status_code=403, detail="Forbidden")
        new_no = str(body.get("ebill_no") or "").strip()
        if not new_no:
            raise HTTPException(status_code=400, detail="ebill_no required")
        eb = await db.dms_ebills.find_one({"id": ebid}, {"_id": 0})
        if not eb:
            raise HTTPException(status_code=404, detail="e-Bill not found")
        await _check_fy_lock(eb.get("created_at"), "e-Bill")
        dup = await db.dms_ebills.find_one({"ebill_no": new_no, "id": {"$ne": ebid}}, {"_id": 0, "id": 1})
        if dup:
            raise HTTPException(status_code=400, detail="ebill_no already used")
        await db.dms_ebills.update_one({"id": ebid}, {"$set": {"ebill_no": new_no, "ebill_no_edited_at": _now(), "ebill_no_edited_by": user["id"]}})
        return {"ok": True, "id": ebid, "ebill_no": new_no}

    @router.put("/retailer-bills/{bid}/number")
    async def update_retailer_bill_number(bid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        b = await db.dms_retailer_bills.find_one({"id": bid}, {"_id": 0})
        if not b:
            raise HTTPException(status_code=404, detail="Bill not found")
        if role not in ("owner", "owner_accountant", "super_admin") and not (role in ("distributor", "distributor_accountant") and user.get("distributor_id") == b.get("distributor_id")):
            raise HTTPException(status_code=403, detail="Forbidden")
        await _check_fy_lock(b.get("created_at"), "Bill")
        new_no = str(body.get("bill_no") or "").strip()
        if not new_no:
            raise HTTPException(status_code=400, detail="bill_no required")
        dup = await db.dms_retailer_bills.find_one({"bill_no": new_no, "id": {"$ne": bid}}, {"_id": 0, "id": 1})
        if dup:
            raise HTTPException(status_code=400, detail="bill_no already used")
        await db.dms_retailer_bills.update_one({"id": bid}, {"$set": {"bill_no": new_no, "bill_no_edited_at": _now(), "bill_no_edited_by": user["id"]}})
        return {"ok": True, "id": bid, "bill_no": new_no}

    # =========================================================================
    # PRICE CIRCULAR  (source-of-truth for pricing history; separate from Product Master)
    # =========================================================================
    @router.get("/price-circulars")
    async def list_price_circulars(user: dict = Depends(get_current_user)):
        headers = await db.dms_price_circulars.find(
            {"kind": {"$ne": "line"}}, {"_id": 0}
        ).sort([("batch_no", -1)]).to_list(500)
        # attach line counts
        for h in headers:
            h["lines_count"] = await db.dms_price_circulars.count_documents(
                {"kind": "line", "circular_id": h["id"]}
            )
        return {"data": headers, "count": len(headers)}

    @router.get("/price-circulars/{cid}")
    async def get_price_circular(cid: str, user: dict = Depends(get_current_user)):
        header = await db.dms_price_circulars.find_one({"id": cid}, {"_id": 0})
        if not header or header.get("kind") == "line":
            raise HTTPException(status_code=404, detail="Circular not found")
        lines = await db.dms_price_circulars.find(
            {"kind": "line", "circular_id": cid}, {"_id": 0}
        ).to_list(2000)
        # enrich with product info
        pids = [ln["product_id"] for ln in lines]
        prods = await db.dms_products.find({"id": {"$in": pids}}, {"_id": 0}).to_list(2000)
        pmap = {p["id"]: p for p in prods}
        cats = {c["id"]: c["name"] async for c in db.dms_categories.find({}, {"_id": 0, "id": 1, "name": 1})}
        for ln in lines:
            p = pmap.get(ln["product_id"], {})
            ln["material_description"] = p.get("material_description", p.get("name", ""))
            ln["grade_specs"] = p.get("grade_specs", "-")
            ln["pack_size"] = p.get("pack_size", "")
            ln["category_name"] = cats.get(p.get("category_id"), "")
        header["lines"] = lines
        return _clean(header)

    async def _publish_circular(title: str, eff_date: str, notes: str,
                                lines: List[Dict[str, Any]], created_by: str) -> Dict[str, Any]:
        """Core Price Circular publish logic — shared by the API endpoint and the
        file importer. Returns the header doc (with lines_count)."""
        # next batch number
        latest = await db.dms_price_circulars.find_one(
            {"kind": {"$ne": "line"}},
            {"_id": 0, "batch_no": 1},
            sort=[("batch_no", -1)],
        )
        next_batch = int((latest or {}).get("batch_no", 0)) + 1
        circular_id = _nid("pcir")

        header_doc = {
            "id": circular_id,
            "tenant_id": DMS_TENANT_ID,
            "title": title,
            "effective_date": eff_date,
            "batch_no": next_batch,
            "batch_label": f"Batch {next_batch} — {title}",
            "is_active": True,
            "notes": (notes or "").strip(),
            "created_by": created_by,
            "created_at": _now(),
        }
        await db.dms_price_circulars.insert_one(header_doc)

        included_product_ids = []
        inserted_lines = 0
        for ln in lines:
            pid = ln.get("product_id")
            if not pid:
                continue
            product = await db.dms_products.find_one({"id": pid}, {"_id": 0})
            if not product:
                continue
            try:
                dlp = float(ln.get("dlp"))
            except Exception:
                raise HTTPException(status_code=400, detail=f"Invalid DLP for product {pid}")
            try:
                mrp = float(ln.get("mrp") or 0)
            except Exception:
                mrp = 0.0
            try:
                margin = float(ln.get("distributor_margin_pct") or 0)
            except Exception:
                margin = 0.0

            # Deactivate the previous active line for this product
            await db.dms_price_circulars.update_many(
                {"kind": "line", "product_id": pid, "is_active": True},
                {"$set": {"is_active": False, "deactivated_at": _now()}},
            )

            # New circular line
            await db.dms_price_circulars.insert_one({
                "id": _nid("pcl"),
                "tenant_id": DMS_TENANT_ID,
                "kind": "line",
                "circular_id": circular_id,
                "product_id": pid,
                "effective_date": eff_date,
                "batch_no": next_batch,
                "mrp": mrp,
                "dlp": dlp,
                "distributor_margin_pct": margin,
                "cash_coupon": (ln.get("cash_coupon") or "").strip(),
                "foc_benefits": (ln.get("foc_benefits") or "").strip(),
                "monthly_gift": (ln.get("monthly_gift") or "").strip(),
                "trade_discount": (ln.get("trade_discount") or "").strip(),
                "is_active": True,
                "created_at": _now(),
            })
            inserted_lines += 1
            included_product_ids.append(pid)

            # Also mirror new DLP into product.unit_price so existing order code works
            old_price = float(product.get("unit_price") or 0)
            if dlp != old_price:
                await db.dms_products.update_one(
                    {"id": pid},
                    {"$set": {"previous_price": old_price, "unit_price": dlp, "updated_at": _now()}},
                )
                # Close open legacy price-batch row + open a new one (for price-history endpoint)
                await db.dms_price_batches.update_one(
                    {"product_id": pid, "to_date": None},
                    {"$set": {"to_date": _now()}}
                )
                await db.dms_price_batches.insert_one({
                    "id": _nid("pb"),
                    "tenant_id": DMS_TENANT_ID,
                    "product_id": pid,
                    "price": dlp,
                    "from_date": _now(),
                    "to_date": None,
                    "created_at": _now(),
                    "circular_id": circular_id,
                    "batch_no": next_batch,
                })

        if inserted_lines == 0:
            # rollback header
            await db.dms_price_circulars.delete_one({"id": circular_id})
            raise HTTPException(status_code=400, detail="No valid product lines")

        header_doc["lines_count"] = inserted_lines
        return _clean(header_doc)

    @router.post("/price-circulars")
    async def create_price_circular(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        """Create a new Price Circular batch. See _publish_circular for behaviour.

        Body: title, effective_date, notes, lines[{product_id, mrp, dlp,
               distributor_margin_pct, cash_coupon, foc_benefits, monthly_gift, trade_discount}]
        """
        title = (body.get("title") or "").strip()
        eff_date = (body.get("effective_date") or "").strip()
        lines = body.get("lines") or []
        if not title:
            raise HTTPException(status_code=400, detail="title required")
        if not eff_date:
            raise HTTPException(status_code=400, detail="effective_date required")
        if not lines:
            raise HTTPException(status_code=400, detail="At least one line required")
        return await _publish_circular(title, eff_date, (body.get("notes") or ""),
                                       lines, user.get("id"))

    # =========================================================================
    # SMART PRICE-LIST IMPORT (Excel / CSV / PDF) — GO OIL circular format
    # =========================================================================
    @router.get("/owner/products/import-template")
    async def products_import_template(user: dict = Depends(owner_or_accountant)):
        from fastapi.responses import Response
        import dms_price_import
        data = dms_price_import.build_template_xlsx()
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="GO_OIL_price_list_template.xlsx"'},
        )

    @router.post("/owner/products/import-circular/preview")
    async def products_import_circular_preview(
        file: UploadFile = File(...),
        user: dict = Depends(owner_or_accountant),
    ):
        """Parse-only: return how many categories/products were detected (no DB writes)."""
        import dms_price_import
        raw = await file.read()
        try:
            parsed = dms_price_import.parse_price_list(raw, file.filename or "")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read the file: {e}")
        products = [i for i in parsed["items"] if i["type"] == "product"]
        categories = [i["name"] for i in parsed["items"] if i["type"] == "category"]
        sample = [{
            "material_description": p["material_description"],
            "grade_specs": p["grade_specs"], "pack_size": p["pack_size"],
            "mrp": p["mrp"], "dlp": p["dlp"],
            "distributor_margin_pct": p["distributor_margin_pct"],
        } for p in products[:8]]
        return {
            "ok": parsed["product_count"] > 0,
            "source": parsed["source"],
            "product_count": parsed["product_count"],
            "category_count": parsed["category_count"],
            "categories": categories[:30],
            "sample": sample,
            "warnings": parsed.get("warnings", []),
        }

    @router.post("/owner/products/import-circular")
    async def products_import_circular(
        file: UploadFile = File(...),
        title: Optional[str] = Form(None),
        effective_date: Optional[str] = Form(None),
        user: dict = Depends(owner_or_accountant),
    ):
        """Import products + pricing from a GO OIL price-list file (xlsx/csv/pdf).

        Categories appear as full-width header rows; each product row carries
        MRP, DLP, distributor margin, cash coupon, FOC, monthly gift, trade
        discount. Creates/updates products and publishes a new Price Circular.
        """
        import dms_price_import
        raw = await file.read()
        try:
            parsed = dms_price_import.parse_price_list(raw, file.filename or "")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read the file: {e}")

        if parsed["product_count"] == 0:
            msg = " ".join(parsed.get("warnings") or []) or \
                "No product rows detected. Make sure the header row has MATERIAL DESCRIPTION and DLP/MRP columns."
            raise HTTPException(status_code=400, detail=msg)

        existing_skus: set = set()
        async for p in db.dms_products.find({}, {"_id": 0, "sku_code": 1}):
            if p.get("sku_code"):
                existing_skus.add(p["sku_code"])
        cats = {c["name"].strip().lower(): c["id"]
                async for c in db.dms_categories.find({}, {"_id": 0, "name": 1, "id": 1})}

        current_cat_id: Optional[str] = None
        created, updated = 0, 0
        circular_lines: List[Dict[str, Any]] = []

        async def _ensure_cat(name: str) -> str:
            key = name.strip().lower()
            cid = cats.get(key)
            if not cid:
                cid = _nid("cat")
                await db.dms_categories.insert_one(
                    {"id": cid, "name": name.strip(), "created_at": _now()})
                cats[key] = cid
            return cid

        for it in parsed["items"]:
            if it["type"] == "category":
                current_cat_id = await _ensure_cat(it["name"])
                continue
            if not current_cat_id:
                current_cat_id = await _ensure_cat("Uncategorized")

            md = it["material_description"]
            grade = it["grade_specs"]
            pack = it["pack_size"]
            name = " ".join(x for x in [md, grade, pack] if x).strip() or md
            dlp = float(it["dlp"] or 0)

            existing = await db.dms_products.find_one({
                "material_description": md, "grade_specs": grade,
                "pack_size": pack, "category_id": current_cat_id}, {"_id": 0})
            if existing:
                pid = existing["id"]
                await db.dms_products.update_one({"id": pid}, {"$set": {
                    "name": name, "material_description": md, "grade_specs": grade,
                    "pack_size": pack, "mrp": _round(it.get("mrp") or 0),
                    "active": True, "updated_at": _now()}})
                updated += 1
            else:
                pid = _nid("prd")
                sku = dms_price_import.make_sku(md, pack, existing_skus)
                await db.dms_products.insert_one({
                    "id": pid, "name": name, "category_id": current_cat_id,
                    "sku_code": sku, "material_description": md,
                    "grade_specs": grade, "pack_size": pack, "description": "",
                    "box_qty": 1, "unit_price": _round(dlp), "previous_price": None,
                    "mrp": _round(it.get("mrp") or 0), "hsn": "", "gst_pct": 0.0,
                    "coupons_per_box": 100, "points_value": 10,
                    "active": True, "created_at": _now(),
                })
                await db.dms_price_batches.insert_one({
                    "id": _nid("pb"), "product_id": pid, "price": _round(dlp),
                    "from_date": _now(), "to_date": None, "created_at": _now()})
                created += 1

            circular_lines.append({
                "product_id": pid, "mrp": it["mrp"], "dlp": dlp,
                "distributor_margin_pct": it["distributor_margin_pct"],
                "cash_coupon": it["cash_coupon"], "foc_benefits": it["foc_benefits"],
                "monthly_gift": it["monthly_gift"],
                "trade_discount": str(it.get("trade_discount") or ""),
            })

        circular = None
        if circular_lines:
            ttl = (title or "").strip() or f"Imported — {(file.filename or 'price list')}"
            eff = (effective_date or "").strip() or _now()[:10]
            circular = await _publish_circular(
                ttl, eff, "Imported via file upload", circular_lines, user.get("id"))

        return {
            "ok": True,
            "created": created,
            "updated": updated,
            "categories": parsed["category_count"],
            "products_parsed": parsed["product_count"],
            "source": parsed["source"],
            "circular_id": (circular or {}).get("id"),
            "circular_batch_no": (circular or {}).get("batch_no"),
            "warnings": parsed.get("warnings", []),
        }

    @router.get("/products/{pid}/circular-history")
    async def product_circular_history(pid: str, user: dict = Depends(get_current_user)):
        """Full pricing history from all circulars for one product (latest first)."""
        rows = await db.dms_price_circulars.find(
            {"kind": "line", "product_id": pid}, {"_id": 0}
        ).sort([("batch_no", -1)]).to_list(500)
        # attach circular title
        cids = list({r["circular_id"] for r in rows})
        headers = await db.dms_price_circulars.find(
            {"id": {"$in": cids}, "kind": {"$ne": "line"}},
            {"_id": 0, "id": 1, "title": 1, "batch_label": 1}
        ).to_list(500)
        hmap = {h["id"]: h for h in headers}
        for r in rows:
            h = hmap.get(r["circular_id"], {})
            r["circular_title"] = h.get("title", "")
            r["batch_label"] = h.get("batch_label", "")
        return {"data": rows}

    @router.get("/price-circulars/{cid}/active-lines")
    async def circular_active_lines(cid: str, user: dict = Depends(get_current_user)):
        """Helper — for a specific circular, which lines are still the *current* active price."""
        lines = await db.dms_price_circulars.find(
            {"kind": "line", "circular_id": cid, "is_active": True}, {"_id": 0}
        ).to_list(2000)
        return {"data": lines, "count": len(lines)}

    # =========================================================================
    # PHASE 2B — CASH & BANK MANAGEMENT (standalone; no auto-link to payments)
    # =========================================================================
    owner_or_owner_acct = _guard("owner", "owner_accountant")

    async def _stop_sale_enabled() -> bool:
        s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0, "stop_sale_on_negative": 1}) or {}
        # default True
        v = s.get("stop_sale_on_negative")
        return True if v is None else bool(v)

    # -- Bank Accounts --
    @router.get("/bank-accounts")
    async def list_bank_accounts(user: dict = Depends(owner_or_owner_acct)):
        rows = await db.dms_bank_accounts.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
        return {"data": rows, "count": len(rows)}

    @router.post("/bank-accounts")
    async def create_bank_account(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        name = str(body.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="name required")
        try:
            opening = float(body.get("opening_balance") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="opening_balance must be a number")
        doc = {
            "id": _nid("bank"),
            "name": name,
            "account_number": str(body.get("account_number") or "").strip(),
            "ifsc": str(body.get("ifsc") or "").strip(),
            "branch": str(body.get("branch") or "").strip(),
            "opening_balance": _round(opening),
            "current_balance": _round(opening),
            "notes": str(body.get("notes") or "").strip(),
            "active": True,
            "created_by": user["id"],
            "created_at": _now(),
        }
        await db.dms_bank_accounts.insert_one(doc)
        doc.pop("_id", None)
        return _clean(doc)

    @router.put("/bank-accounts/{bid}")
    async def update_bank_account(bid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        cur = await db.dms_bank_accounts.find_one({"id": bid}, {"_id": 0})
        if not cur:
            raise HTTPException(status_code=404, detail="Bank account not found")
        upd: Dict[str, Any] = {}
        for k in ("name", "account_number", "ifsc", "branch", "notes"):
            if k in body:
                upd[k] = str(body.get(k) or "").strip()
        if "active" in body:
            upd["active"] = bool(body.get("active"))
        upd["updated_at"] = _now()
        await db.dms_bank_accounts.update_one({"id": bid}, {"$set": upd})
        return {"ok": True}

    @router.delete("/bank-accounts/{bid}")
    async def delete_bank_account(bid: str, user: dict = Depends(owner_only)):
        # only allow delete if no transactions
        c = await db.dms_bank_transactions.count_documents({"bank_account_id": bid})
        if c > 0:
            raise HTTPException(status_code=400, detail=f"Cannot delete: {c} transactions exist. Deactivate instead.")
        await db.dms_bank_accounts.delete_one({"id": bid})
        return {"ok": True}

    # -- Bank Transactions --
    @router.get("/bank-transactions")
    async def list_bank_transactions(user: dict = Depends(owner_or_owner_acct), account_id: Optional[str] = None,
                                     start: Optional[str] = None, end: Optional[str] = None, type: Optional[str] = None):
        q: Dict[str, Any] = {}
        if account_id: q["bank_account_id"] = account_id
        if type: q["type"] = type
        if start: q.setdefault("date", {})["$gte"] = start
        if end: q.setdefault("date", {})["$lte"] = end
        rows = await db.dms_bank_transactions.find(q, {"_id": 0}).sort("date", -1).to_list(2000)
        return {"data": rows, "count": len(rows)}

    @router.post("/bank-transactions")
    async def create_bank_transaction(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        aid = str(body.get("bank_account_id") or "").strip()
        acct = await db.dms_bank_accounts.find_one({"id": aid}, {"_id": 0})
        if not acct:
            raise HTTPException(status_code=400, detail="bank_account_id invalid")
        typ = str(body.get("type") or "").strip().lower()
        if typ not in ("deposit", "withdrawal"):
            raise HTTPException(status_code=400, detail="type must be deposit or withdrawal")
        try:
            amt = float(body.get("amount") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="amount must be a number")
        if amt <= 0:
            raise HTTPException(status_code=400, detail="amount must be > 0")
        date = str(body.get("date") or _now()[:10]).strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        await _check_fy_lock(date, "bank transaction")
        # compute new balance
        delta = amt if typ == "deposit" else -amt
        new_balance = _round(float(acct.get("current_balance") or 0) + delta)
        doc = {
            "id": _nid("btx"),
            "bank_account_id": aid,
            "bank_account_name": acct.get("name"),
            "date": date,
            "type": typ,
            "amount": _round(amt),
            "reference": str(body.get("reference") or "").strip(),
            "notes": str(body.get("notes") or "").strip(),
            "balance_after": new_balance,
            "created_by": user["id"],
            "created_by_name": user.get("name"),
            "created_at": _now(),
        }
        await db.dms_bank_transactions.insert_one(doc)
        await db.dms_bank_accounts.update_one({"id": aid}, {"$set": {"current_balance": new_balance, "updated_at": _now()}})
        doc.pop("_id", None)
        return _clean(doc)

    @router.delete("/bank-transactions/{tid}")
    async def delete_bank_transaction(tid: str, user: dict = Depends(owner_only)):
        tx = await db.dms_bank_transactions.find_one({"id": tid}, {"_id": 0})
        if not tx:
            raise HTTPException(status_code=404, detail="Transaction not found")
        await _check_fy_lock(tx.get("date"), "bank transaction")
        # reverse balance
        delta = float(tx["amount"]) if tx["type"] == "deposit" else -float(tx["amount"])
        acct = await db.dms_bank_accounts.find_one({"id": tx["bank_account_id"]}, {"_id": 0})
        if acct:
            new_balance = _round(float(acct.get("current_balance") or 0) - delta)
            await db.dms_bank_accounts.update_one({"id": tx["bank_account_id"]}, {"$set": {"current_balance": new_balance}})
        await db.dms_bank_transactions.delete_one({"id": tid})
        return {"ok": True}

    # -- Cash Register --
    @router.get("/cash-register")
    async def list_cash_register(user: dict = Depends(owner_or_owner_acct), start: Optional[str] = None,
                                 end: Optional[str] = None, type: Optional[str] = None):
        q: Dict[str, Any] = {}
        if type: q["type"] = type
        if start: q.setdefault("date", {})["$gte"] = start
        if end: q.setdefault("date", {})["$lte"] = end
        rows = await db.dms_cash_register.find(q, {"_id": 0}).sort("date", -1).to_list(2000)
        # compute current balance
        agg = await db.dms_cash_register.aggregate([
            {"$group": {"_id": "$type", "total": {"$sum": "$amount"}}}
        ]).to_list(10)
        totals = {r["_id"]: r["total"] for r in agg}
        current_balance = _round((totals.get("in") or 0) - (totals.get("out") or 0))
        return {"data": rows, "count": len(rows), "current_balance": current_balance}

    @router.post("/cash-register")
    async def create_cash_entry(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        typ = str(body.get("type") or "").strip().lower()
        if typ not in ("in", "out"):
            raise HTTPException(status_code=400, detail="type must be 'in' or 'out'")
        try:
            amt = float(body.get("amount") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="amount must be a number")
        if amt <= 0:
            raise HTTPException(status_code=400, detail="amount must be > 0")
        date = str(body.get("date") or _now()[:10]).strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        await _check_fy_lock(date, "cash entry")
        # compute new balance
        agg = await db.dms_cash_register.aggregate([
            {"$group": {"_id": "$type", "total": {"$sum": "$amount"}}}
        ]).to_list(10)
        totals = {r["_id"]: r["total"] for r in agg}
        current_balance = (totals.get("in") or 0) - (totals.get("out") or 0)
        delta = amt if typ == "in" else -amt
        new_balance = _round(current_balance + delta)
        doc = {
            "id": _nid("cash"),
            "date": date,
            "type": typ,
            "amount": _round(amt),
            "reference": str(body.get("reference") or "").strip(),
            "notes": str(body.get("notes") or "").strip(),
            "balance_after": new_balance,
            "created_by": user["id"],
            "created_by_name": user.get("name"),
            "created_at": _now(),
        }
        await db.dms_cash_register.insert_one(doc)
        doc.pop("_id", None)
        return _clean(doc)

    @router.delete("/cash-register/{cid}")
    async def delete_cash_entry(cid: str, user: dict = Depends(owner_only)):
        cur = await db.dms_cash_register.find_one({"id": cid}, {"_id": 0})
        if not cur:
            raise HTTPException(status_code=404, detail="Entry not found")
        await _check_fy_lock(cur.get("date"), "cash entry")
        await db.dms_cash_register.delete_one({"id": cid})
        return {"ok": True}

    # -- Cheques --
    @router.get("/cheques")
    async def list_cheques(user: dict = Depends(owner_or_owner_acct), direction: Optional[str] = None,
                           status: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None):
        q: Dict[str, Any] = {}
        if direction: q["direction"] = direction
        if status: q["status"] = status
        if start: q.setdefault("date", {})["$gte"] = start
        if end: q.setdefault("date", {})["$lte"] = end
        rows = await db.dms_cheques.find(q, {"_id": 0}).sort("date", -1).to_list(2000)
        return {"data": rows, "count": len(rows)}

    @router.post("/cheques")
    async def create_cheque(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        cheque_no = str(body.get("cheque_no") or "").strip()
        if not cheque_no:
            raise HTTPException(status_code=400, detail="cheque_no required")
        direction = str(body.get("direction") or "").strip().lower()
        if direction not in ("received", "issued"):
            raise HTTPException(status_code=400, detail="direction must be received or issued")
        try:
            amt = float(body.get("amount") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="amount must be a number")
        if amt <= 0:
            raise HTTPException(status_code=400, detail="amount must be > 0")
        date = str(body.get("date") or _now()[:10]).strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        await _check_fy_lock(date, "cheque")
        status_val = str(body.get("status") or "pending").strip().lower()
        if status_val not in ("pending", "cleared", "bounced", "cancelled"):
            raise HTTPException(status_code=400, detail="status must be pending/cleared/bounced/cancelled")
        doc = {
            "id": _nid("chq"),
            "cheque_no": cheque_no,
            "date": date,
            "direction": direction,
            "party_name": str(body.get("party_name") or "").strip(),
            "amount": _round(amt),
            "bank_name": str(body.get("bank_name") or "").strip(),
            "status": status_val,
            "notes": str(body.get("notes") or "").strip(),
            "created_by": user["id"],
            "created_by_name": user.get("name"),
            "created_at": _now(),
        }
        await db.dms_cheques.insert_one(doc)
        doc.pop("_id", None)
        return _clean(doc)

    @router.put("/cheques/{cqid}")
    async def update_cheque(cqid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        cur = await db.dms_cheques.find_one({"id": cqid}, {"_id": 0})
        if not cur:
            raise HTTPException(status_code=404, detail="Cheque not found")
        await _check_fy_lock(cur.get("date"), "cheque")
        upd: Dict[str, Any] = {}
        for k in ("cheque_no", "party_name", "bank_name", "notes"):
            if k in body:
                upd[k] = str(body.get(k) or "").strip()
        if "amount" in body:
            try:
                amt = float(body["amount"])
            except Exception:
                raise HTTPException(status_code=400, detail="amount must be a number")
            if amt <= 0:
                raise HTTPException(status_code=400, detail="amount must be > 0")
            upd["amount"] = _round(amt)
        if "status" in body:
            sv = str(body.get("status") or "").strip().lower()
            if sv not in ("pending", "cleared", "bounced", "cancelled"):
                raise HTTPException(status_code=400, detail="invalid status")
            upd["status"] = sv
        if "date" in body:
            d = str(body.get("date") or "").strip()
            try:
                datetime.strptime(d, "%Y-%m-%d")
            except Exception:
                raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
            await _check_fy_lock(d, "cheque")
            upd["date"] = d
        upd["updated_at"] = _now()
        await db.dms_cheques.update_one({"id": cqid}, {"$set": upd})
        return {"ok": True}

    @router.delete("/cheques/{cqid}")
    async def delete_cheque(cqid: str, user: dict = Depends(owner_only)):
        cur = await db.dms_cheques.find_one({"id": cqid}, {"_id": 0})
        if not cur:
            raise HTTPException(status_code=404, detail="Cheque not found")
        await _check_fy_lock(cur.get("date"), "cheque")
        await db.dms_cheques.delete_one({"id": cqid})
        return {"ok": True}

    # -- Loan Accounts --
    @router.get("/loan-accounts")
    async def list_loans(user: dict = Depends(owner_or_owner_acct)):
        rows = await db.dms_loan_accounts.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
        return {"data": rows, "count": len(rows)}

    @router.post("/loan-accounts")
    async def create_loan(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        name = str(body.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="name required")
        try:
            principal = float(body.get("principal") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="principal must be a number")
        if principal <= 0:
            raise HTTPException(status_code=400, detail="principal must be > 0")
        try:
            rate = float(body.get("interest_rate") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="interest_rate must be a number")
        start_date = str(body.get("start_date") or _now()[:10]).strip()
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="start_date must be YYYY-MM-DD")
        try:
            tenure = int(body.get("tenure_months") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="tenure_months must be integer")
        doc = {
            "id": _nid("loan"),
            "name": name,
            "lender_name": str(body.get("lender_name") or "").strip(),
            "principal": _round(principal),
            "interest_rate": _round(rate),
            "start_date": start_date,
            "tenure_months": tenure,
            "outstanding": _round(principal),  # start with principal
            "notes": str(body.get("notes") or "").strip(),
            "active": True,
            "created_by": user["id"],
            "created_at": _now(),
        }
        await db.dms_loan_accounts.insert_one(doc)
        # auto-log a disbursement txn
        await db.dms_loan_transactions.insert_one({
            "id": _nid("lntx"),
            "loan_account_id": doc["id"],
            "date": start_date,
            "type": "disbursement",
            "amount": _round(principal),
            "notes": "Initial disbursement",
            "outstanding_after": _round(principal),
            "created_by": user["id"],
            "created_at": _now(),
        })
        doc.pop("_id", None)
        return _clean(doc)

    @router.put("/loan-accounts/{lid}")
    async def update_loan(lid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        cur = await db.dms_loan_accounts.find_one({"id": lid}, {"_id": 0})
        if not cur:
            raise HTTPException(status_code=404, detail="Loan not found")
        upd: Dict[str, Any] = {}
        for k in ("name", "lender_name", "notes"):
            if k in body:
                upd[k] = str(body.get(k) or "").strip()
        if "interest_rate" in body:
            try:
                upd["interest_rate"] = _round(float(body["interest_rate"]))
            except Exception:
                raise HTTPException(status_code=400, detail="interest_rate must be a number")
        if "tenure_months" in body:
            try:
                upd["tenure_months"] = int(body["tenure_months"])
            except Exception:
                raise HTTPException(status_code=400, detail="tenure_months must be integer")
        if "active" in body:
            upd["active"] = bool(body["active"])
        upd["updated_at"] = _now()
        await db.dms_loan_accounts.update_one({"id": lid}, {"$set": upd})
        return {"ok": True}

    @router.delete("/loan-accounts/{lid}")
    async def delete_loan(lid: str, user: dict = Depends(owner_only)):
        c = await db.dms_loan_transactions.count_documents({"loan_account_id": lid})
        if c > 1:  # allow deletion if only initial disbursement
            raise HTTPException(status_code=400, detail=f"Cannot delete: {c} transactions exist")
        await db.dms_loan_transactions.delete_many({"loan_account_id": lid})
        await db.dms_loan_accounts.delete_one({"id": lid})
        return {"ok": True}

    # -- Loan Transactions --
    @router.get("/loan-transactions")
    async def list_loan_transactions(user: dict = Depends(owner_or_owner_acct), loan_id: Optional[str] = None):
        q: Dict[str, Any] = {}
        if loan_id: q["loan_account_id"] = loan_id
        rows = await db.dms_loan_transactions.find(q, {"_id": 0}).sort("date", -1).to_list(2000)
        return {"data": rows, "count": len(rows)}

    @router.post("/loan-transactions")
    async def create_loan_transaction(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        lid = str(body.get("loan_account_id") or "").strip()
        loan = await db.dms_loan_accounts.find_one({"id": lid}, {"_id": 0})
        if not loan:
            raise HTTPException(status_code=400, detail="loan_account_id invalid")
        typ = str(body.get("type") or "").strip().lower()
        if typ not in ("disbursement", "repayment", "interest"):
            raise HTTPException(status_code=400, detail="type must be disbursement/repayment/interest")
        try:
            amt = float(body.get("amount") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="amount must be a number")
        if amt <= 0:
            raise HTTPException(status_code=400, detail="amount must be > 0")
        date = str(body.get("date") or _now()[:10]).strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        await _check_fy_lock(date, "loan transaction")
        # outstanding: disbursement/interest increase; repayment decreases
        delta = -amt if typ == "repayment" else amt
        new_outstanding = _round(float(loan.get("outstanding") or 0) + delta)
        doc = {
            "id": _nid("lntx"),
            "loan_account_id": lid,
            "date": date,
            "type": typ,
            "amount": _round(amt),
            "notes": str(body.get("notes") or "").strip(),
            "outstanding_after": new_outstanding,
            "created_by": user["id"],
            "created_by_name": user.get("name"),
            "created_at": _now(),
        }
        await db.dms_loan_transactions.insert_one(doc)
        await db.dms_loan_accounts.update_one({"id": lid}, {"$set": {"outstanding": new_outstanding, "updated_at": _now()}})
        doc.pop("_id", None)
        return _clean(doc)

    # =========================================================================
    # PHASE 2B — GODOWN MANAGEMENT + INVENTORY
    # =========================================================================

    async def _adjust_godown_stock(godown_id: str, product_id: str, delta_boxes: int, reason: str, ref: str = ""):
        cur = await db.dms_godown_inventory.find_one({"godown_id": godown_id, "product_id": product_id})
        if cur:
            new_qty = int(cur.get("qty_boxes", 0)) + int(delta_boxes)
            await db.dms_godown_inventory.update_one(
                {"godown_id": godown_id, "product_id": product_id},
                {"$set": {"qty_boxes": max(new_qty, 0), "updated_at": _now()}},
            )
        else:
            await db.dms_godown_inventory.insert_one({
                "id": _nid("ginv"),
                "godown_id": godown_id,
                "product_id": product_id,
                "qty_boxes": max(int(delta_boxes), 0),
                "updated_at": _now(),
            })
        await db.dms_stock_ledger.insert_one({
            "id": _nid("sl"),
            "scope": "godown",
            "godown_id": godown_id,
            "product_id": product_id,
            "delta_boxes": int(delta_boxes),
            "reason": reason,
            "reference": ref,
            "at": _now(),
        })

    async def _get_godown_stock(godown_id: str, product_id: str) -> int:
        doc = await db.dms_godown_inventory.find_one({"godown_id": godown_id, "product_id": product_id}, {"_id": 0, "qty_boxes": 1})
        return int(doc.get("qty_boxes", 0)) if doc else 0

    @router.get("/godowns")
    async def list_godowns(user: dict = Depends(owner_or_owner_acct)):
        rows = await db.dms_godowns.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
        # attach total_boxes per godown
        for r in rows:
            agg = await db.dms_godown_inventory.aggregate([
                {"$match": {"godown_id": r["id"]}},
                {"$group": {"_id": None, "total_boxes": {"$sum": "$qty_boxes"}}},
            ]).to_list(2)
            r["total_boxes"] = int((agg[0]["total_boxes"] if agg else 0) or 0)
        return {"data": rows, "count": len(rows)}

    @router.post("/godowns")
    async def create_godown(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        name = str(body.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="name required")
        doc = {
            "id": _nid("gdn"),
            "name": name,
            "address": str(body.get("address") or "").strip(),
            "manager_name": str(body.get("manager_name") or "").strip(),
            "phone": str(body.get("phone") or "").strip(),
            "capacity_boxes": int(body.get("capacity_boxes") or 0),
            "notes": str(body.get("notes") or "").strip(),
            "active": True,
            "created_by": user["id"],
            "created_at": _now(),
        }
        await db.dms_godowns.insert_one(doc)
        doc.pop("_id", None)
        return _clean(doc)

    @router.put("/godowns/{gid}")
    async def update_godown(gid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        cur = await db.dms_godowns.find_one({"id": gid}, {"_id": 0})
        if not cur:
            raise HTTPException(status_code=404, detail="Godown not found")
        upd: Dict[str, Any] = {}
        for k in ("name", "address", "manager_name", "phone", "notes"):
            if k in body:
                upd[k] = str(body.get(k) or "").strip()
        if "capacity_boxes" in body:
            try:
                upd["capacity_boxes"] = int(body["capacity_boxes"])
            except Exception:
                raise HTTPException(status_code=400, detail="capacity_boxes must be integer")
        if "active" in body:
            upd["active"] = bool(body["active"])
        upd["updated_at"] = _now()
        await db.dms_godowns.update_one({"id": gid}, {"$set": upd})
        return {"ok": True}

    @router.delete("/godowns/{gid}")
    async def delete_godown(gid: str, user: dict = Depends(owner_only)):
        # only allow delete if no stock
        agg = await db.dms_godown_inventory.aggregate([
            {"$match": {"godown_id": gid}},
            {"$group": {"_id": None, "total_boxes": {"$sum": "$qty_boxes"}}},
        ]).to_list(2)
        total = int((agg[0]["total_boxes"] if agg else 0) or 0)
        if total > 0:
            raise HTTPException(status_code=400, detail=f"Cannot delete: {total} boxes in godown. Transfer stock first.")
        await db.dms_godown_inventory.delete_many({"godown_id": gid})
        await db.dms_godowns.delete_one({"id": gid})
        return {"ok": True}

    @router.get("/godowns/{gid}/inventory")
    async def get_godown_inventory(gid: str, user: dict = Depends(owner_or_owner_acct)):
        godown = await db.dms_godowns.find_one({"id": gid}, {"_id": 0})
        if not godown:
            raise HTTPException(status_code=404, detail="Godown not found")
        rows = await db.dms_godown_inventory.find({"godown_id": gid}, {"_id": 0}).to_list(2000)
        pids = [r["product_id"] for r in rows]
        prods = {p["id"]: p async for p in db.dms_products.find({"id": {"$in": pids}}, {"_id": 0})}
        for r in rows:
            p = prods.get(r["product_id"], {})
            r["product_name"] = p.get("name", "")
            r["sku_code"] = p.get("sku_code", "")
            r["material_description"] = p.get("material_description", "")
            r["pack_size"] = p.get("pack_size", "")
            r["unit_price"] = p.get("unit_price", 0)
            r["value"] = _round((r.get("qty_boxes", 0) or 0) * (p.get("unit_price", 0) or 0))
            # Phase 2C: low-stock flag
            rl = int(r.get("reorder_level_boxes") or 0)
            r["reorder_level_boxes"] = rl
            r["low_stock"] = bool(rl > 0 and int(r.get("qty_boxes", 0) or 0) <= rl)
        rows.sort(key=lambda x: x.get("product_name", ""))
        return {
            "godown": godown,
            "data": rows,
            "total_value": _round(sum(r.get("value", 0) for r in rows)),
            "total_boxes": int(sum(r.get("qty_boxes", 0) for r in rows)),
        }

    # =========================================================================
    # PHASE 2B — STOCK TRANSFER (owner <-> godown, godown <-> godown)
    # =========================================================================
    @router.get("/stock-transfers")
    async def list_stock_transfers(user: dict = Depends(owner_or_owner_acct), start: Optional[str] = None, end: Optional[str] = None):
        q: Dict[str, Any] = {}
        if start: q.setdefault("date", {})["$gte"] = start
        if end: q.setdefault("date", {})["$lte"] = end
        rows = await db.dms_stock_transfers.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
        return {"data": rows, "count": len(rows)}

    @router.get("/stock-transfers/{tid}")
    async def get_stock_transfer(tid: str, user: dict = Depends(owner_or_owner_acct)):
        row = await db.dms_stock_transfers.find_one({"id": tid}, {"_id": 0})
        if not row:
            raise HTTPException(status_code=404, detail="Transfer not found")
        return row

    @router.post("/stock-transfers")
    async def create_stock_transfer(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        """Body: {date, from_type: 'owner'|'godown', from_godown_id?, to_type: 'godown'|'owner', to_godown_id?, items: [{product_id, qty_boxes}], notes}"""
        date = str(body.get("date") or _now()[:10]).strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        await _check_fy_lock(date, "stock transfer")
        from_type = str(body.get("from_type") or "").strip().lower()
        to_type = str(body.get("to_type") or "").strip().lower()
        if from_type not in ("owner", "godown") or to_type not in ("owner", "godown"):
            raise HTTPException(status_code=400, detail="from_type/to_type must be owner or godown")
        if from_type == "owner" and to_type == "owner":
            raise HTTPException(status_code=400, detail="Source and destination cannot both be owner")
        from_gid = str(body.get("from_godown_id") or "").strip() or None
        to_gid = str(body.get("to_godown_id") or "").strip() or None
        if from_type == "godown" and not from_gid:
            raise HTTPException(status_code=400, detail="from_godown_id required")
        if to_type == "godown" and not to_gid:
            raise HTTPException(status_code=400, detail="to_godown_id required")
        if from_type == "godown" and to_type == "godown" and from_gid == to_gid:
            raise HTTPException(status_code=400, detail="Source and destination godowns cannot be the same")
        # verify godowns exist
        from_godown = None; to_godown = None
        if from_gid:
            from_godown = await db.dms_godowns.find_one({"id": from_gid}, {"_id": 0})
            if not from_godown:
                raise HTTPException(status_code=400, detail="from_godown_id invalid")
        if to_gid:
            to_godown = await db.dms_godowns.find_one({"id": to_gid}, {"_id": 0})
            if not to_godown:
                raise HTTPException(status_code=400, detail="to_godown_id invalid")
        items = body.get("items") or []
        if not items:
            raise HTTPException(status_code=400, detail="items required")
        # normalize + validate stock at source
        norm_items: List[Dict[str, Any]] = []
        pids = [str(i.get("product_id") or "").strip() for i in items if i.get("product_id")]
        prods = {p["id"]: p async for p in db.dms_products.find({"id": {"$in": pids}}, {"_id": 0})}
        for it in items:
            pid = str(it.get("product_id") or "").strip()
            try:
                qty = int(it.get("qty_boxes") or 0)
            except Exception:
                raise HTTPException(status_code=400, detail="qty_boxes must be integer")
            if qty <= 0:
                raise HTTPException(status_code=400, detail=f"qty_boxes must be > 0 for {pid}")
            p = prods.get(pid)
            if not p:
                raise HTTPException(status_code=400, detail=f"product_id invalid: {pid}")
            # verify source has enough stock
            if from_type == "owner":
                avail = await _get_owner_stock(pid)
            else:
                avail = await _get_godown_stock(from_gid, pid)
            if qty > avail:
                raise HTTPException(status_code=400, detail=f"Insufficient stock for {p.get('name', pid)}: available {avail}, requested {qty}")
            norm_items.append({
                "product_id": pid,
                "product_name": p.get("name", ""),
                "sku_code": p.get("sku_code", ""),
                "material_description": p.get("material_description", ""),
                "pack_size": p.get("pack_size", ""),
                "qty_boxes": qty,
            })
        # Build transfer_no like ST-{yymmdd}-{count+1}
        cnt = await db.dms_stock_transfers.count_documents({})
        transfer_no = f"ST-{datetime.now().strftime('%y%m%d')}-{cnt + 1:04d}"
        transfer = {
            "id": _nid("stx"),
            "transfer_no": transfer_no,
            "date": date,
            "from_type": from_type,
            "from_godown_id": from_gid,
            "from_godown_name": (from_godown or {}).get("name") if from_godown else "Owner Warehouse",
            "to_type": to_type,
            "to_godown_id": to_gid,
            "to_godown_name": (to_godown or {}).get("name") if to_godown else "Owner Warehouse",
            "items": norm_items,
            "total_boxes": int(sum(i["qty_boxes"] for i in norm_items)),
            "notes": str(body.get("notes") or "").strip(),
            "created_by": user["id"],
            "created_by_name": user.get("name"),
            "created_at": _now(),
        }
        await db.dms_stock_transfers.insert_one(transfer)
        # apply movements
        for it in norm_items:
            if from_type == "owner":
                await _adjust_owner_stock(it["product_id"], -it["qty_boxes"], "stock_transfer_out", transfer_no)
            else:
                await _adjust_godown_stock(from_gid, it["product_id"], -it["qty_boxes"], "stock_transfer_out", transfer_no)
            if to_type == "owner":
                await _adjust_owner_stock(it["product_id"], it["qty_boxes"], "stock_transfer_in", transfer_no)
            else:
                await _adjust_godown_stock(to_gid, it["product_id"], it["qty_boxes"], "stock_transfer_in", transfer_no)
        transfer.pop("_id", None)
        return _clean(transfer)

    # =========================================================================
    # PHASE 2B — settings endpoint extension for stop_sale_on_negative
    # =========================================================================
    @router.put("/settings/stop-sale")
    async def toggle_stop_sale(body: Dict[str, Any] = Body(...), user: dict = Depends(owner_only)):
        v = bool(body.get("enabled", True))
        await db.dms_settings.update_one({"id": "global"}, {"$set": {"stop_sale_on_negative": v, "updated_at": _now()}}, upsert=True)
        return {"ok": True, "stop_sale_on_negative": v}

    # =========================================================================
    # PHASE 2C — Import / Export (Parties + Items importable; Sales + Payments export-only)
    # =========================================================================
    def _xlsx_response(wb, filename: str):
        from io import BytesIO
        from fastapi.responses import Response
        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'},
        )

    # -- Parties Export --
    @router.get("/parties/export")
    async def export_parties(user: dict = Depends(owner_or_accountant)):
        from openpyxl import Workbook
        wb = Workbook()
        # Distributors sheet
        ws1 = wb.active; ws1.title = "Distributors"
        ws1.append(["name", "code", "email", "phone", "address", "gstin", "credit_limit", "active"])
        async for d in db.dms_distributors.find({}, {"_id": 0}):
            ws1.append([d.get("name"), d.get("code", ""), d.get("email", ""), d.get("phone", ""),
                        d.get("address", ""), d.get("gstin", ""), d.get("credit_limit", 0), bool(d.get("active", True))])
        # Retailers sheet
        ws2 = wb.create_sheet("Retailers")
        ws2.append(["name", "code", "email", "phone", "address", "gstin", "distributor_email", "active"])
        d_by_id = {d["id"]: d async for d in db.dms_distributors.find({}, {"_id": 0, "id": 1, "email": 1})}
        async for r in db.dms_retailers.find({}, {"_id": 0}):
            d = d_by_id.get(r.get("distributor_id"), {})
            ws2.append([r.get("name"), r.get("code", ""), r.get("email", ""), r.get("phone", ""),
                        r.get("address", ""), r.get("gstin", ""), d.get("email", ""), bool(r.get("active", True))])
        for w in [ws1, ws2]:
            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                w.column_dimensions[col_letter].width = 22
        return _xlsx_response(wb, "parties")

    # -- Parties Import --
    @router.post("/parties/import")
    async def import_parties(file: UploadFile = File(...), user: dict = Depends(owner_or_accountant)):
        """Two sheets: 'Distributors' and 'Retailers'. See export for column layout."""
        from openpyxl import load_workbook
        from io import BytesIO
        raw = await file.read()
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read xlsx: {e}")
        summary = {"distributors": {"created": 0, "updated": 0, "skipped": 0, "errors": []},
                   "retailers": {"created": 0, "updated": 0, "skipped": 0, "errors": []}}

        # Distributors
        if "Distributors" in wb.sheetnames:
            ws = wb["Distributors"]
            header = [str(h or "").strip().lower() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            col = {h: i for i, h in enumerate(header)}
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(v is None or v == "" for v in row):
                    continue
                try:
                    name = str(row[col.get("name", -1)] or "").strip()
                    email = str(row[col.get("email", -1)] or "").strip().lower()
                    if not name:
                        summary["distributors"]["skipped"] += 1
                        summary["distributors"]["errors"].append(f"Row {i}: missing name")
                        continue
                    doc = {
                        "name": name,
                        "code": str(row[col.get("code", -1)] or "").strip(),
                        "email": email,
                        "phone": str(row[col.get("phone", -1)] or "").strip(),
                        "address": str(row[col.get("address", -1)] or "").strip(),
                        "gstin": str(row[col.get("gstin", -1)] or "").strip(),
                        "credit_limit": float(row[col.get("credit_limit", -1)] or 0) if col.get("credit_limit", -1) >= 0 else 0,
                        "active": bool(row[col.get("active", -1)]) if col.get("active", -1) >= 0 and row[col.get("active", -1)] is not None else True,
                    }
                    existing = None
                    if email:
                        existing = await db.dms_distributors.find_one({"email": email}, {"_id": 0, "id": 1})
                    if existing:
                        await db.dms_distributors.update_one({"id": existing["id"]}, {"$set": {**doc, "updated_at": _now()}})
                        summary["distributors"]["updated"] += 1
                    else:
                        doc["id"] = _nid("dist"); doc["created_at"] = _now()
                        await db.dms_distributors.insert_one(doc)
                        summary["distributors"]["created"] += 1
                except Exception as e:
                    summary["distributors"]["skipped"] += 1
                    summary["distributors"]["errors"].append(f"Row {i}: {e}")

        # Retailers
        if "Retailers" in wb.sheetnames:
            ws = wb["Retailers"]
            header = [str(h or "").strip().lower() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            col = {h: i for i, h in enumerate(header)}
            # cache distributor lookup by email
            d_by_email = {d["email"].lower(): d["id"] async for d in db.dms_distributors.find({"email": {"$ne": None}}, {"_id": 0, "id": 1, "email": 1})}
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(v is None or v == "" for v in row):
                    continue
                try:
                    name = str(row[col.get("name", -1)] or "").strip()
                    email = str(row[col.get("email", -1)] or "").strip().lower()
                    d_email = str(row[col.get("distributor_email", -1)] or "").strip().lower()
                    if not name or not d_email:
                        summary["retailers"]["skipped"] += 1
                        summary["retailers"]["errors"].append(f"Row {i}: missing name or distributor_email")
                        continue
                    d_id = d_by_email.get(d_email)
                    if not d_id:
                        summary["retailers"]["skipped"] += 1
                        summary["retailers"]["errors"].append(f"Row {i}: unknown distributor_email '{d_email}'")
                        continue
                    doc = {
                        "name": name,
                        "code": str(row[col.get("code", -1)] or "").strip(),
                        "email": email,
                        "phone": str(row[col.get("phone", -1)] or "").strip(),
                        "address": str(row[col.get("address", -1)] or "").strip(),
                        "gstin": str(row[col.get("gstin", -1)] or "").strip(),
                        "distributor_id": d_id,
                        "active": bool(row[col.get("active", -1)]) if col.get("active", -1) >= 0 and row[col.get("active", -1)] is not None else True,
                    }
                    existing = None
                    if email:
                        existing = await db.dms_retailers.find_one({"email": email}, {"_id": 0, "id": 1})
                    if existing:
                        await db.dms_retailers.update_one({"id": existing["id"]}, {"$set": {**doc, "updated_at": _now()}})
                        summary["retailers"]["updated"] += 1
                    else:
                        doc["id"] = _nid("ret"); doc["created_at"] = _now()
                        await db.dms_retailers.insert_one(doc)
                        summary["retailers"]["created"] += 1
                except Exception as e:
                    summary["retailers"]["skipped"] += 1
                    summary["retailers"]["errors"].append(f"Row {i}: {e}")

        return summary

    # -- Sale Bills Export (both primary e-bills and secondary retailer bills) --
    @router.get("/sale-bills/export")
    async def export_sale_bills(user: dict = Depends(owner_or_accountant), start: Optional[str] = None, end: Optional[str] = None):
        from openpyxl import Workbook
        wb = Workbook()
        # Primary e-bills sheet
        ws1 = wb.active; ws1.title = "Primary_eBills"
        ws1.append(["ebill_no", "date", "distributor", "order_no", "subtotal", "gst_total", "total", "status"])
        q = {}
        if start: q.setdefault("created_at", {})["$gte"] = start
        if end: q.setdefault("created_at", {})["$lte"] = end + "T23:59:59"
        async for eb in db.dms_ebills.find(q, {"_id": 0}):
            ws1.append([eb.get("ebill_no"), (eb.get("created_at") or "")[:10], eb.get("distributor_name"),
                        eb.get("order_no"), eb.get("subtotal"), eb.get("gst_total"), eb.get("total"), eb.get("status")])
        # Retailer bills sheet
        ws2 = wb.create_sheet("Retailer_Bills")
        ws2.append(["bill_no", "date", "distributor", "retailer", "order_no", "subtotal", "gst_total", "total", "status"])
        d_by_id = {d["id"]: d["name"] async for d in db.dms_distributors.find({}, {"_id": 0, "id": 1, "name": 1})}
        r_by_id = {r["id"]: r["name"] async for r in db.dms_retailers.find({}, {"_id": 0, "id": 1, "name": 1})}
        async for b in db.dms_retailer_bills.find(q, {"_id": 0}):
            ws2.append([b.get("bill_no"), (b.get("created_at") or "")[:10],
                        d_by_id.get(b.get("distributor_id"), ""),
                        r_by_id.get(b.get("retailer_id"), ""),
                        b.get("order_no", ""), b.get("subtotal"), b.get("gst_total"), b.get("total"), b.get("status")])
        for w in [ws1, ws2]:
            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
                w.column_dimensions[col_letter].width = 20
        return _xlsx_response(wb, "sale_bills")

    # -- Payments Export --
    @router.get("/payments/export")
    async def export_payments(user: dict = Depends(owner_or_accountant), start: Optional[str] = None, end: Optional[str] = None):
        from openpyxl import Workbook
        wb = Workbook()
        # Primary Payments (kind=payment in dms_primary_ledger)
        ws1 = wb.active; ws1.title = "Primary_Payments"
        ws1.append(["date", "distributor", "amount", "reference", "description"])
        q = {"kind": "payment"}
        if start: q.setdefault("at", {})["$gte"] = start
        if end: q.setdefault("at", {})["$lte"] = end + "T23:59:59"
        d_by_id = {d["id"]: d["name"] async for d in db.dms_distributors.find({}, {"_id": 0, "id": 1, "name": 1})}
        async for e in db.dms_primary_ledger.find(q, {"_id": 0}):
            ws1.append([(e.get("at") or "")[:10], d_by_id.get(e.get("distributor_id"), ""),
                        e.get("amount"), e.get("reference_no", ""), e.get("description", "")])
        # Secondary Payments (kind=payment in dms_retailer_ledger)
        ws2 = wb.create_sheet("Secondary_Payments")
        ws2.append(["date", "distributor", "retailer", "amount", "reference", "description"])
        r_by_id = {r["id"]: r["name"] async for r in db.dms_retailers.find({}, {"_id": 0, "id": 1, "name": 1})}
        async for e in db.dms_retailer_ledger.find(q, {"_id": 0}):
            ws2.append([(e.get("at") or "")[:10],
                        d_by_id.get(e.get("distributor_id"), ""),
                        r_by_id.get(e.get("retailer_id"), ""),
                        e.get("amount"), e.get("reference_no", ""), e.get("description", "")])
        for w in [ws1, ws2]:
            for col_letter in ["A", "B", "C", "D", "E", "F"]:
                w.column_dimensions[col_letter].width = 20
        return _xlsx_response(wb, "payments")

    # -- Import Templates --
    @router.get("/sale-bills/import-template")
    async def sale_bills_import_template(user: dict = Depends(owner_or_accountant)):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active; ws.title = "SaleBills"
        ws.append(["bill_no", "date", "distributor_email", "retailer_email", "subtotal", "gst_total", "total"])
        ws.append(["INV-1001", datetime.now().strftime("%Y-%m-%d"), "distributor1@gooil.com", "retailer1@gooil.com", 1000, 180, 1180])
        for c in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[c].width = 22
        return _xlsx_response(wb, "sale_bills_template")

    @router.get("/payments/import-template")
    async def payments_import_template(user: dict = Depends(owner_or_accountant)):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active; ws.title = "Payments"
        ws.append(["type", "date", "distributor_email", "retailer_email", "amount", "method", "reference", "description"])
        ws.append(["secondary", datetime.now().strftime("%Y-%m-%d"), "distributor1@gooil.com", "retailer1@gooil.com", 500, "cash", "PMT-1001", "On account"])
        ws.append(["primary", datetime.now().strftime("%Y-%m-%d"), "distributor1@gooil.com", "", 5000, "bank_transfer", "PMT-1002", "Against invoices"])
        for c in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[c].width = 20
        return _xlsx_response(wb, "payments_template")

    # -- Sale Bills Import --
    @router.post("/sale-bills/import")
    async def import_sale_bills(file: UploadFile = File(...), user: dict = Depends(owner_or_accountant)):
        """Import historical retailer sale bills. Sheet 'SaleBills' with columns:
        bill_no, date, distributor_email, retailer_email, subtotal, gst_total, total.
        Creates a retailer bill + a retailer-ledger invoice entry (idempotent by bill_no)."""
        from openpyxl import load_workbook
        from io import BytesIO
        raw = await file.read()
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read xlsx: {e}")
        ws = wb["SaleBills"] if "SaleBills" in wb.sheetnames else wb.active
        header = [str(h or "").strip().lower() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        col = {h: i for i, h in enumerate(header)}
        d_by_email = {d.get("email", "").lower(): d async for d in db.dms_distributors.find({}, {"_id": 0, "id": 1, "email": 1})}
        r_by_email = {r.get("email", "").lower(): r async for r in db.dms_retailers.find({}, {"_id": 0, "id": 1, "email": 1, "distributor_id": 1})}
        summary = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or v == "" for v in row):
                continue
            try:
                def _g(k):
                    idx = col.get(k, -1)
                    return row[idx] if 0 <= idx < len(row) else None
                bill_no = str(_g("bill_no") or "").strip()
                d_email = str(_g("distributor_email") or "").strip().lower()
                r_email = str(_g("retailer_email") or "").strip().lower()
                d = d_by_email.get(d_email); r = r_by_email.get(r_email)
                if not d or not r:
                    summary["skipped"] += 1; summary["errors"].append(f"Row {i}: unknown distributor/retailer email"); continue
                if not bill_no:
                    bill_no = await _next_no("retailer_invoice", "INV", 4)
                subtotal = float(_g("subtotal") or 0)
                gst_total = float(_g("gst_total") or 0)
                total = float(_g("total") or (subtotal + gst_total))
                date = str(_g("date") or _now()[:10])[:10]
                existing = await db.dms_retailer_bills.find_one({"bill_no": bill_no}, {"_id": 0, "id": 1})
                if existing:
                    summary["skipped"] += 1; summary["errors"].append(f"Row {i}: bill_no '{bill_no}' already exists"); continue
                bill = {
                    "id": _nid("rb"), "bill_no": bill_no, "order_no": None, "order_id": None,
                    "retailer_id": r["id"], "distributor_id": d["id"],
                    "items": [], "subtotal": _round(subtotal), "gst_total": _round(gst_total),
                    "total": _round(total), "status": "issued", "imported": True,
                    "created_at": date + "T00:00:00",
                }
                await db.dms_retailer_bills.insert_one(bill)
                await db.dms_retailer_ledger.insert_one({
                    "id": _nid("rle"), "distributor_id": d["id"], "retailer_id": r["id"],
                    "kind": "invoice", "reference_id": bill["id"], "reference_no": bill_no,
                    "amount": _round(total), "description": f"Imported sale bill {bill_no}", "at": date + "T00:00:00",
                })
                summary["created"] += 1
            except Exception as e:
                summary["skipped"] += 1; summary["errors"].append(f"Row {i}: {e}")
        return summary

    # -- Payments Import --
    @router.post("/payments/import")
    async def import_payments(file: UploadFile = File(...), user: dict = Depends(owner_or_accountant)):
        """Import payments. Sheet 'Payments' with columns:
        type(primary/secondary), date, distributor_email, retailer_email, amount, method, reference, description.
        Posts to primary or retailer ledger as kind=payment (idempotent by reference within party)."""
        from openpyxl import load_workbook
        from io import BytesIO
        raw = await file.read()
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read xlsx: {e}")
        ws = wb["Payments"] if "Payments" in wb.sheetnames else wb.active
        header = [str(h or "").strip().lower() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        col = {h: i for i, h in enumerate(header)}
        d_by_email = {d.get("email", "").lower(): d async for d in db.dms_distributors.find({}, {"_id": 0, "id": 1, "email": 1})}
        r_by_email = {r.get("email", "").lower(): r async for r in db.dms_retailers.find({}, {"_id": 0, "id": 1, "email": 1, "distributor_id": 1})}
        summary = {"created": 0, "skipped": 0, "errors": []}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or v == "" for v in row):
                continue
            try:
                def _g(k):
                    idx = col.get(k, -1)
                    return row[idx] if 0 <= idx < len(row) else None
                ptype = str(_g("type") or "").strip().lower()
                d_email = str(_g("distributor_email") or "").strip().lower()
                r_email = str(_g("retailer_email") or "").strip().lower()
                amount = _round(float(_g("amount") or 0))
                if amount <= 0:
                    summary["skipped"] += 1; summary["errors"].append(f"Row {i}: amount must be > 0"); continue
                method = str(_g("method") or "cash").strip()
                reference = str(_g("reference") or "").strip()
                desc = str(_g("description") or "Imported payment").strip()
                date = str(_g("date") or _now()[:10])[:10]
                d = d_by_email.get(d_email)
                if not ptype:
                    ptype = "secondary" if r_email else "primary"
                if ptype == "secondary":
                    r = r_by_email.get(r_email)
                    if not r:
                        summary["skipped"] += 1; summary["errors"].append(f"Row {i}: unknown retailer email"); continue
                    if reference and await db.dms_retailer_ledger.find_one({"retailer_id": r["id"], "reference_no": reference, "kind": "payment"}, {"_id": 0, "id": 1}):
                        summary["skipped"] += 1; summary["errors"].append(f"Row {i}: reference '{reference}' already exists"); continue
                    await db.dms_retailer_ledger.insert_one({
                        "id": _nid("rle"), "distributor_id": (d or {}).get("id") or r.get("distributor_id"), "retailer_id": r["id"],
                        "kind": "payment", "reference_no": reference or f"PMT-{i}", "amount": amount,
                        "method": method, "description": desc, "at": date + "T00:00:00", "imported": True,
                    })
                    summary["created"] += 1
                else:  # primary
                    if not d:
                        summary["skipped"] += 1; summary["errors"].append(f"Row {i}: unknown distributor email"); continue
                    if reference and await db.dms_primary_ledger.find_one({"distributor_id": d["id"], "reference_no": reference, "kind": "payment"}, {"_id": 0, "id": 1}):
                        summary["skipped"] += 1; summary["errors"].append(f"Row {i}: reference '{reference}' already exists"); continue
                    await db.dms_primary_ledger.insert_one({
                        "id": _nid("ple"), "distributor_id": d["id"],
                        "kind": "payment", "reference_no": reference or f"PMT-{i}", "amount": amount,
                        "method": method, "description": desc, "at": date + "T00:00:00", "imported": True,
                    })
                    summary["created"] += 1
            except Exception as e:
                summary["skipped"] += 1; summary["errors"].append(f"Row {i}: {e}")
        return summary

    # =========================================================================
    # PHASE 2C — Direct +Add Sales invoice (no sales order required)
    # =========================================================================
    @router.post("/direct-sales")
    async def create_direct_sale(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        """Directly create a sale bill without a preceding secondary order.
        RBAC (Task 3 — "bill for everyone"):
          - owner / super_admin: any distributor+retailer
          - distributor / distributor_accountant: own retailers only
          - salesperson: retailers under an assigned distributor only
          - retailer: counter-sale to a walk-in customer (self), optional customer{}
        """
        role = user.get("role")
        if role not in ("owner", "super_admin", "distributor", "distributor_accountant",
                        "salesperson", "team_leader", "retailer"):
            raise HTTPException(status_code=403, detail="Forbidden")
        skip_inventory = False
        if role == "retailer":
            # Retailer counter-sale: bill their own store's customer
            rid = str(user.get("retailer_id") or "").strip()
            if not rid:
                raise HTTPException(status_code=400, detail="Retailer profile missing")
            retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
            if not retailer:
                raise HTTPException(status_code=400, detail="Retailer not found")
            did = str(retailer.get("distributor_id") or "").strip()
            dist = await db.dms_distributors.find_one({"id": did}, {"_id": 0}) if did else None
            skip_inventory = True  # retailer already owns the stock; no distributor inventory move
        else:
            did = str(body.get("distributor_id") or user.get("distributor_id") or "").strip()
            rid = str(body.get("retailer_id") or "").strip()
            if not did or not rid:
                raise HTTPException(status_code=400, detail="distributor_id and retailer_id required")
            if role in ("distributor", "distributor_accountant") and did != user.get("distributor_id"):
                raise HTTPException(status_code=403, detail="Cannot create bill for another distributor")
            if role in ("salesperson", "team_leader"):
                assigned = await db.dms_sp_assignments.find_one(
                    {"salesperson_id": user["id"], "distributor_id": did}, {"_id": 0}) if role == "salesperson" else \
                    await db.dms_tl_assignments.find_one({"team_leader_id": user["id"], "distributor_id": did}, {"_id": 0})
                if not assigned:
                    raise HTTPException(status_code=403, detail="You are not assigned to this distributor")
            dist = await db.dms_distributors.find_one({"id": did}, {"_id": 0})
            retailer = await db.dms_retailers.find_one({"id": rid}, {"_id": 0})
            if not dist or not retailer:
                raise HTTPException(status_code=400, detail="Invalid distributor_id or retailer_id")
            if retailer.get("distributor_id") != did:
                raise HTTPException(status_code=400, detail="Retailer does not belong to this distributor")
        date = str(body.get("date") or _now()[:10]).strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        await _check_fy_lock(date, "direct sale")
        items = body.get("items") or []
        if not items:
            raise HTTPException(status_code=400, detail="items required")
        # Load products
        pids = [str(i.get("product_id") or "").strip() for i in items if i.get("product_id")]
        prods = {p["id"]: p async for p in db.dms_products.find({"id": {"$in": pids}}, {"_id": 0})}
        settings = await _get_settings()
        gst_default = float(settings.get("gst_pct") or 0)
        norm_items: List[Dict[str, Any]] = []
        subtotal = 0.0; gst_total = 0.0
        stop_sale = await _stop_sale_enabled()
        for it in items:
            pid = str(it.get("product_id") or "").strip()
            p = prods.get(pid)
            if not p:
                raise HTTPException(status_code=400, detail=f"Invalid product_id: {pid}")
            try:
                qb = int(it.get("qty_boxes") or 0); qp = int(it.get("qty_pcs") or 0)
            except Exception:
                raise HTTPException(status_code=400, detail="qty must be integer")
            if qb <= 0 and qp <= 0:
                raise HTTPException(status_code=400, detail=f"quantities must be > 0 for {p['name']}")
            # price: prefer body, else retailer_prices, else product unit_price
            box_price = float(it.get("box_price") or 0)
            if box_price <= 0:
                rp = await db.dms_retailer_prices.find_one({"distributor_id": did, "product_id": pid}, {"_id": 0})
                box_price = float(rp["selling_price"]) if rp else float(p.get("unit_price", 0))
            box_qty = int(p.get("box_qty", 1) or 1)
            pcs_price = _round(box_price / max(box_qty, 1))
            line_sub = _round(qb * box_price + qp * pcs_price)
            line_gst = _round(line_sub * gst_default / 100)
            line_total = _round(line_sub + line_gst)
            # stock check
            if stop_sale and not skip_inventory:
                need_boxes = qb + (qp // max(box_qty, 1)) + (1 if qp % max(box_qty, 1) > 0 else 0)
                inv = await db.dms_distributor_inventory.find_one({"distributor_id": did, "product_id": pid}, {"_id": 0, "qty_boxes": 1})
                avail = int((inv or {}).get("qty_boxes", 0) or 0)
                if need_boxes > avail:
                    raise HTTPException(status_code=400, detail=f"Insufficient distributor stock for {p['name']}: available {avail} boxes, need {need_boxes}")
            norm_items.append({
                "product_id": pid, "product_name": p.get("name", ""), "sku_code": p.get("sku_code", ""),
                "hsn": p.get("hsn", ""),
                "box_qty": box_qty, "box_price": _round(box_price), "pcs_price": pcs_price,
                "gst_pct": gst_default, "qty_boxes_dispatched": qb, "qty_pcs_dispatched": qp,
                "dispatched_qty_boxes": qb, "dispatched_qty_pcs": qp,
                "line_subtotal": line_sub, "line_gst": line_gst, "line_total": line_total,
            })
            subtotal += line_sub; gst_total += line_gst
            # decrement inventory (skip for retailer counter-sale)
            need_boxes = qb + (qp // max(box_qty, 1)) + (1 if qp % max(box_qty, 1) > 0 else 0)
            if need_boxes > 0 and not skip_inventory:
                await db.dms_distributor_inventory.update_one(
                    {"distributor_id": did, "product_id": pid},
                    {"$inc": {"qty_boxes": -need_boxes}, "$set": {"updated_at": _now()}},
                    upsert=True,
                )
                await db.dms_stock_ledger.insert_one({
                    "id": _nid("sl"), "scope": "distributor", "distributor_id": did,
                    "product_id": pid, "delta_boxes": -need_boxes,
                    "reason": "direct_sale", "reference": "", "at": _now(),
                })
        total = _round(subtotal + gst_total)
        # create bill (no order)
        bill_no = str(body.get("bill_no") or f"DS-{datetime.now().strftime('%y%m%d%H%M%S')}").strip()
        # duplicate check
        dup = await db.dms_retailer_bills.find_one({"bill_no": bill_no}, {"_id": 0})
        if dup:
            raise HTTPException(status_code=400, detail=f"Bill number '{bill_no}' already exists")
        customer = body.get("customer") or {}
        if not isinstance(customer, dict):
            customer = {}
        transport = body.get("transport") or {}
        if not isinstance(transport, dict):
            transport = {}
        bill = {
            "id": _nid("rb"), "bill_no": bill_no, "date": date,
            "order_id": None, "order_no": None,
            "retailer_id": rid, "distributor_id": did,
            "items": norm_items, "subtotal": _round(subtotal), "gst_total": _round(gst_total), "total": total,
            "status": "issued", "source": "direct_sale",
            "customer": {
                "name": str(customer.get("name") or "").strip(),
                "phone": str(customer.get("phone") or "").strip(),
                "address": str(customer.get("address") or "").strip(),
                "gstin": str(customer.get("gstin") or "").strip(),
            },
            "transport": {
                "mode": str(transport.get("mode") or "").strip(),
                "vehicle_no": str(transport.get("vehicle_no") or "").strip(),
                "transporter": str(transport.get("transporter") or "").strip(),
                "lr_no": str(transport.get("lr_no") or "").strip(),
            },
            "notes": str(body.get("notes") or "").strip(),
            "created_by": user["id"], "created_by_name": user.get("name"),
            "created_by_role": role,
            "created_at": _now(),
        }
        await db.dms_retailer_bills.insert_one(bill)
        # Retailer counter-sale is to a walk-in customer, NOT a purchase from the
        # distributor — so it must NOT affect the retailer↔distributor ledger.
        if role != "retailer":
            await db.dms_retailer_ledger.insert_one({
                "id": _nid("rle"), "distributor_id": did, "retailer_id": rid,
                "kind": "invoice", "reference_id": bill["id"], "reference_no": bill["bill_no"],
                "amount": total, "description": f"Direct sale bill {bill_no}", "at": _now(),
            })
        bill.pop("_id", None)
        return _clean(bill)

    # =========================================================================
    # PHASE 2C — PO PDF (Purchase Order = primary order print view)
    # =========================================================================
    @router.get("/print/purchase-order/{oid}")
    async def print_purchase_order(oid: str, user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        order = await db.dms_primary_orders.find_one({"id": oid}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if role in ("distributor", "distributor_accountant") and order["distributor_id"] != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        dist = await db.dms_distributors.find_one({"id": order["distributor_id"]}, {"_id": 0})
        order["distributor"] = dist
        s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0, "invoice_terms": 1, "invoice_message": 1, "company_name": 1}) or {}
        order["invoice_terms"] = s.get("invoice_terms") or ""
        order["invoice_message"] = s.get("invoice_message") or ""
        order["company_name"] = s.get("company_name") or "GO OIL Lubricants"
        order["doc_type"] = "Purchase Order"
        return order

    # =========================================================================
    # PHASE 2C — Document stubs (Estimate / Delivery Challan / Sale Return / Credit / Debit Note)
    # =========================================================================
    DOC_TYPES = {"estimate", "delivery_challan", "sale_return", "credit_note", "debit_note"}
    DOC_LABELS = {
        "estimate": "Estimate/Quotation", "delivery_challan": "Delivery Challan",
        "sale_return": "Sale Return", "credit_note": "Credit Note", "debit_note": "Debit Note",
    }
    DOC_PREFIX = {"estimate": "EST", "delivery_challan": "DC", "sale_return": "SR", "credit_note": "CN", "debit_note": "DN"}

    async def _party_lookup(party_type: str, party_id: str) -> Dict[str, Any]:
        col = db.dms_retailers if party_type == "retailer" else db.dms_distributors
        return await col.find_one({"id": party_id}, {"_id": 0}) or {}

    @router.get("/documents")
    async def list_documents(user: dict = Depends(get_current_user), type: Optional[str] = None,
                             start: Optional[str] = None, end: Optional[str] = None):
        role = user.get("role")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        q: Dict[str, Any] = {}
        if type:
            if type not in DOC_TYPES:
                raise HTTPException(status_code=400, detail=f"type must be one of {sorted(DOC_TYPES)}")
            q["type"] = type
        if start: q.setdefault("date", {})["$gte"] = start
        if end: q.setdefault("date", {})["$lte"] = end
        # Distributor scope
        if role in ("distributor", "distributor_accountant"):
            q["distributor_id"] = user.get("distributor_id")
        rows = await db.dms_documents.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return {"data": rows, "count": len(rows)}

    @router.post("/documents")
    async def create_document(body: Dict[str, Any] = Body(...), user: dict = Depends(get_current_user)):
        role = user.get("role")
        if role not in ("owner", "super_admin", "distributor", "distributor_accountant"):
            raise HTTPException(status_code=403, detail="Forbidden")
        doc_type = str(body.get("type") or "").strip().lower()
        if doc_type not in DOC_TYPES:
            raise HTTPException(status_code=400, detail=f"type must be one of {sorted(DOC_TYPES)}")
        # Delivery Challan is now auto-generated during the Dispatch flow — not created manually.
        if doc_type == "delivery_challan":
            raise HTTPException(status_code=400, detail="Delivery Challan is generated automatically after Dispatch. It cannot be created here.")
        date = str(body.get("date") or _now()[:10]).strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except Exception:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        await _check_fy_lock(date, f"{DOC_LABELS[doc_type]}")
        party_type = str(body.get("party_type") or "").strip().lower()
        if party_type not in ("retailer", "distributor"):
            raise HTTPException(status_code=400, detail="party_type must be retailer or distributor")
        party_id = str(body.get("party_id") or "").strip()
        party = await _party_lookup(party_type, party_id)
        if not party:
            raise HTTPException(status_code=400, detail=f"Invalid {party_type} party_id")
        if role in ("distributor", "distributor_accountant"):
            # distributor scope: only own retailers OR self
            if party_type == "retailer" and party.get("distributor_id") != user.get("distributor_id"):
                raise HTTPException(status_code=403, detail="Retailer not under your distributor")
            if party_type == "distributor" and party.get("id") != user.get("distributor_id"):
                raise HTTPException(status_code=403, detail="Cannot create doc for another distributor")
        # Items (product-agnostic — free text allowed)
        items = body.get("items") or []
        norm_items: List[Dict[str, Any]] = []
        subtotal = 0.0
        for it in items:
            desc = str(it.get("description") or it.get("product_name") or "").strip()
            try:
                qty = float(it.get("qty") or 0); rate = float(it.get("rate") or 0)
            except Exception:
                raise HTTPException(status_code=400, detail="qty and rate must be numbers")
            amt = _round(qty * rate)
            norm_items.append({
                "description": desc, "product_id": str(it.get("product_id") or "") or None,
                "qty": qty, "rate": _round(rate), "amount": amt,
            })
            subtotal += amt
        gst_pct = float(body.get("gst_pct") or 0)
        gst_total = _round(subtotal * gst_pct / 100)
        total = _round(subtotal + gst_total)
        # allow a lump-sum amount (e.g. Credit/Debit notes without line items)
        if total <= 0 and body.get("amount"):
            total = _round(float(body.get("amount")))
            subtotal = total
        # doc_no
        cnt = await db.dms_documents.count_documents({"type": doc_type})
        doc_no = str(body.get("doc_no") or f"{DOC_PREFIX[doc_type]}-{datetime.now().strftime('%y%m%d')}-{cnt + 1:04d}").strip()
        # duplicate check within type
        dup = await db.dms_documents.find_one({"type": doc_type, "doc_no": doc_no}, {"_id": 0, "id": 1})
        if dup:
            raise HTTPException(status_code=400, detail=f"Document number '{doc_no}' already exists for {DOC_LABELS[doc_type]}")
        doc = {
            "id": _nid("doc"), "type": doc_type, "doc_no": doc_no, "date": date,
            "party_type": party_type, "party_id": party_id, "party_name": party.get("name", ""),
            "distributor_id": user.get("distributor_id") if role in ("distributor", "distributor_accountant") else (party.get("distributor_id") if party_type == "retailer" else party.get("id")),
            "items": norm_items, "subtotal": _round(subtotal), "gst_pct": gst_pct,
            "gst_total": gst_total, "total": total,
            "notes": str(body.get("notes") or "").strip(),
            "created_by": user["id"], "created_by_name": user.get("name"),
            "created_at": _now(),
        }
        await db.dms_documents.insert_one(doc)
        doc.pop("_id", None)

        # ── Side-effects (Item 9) ──────────────────────────────────────────
        dist_id = doc["distributor_id"]
        # Sale Return → adjust distributor inventory + stock ledger
        if doc_type == "sale_return" and dist_id:
            # retailer returns to distributor → stock increases (+)
            # distributor returns to company → stock decreases (-)
            sign = 1 if party_type == "retailer" else -1
            for it in norm_items:
                pid = it.get("product_id")
                qty_boxes = int(round(float(it.get("qty") or 0)))
                if not pid or qty_boxes <= 0:
                    continue
                inv = await db.dms_distributor_inventory.find_one({"distributor_id": dist_id, "product_id": pid})
                if inv:
                    new_qty = max(0, int(inv.get("qty_boxes", 0)) + sign * qty_boxes)
                    await db.dms_distributor_inventory.update_one({"id": inv["id"]}, {"$set": {"qty_boxes": new_qty, "updated_at": _now()}})
                elif sign > 0:
                    prod = await db.dms_products.find_one({"id": pid}, {"_id": 0, "name": 1, "sku_code": 1})
                    await db.dms_distributor_inventory.insert_one({
                        "id": _nid("dinv"), "distributor_id": dist_id, "product_id": pid,
                        "product_name": (prod or {}).get("name"), "sku_code": (prod or {}).get("sku_code"),
                        "qty_boxes": qty_boxes, "updated_at": _now(),
                    })
                await db.dms_stock_ledger.insert_one({
                    "id": _nid("sl"), "scope": "distributor", "distributor_id": dist_id,
                    "product_id": pid, "delta_boxes": sign * qty_boxes,
                    "reason": ("sale_return_from_retailer" if sign > 0 else "sale_return_to_company"),
                    "reference": doc_no, "at": _now(),
                })
        # Credit Note / Debit Note → auto post to respective party ledger
        if doc_type in ("credit_note", "debit_note") and total > 0:
            if party_type == "retailer":
                await db.dms_retailer_ledger.insert_one({
                    "id": _nid("rle"), "distributor_id": dist_id, "retailer_id": party_id,
                    "kind": doc_type, "reference_id": doc["id"], "reference_no": doc_no,
                    "amount": total, "description": f"{DOC_LABELS[doc_type]} {doc_no}", "at": _now(),
                })
            else:  # distributor party → primary ledger
                await db.dms_primary_ledger.insert_one({
                    "id": _nid("ple"), "distributor_id": party_id,
                    "kind": doc_type, "reference_id": doc["id"], "reference_no": doc_no,
                    "amount": total, "description": f"{DOC_LABELS[doc_type]} {doc_no}", "at": _now(),
                })
        return _clean(doc)

    @router.get("/documents/{did}")
    async def get_document(did: str, user: dict = Depends(get_current_user)):
        d = await db.dms_documents.find_one({"id": did}, {"_id": 0})
        if not d:
            raise HTTPException(status_code=404, detail="Document not found")
        role = user.get("role")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        if role in ("distributor", "distributor_accountant") and d.get("distributor_id") != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        return d

    @router.get("/documents/{did}/print")
    async def print_document(did: str, user: dict = Depends(get_current_user)):
        d = await db.dms_documents.find_one({"id": did}, {"_id": 0})
        if not d:
            raise HTTPException(status_code=404, detail="Document not found")
        role = user.get("role")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Forbidden")
        if role in ("distributor", "distributor_accountant") and d.get("distributor_id") != user.get("distributor_id"):
            raise HTTPException(status_code=403, detail="Forbidden")
        party = await _party_lookup(d.get("party_type", "retailer"), d.get("party_id", ""))
        d["party"] = party
        d["doc_type_label"] = DOC_LABELS.get(d["type"], d["type"])
        s = await db.dms_settings.find_one({"id": "global"}, {"_id": 0, "invoice_terms": 1, "invoice_message": 1, "company_name": 1}) or {}
        d["invoice_terms"] = s.get("invoice_terms") or ""
        d["invoice_message"] = s.get("invoice_message") or ""
        d["company_name"] = s.get("company_name") or "GO OIL Lubricants"
        return d

    # =========================================================================
    # PHASE 2C — Finance Dashboard Snapshot
    # =========================================================================
    @router.get("/dashboard/finance-snapshot")
    async def finance_snapshot(user: dict = Depends(_guard("owner", "owner_accountant", "super_admin"))):
        agg_bank = await db.dms_bank_accounts.aggregate([{"$group": {"_id": None, "t": {"$sum": "$current_balance"}}}]).to_list(2)
        cash_in_bank = _round((agg_bank[0]["t"] if agg_bank else 0) or 0)
        agg_cash = await db.dms_cash_register.aggregate([{"$group": {"_id": "$type", "t": {"$sum": "$amount"}}}]).to_list(10)
        totals = {r["_id"]: r["t"] for r in agg_cash}
        cash_in_hand = _round((totals.get("in") or 0) - (totals.get("out") or 0))
        agg_loans = await db.dms_loan_accounts.aggregate([{"$group": {"_id": None, "t": {"$sum": "$outstanding"}}}]).to_list(2)
        outstanding_loans = _round((agg_loans[0]["t"] if agg_loans else 0) or 0)
        return {
            "cash_in_bank": cash_in_bank,
            "cash_in_hand": cash_in_hand,
            "outstanding_loans": outstanding_loans,
            "net_liquid": _round(cash_in_bank + cash_in_hand),
            "net_position": _round(cash_in_bank + cash_in_hand - outstanding_loans),
        }

    # =========================================================================
    # PHASE 2C — Godown Reorder Level + Low-Stock alerts
    # =========================================================================
    @router.put("/godowns/{gid}/reorder-level")
    async def set_reorder_level(gid: str, body: Dict[str, Any] = Body(...), user: dict = Depends(owner_or_owner_acct)):
        product_id = str(body.get("product_id") or "").strip()
        if not product_id:
            raise HTTPException(status_code=400, detail="product_id required")
        try:
            level = int(body.get("reorder_level_boxes") or 0)
        except Exception:
            raise HTTPException(status_code=400, detail="reorder_level_boxes must be integer")
        if level < 0:
            raise HTTPException(status_code=400, detail="reorder_level_boxes must be ≥ 0")
        # verify godown exists
        g = await db.dms_godowns.find_one({"id": gid}, {"_id": 0, "id": 1})
        if not g:
            raise HTTPException(status_code=404, detail="Godown not found")
        # upsert inventory row with reorder_level
        cur = await db.dms_godown_inventory.find_one({"godown_id": gid, "product_id": product_id})
        if cur:
            await db.dms_godown_inventory.update_one(
                {"godown_id": gid, "product_id": product_id},
                {"$set": {"reorder_level_boxes": level, "updated_at": _now()}},
            )
        else:
            await db.dms_godown_inventory.insert_one({
                "id": _nid("ginv"), "godown_id": gid, "product_id": product_id,
                "qty_boxes": 0, "reorder_level_boxes": level, "updated_at": _now(),
            })
        return {"ok": True, "reorder_level_boxes": level}

    @router.get("/godowns/low-stock")
    async def list_low_stock(user: dict = Depends(owner_or_owner_acct)):
        """List all godown-inventory rows where qty_boxes ≤ reorder_level_boxes AND reorder_level_boxes > 0."""
        rows = await db.dms_godown_inventory.find(
            {"reorder_level_boxes": {"$gt": 0}, "$expr": {"$lte": ["$qty_boxes", "$reorder_level_boxes"]}},
            {"_id": 0}
        ).to_list(2000)
        # enrich
        gods = {g["id"]: g["name"] async for g in db.dms_godowns.find({}, {"_id": 0, "id": 1, "name": 1})}
        pids = [r["product_id"] for r in rows]
        prods = {p["id"]: p async for p in db.dms_products.find({"id": {"$in": pids}}, {"_id": 0})}
        for r in rows:
            r["godown_name"] = gods.get(r["godown_id"], "")
            p = prods.get(r["product_id"], {})
            r["product_name"] = p.get("name", ""); r["sku_code"] = p.get("sku_code", "")
        return {"data": rows, "count": len(rows)}

    # =========================================================================
    # PHASE 3 — Reports Module
    # =========================================================================
    from dms_reports import (
        REPORT_CATALOG, CATEGORY_ORDER, role_can_see_report,
        get_report_by_id, run_report, run_sale_report,
    )

    @router.get("/reports/catalog")
    async def reports_catalog(user: dict = Depends(get_current_user)):
        role = user.get("role", "")
        if role == "retailer":
            raise HTTPException(status_code=403, detail="Retailers cannot access reports")
        favs = await db.dms_report_favorites.find(
            {"user_id": user["id"]}, {"_id": 0, "report_id": 1}
        ).to_list(200)
        fav_ids = {f["report_id"] for f in favs}
        visible = [r for r in REPORT_CATALOG if role_can_see_report(r, role)]
        groups = []
        for cat_id, cat_label in CATEGORY_ORDER:
            items = [{
                "id": r["id"], "name": r["name"], "description": r["description"],
                "status": r["status"], "filters": r.get("filters", []),
                "is_favorite": r["id"] in fav_ids,
            } for r in visible if r["category"] == cat_id]
            if items:
                groups.append({"category": cat_id, "label": cat_label, "items": items})
        favorites = [{"id": r["id"], "name": r["name"], "category": r["category"],
                      "status": r["status"]} for r in visible if r["id"] in fav_ids]
        return {"groups": groups, "favorites": favorites}

    @router.post("/reports/favorites/toggle/{report_id}")
    async def toggle_favorite(report_id: str, user: dict = Depends(get_current_user)):
        if user.get("role") == "retailer":
            raise HTTPException(status_code=403, detail="Retailers cannot access reports")
        report = get_report_by_id(report_id)
        if not report:
            raise HTTPException(status_code=404, detail="Unknown report_id")
        if not role_can_see_report(report, user.get("role", "")):
            raise HTTPException(status_code=403, detail="Report not accessible for your role")
        existing = await db.dms_report_favorites.find_one(
            {"user_id": user["id"], "report_id": report_id}
        )
        if existing:
            await db.dms_report_favorites.delete_one(
                {"user_id": user["id"], "report_id": report_id}
            )
            return {"ok": True, "is_favorite": False}
        await db.dms_report_favorites.insert_one({
            "id": _nid("rfav"), "user_id": user["id"],
            "report_id": report_id, "at": _now(),
        })
        return {"ok": True, "is_favorite": True}

    # Per-user saved filters (Polish item)
    @router.get("/reports/saved-filters/{report_id}")
    async def list_saved_filters(report_id: str, user: dict = Depends(get_current_user)):
        if user.get("role") == "retailer":
            raise HTTPException(status_code=403, detail="Retailers cannot access reports")
        rows = await db.dms_report_saved_filters.find(
            {"user_id": user["id"], "report_id": report_id}, {"_id": 0}
        ).to_list(50)
        return {"data": rows, "count": len(rows)}

    @router.post("/reports/saved-filters/{report_id}")
    async def save_filter(report_id: str, payload: Dict[str, Any] = Body(...),
                          user: dict = Depends(get_current_user)):
        if user.get("role") == "retailer":
            raise HTTPException(status_code=403, detail="Retailers cannot access reports")
        name = (payload.get("name") or "").strip()
        filters = payload.get("filters") or {}
        if not name:
            raise HTTPException(status_code=400, detail="Name required")
        doc = {
            "id": _nid("sf"),
            "user_id": user["id"],
            "report_id": report_id,
            "name": name,
            "filters": filters,
            "at": _now(),
        }
        await db.dms_report_saved_filters.insert_one(doc)
        return {"ok": True, "id": doc["id"]}

    @router.delete("/reports/saved-filters/{filter_id}")
    async def delete_saved_filter(filter_id: str, user: dict = Depends(get_current_user)):
        if user.get("role") == "retailer":
            raise HTTPException(status_code=403, detail="Retailers cannot access reports")
        r = await db.dms_report_saved_filters.delete_one(
            {"id": filter_id, "user_id": user["id"]}
        )
        return {"deleted": r.deleted_count}

    def _report_guard(report_id: str, user: dict):
        report = get_report_by_id(report_id)
        if not report:
            raise HTTPException(status_code=404, detail="Unknown report_id")
        if report.get("status") != "live":
            raise HTTPException(status_code=400, detail="Report is not yet available")
        if not role_can_see_report(report, user.get("role", "")):
            raise HTTPException(status_code=403, detail="Report not accessible for your role")
        return report

    def _extract_filters(request: Request) -> Dict[str, Any]:
        # Pull all query params — the report engine ignores unknown keys.
        out = {}
        for k, v in request.query_params.items():
            if v is None or v == "":
                continue
            out[k] = v
        return out

    @router.get("/reports/{report_id}/run")
    async def report_run(report_id: str, request: Request,
                         user: dict = Depends(get_current_user)):
        report = _report_guard(report_id, user)
        filters = _extract_filters(request)
        data = await run_report(db, user, report_id, filters)
        data["report"] = {"id": report["id"], "name": report["name"],
                          "category": report["category"]}
        return data

    @router.get("/reports/{report_id}/export")
    async def report_export(report_id: str, request: Request,
                            user: dict = Depends(get_current_user)):
        report = _report_guard(report_id, user)
        from openpyxl import Workbook
        filters = _extract_filters(request)
        data = await run_report(db, user, report_id, filters)
        wb = Workbook(); ws = wb.active
        # Excel sheet title cannot contain \/?*[]: characters
        safe_title = "".join(c for c in report["name"] if c not in "\\/?*[]:")[:31]
        ws.title = safe_title or "Report"
        ws.append([report["name"]])
        param_summary = " | ".join(f"{k}={v}" for k, v in filters.items()) or "(no filters)"
        ws.append([f"Filters: {param_summary}", f"Generated: {_now()}"])
        ws.append([])
        cols = data.get("columns", [])
        ws.append([c["label"] for c in cols])
        for r in data.get("rows", []):
            ws.append([r.get(c["key"], "") for c in cols])
        # Totals row
        tot = data.get("totals", {})
        totals_row = []
        for c in cols:
            k = c["key"]
            if c.get("totals") and k in tot:
                totals_row.append(tot[k])
            else:
                totals_row.append("")
        if any(v != "" for v in totals_row):
            ws.append([])
            # Label the totals row on the first non-currency column
            totals_row[0] = "TOTAL"
            ws.append(totals_row)
        # Column widths — heuristic
        for i, c in enumerate(cols, start=1):
            w = 14 if c.get("type") in ("currency", "number", "int", "pct") else 22
            ws.column_dimensions[chr(64 + i) if i < 27 else "A" + chr(64 + (i - 26))].width = w
        return _xlsx_response(wb, report["id"])

    # Legacy Sale Report endpoints (kept for backward compatibility)
    @router.get("/reports/sale/run")
    async def report_sale_run_legacy(
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        sale_type: str = "both",
        party_id: Optional[str] = None,
        user: dict = Depends(get_current_user),
    ):
        _report_guard("sale", user)
        if sale_type not in ("primary", "secondary", "both"):
            raise HTTPException(status_code=400, detail="sale_type must be primary|secondary|both")
        return await run_sale_report(db, user, date_from=date_from, date_to=date_to,
                                     sale_type=sale_type, party_id=party_id)

    @router.get("/reports/sale/export")
    async def report_sale_export_legacy(
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        sale_type: str = "both",
        party_id: Optional[str] = None,
        user: dict = Depends(get_current_user),
    ):
        _report_guard("sale", user)
        from openpyxl import Workbook
        data = await run_sale_report(db, user, date_from=date_from, date_to=date_to,
                                     sale_type=sale_type, party_id=party_id)
        wb = Workbook(); ws = wb.active; ws.title = "Sale Report"
        ws.append(["Sale Report"])
        ws.append([f"Range: {date_from or '(all time)'} to {date_to or '(today)'}",
                   f"Sale Type: {sale_type}", f"Generated: {_now()}"])
        ws.append([])
        headers = ["Type", "Bill No", "Date", "Order No", "Party Type", "Party Name",
                   "Items", "Subtotal", "GST", "Total"]
        ws.append(headers)
        for r in data["rows"]:
            ws.append([r.get("sale_type", ""), r.get("bill_no", ""),
                       (r.get("date") or "")[:10], r.get("order_no", ""),
                       r.get("party_type", ""), r.get("party_name", ""),
                       r.get("items_count", 0), r.get("subtotal", 0),
                       r.get("gst_total", 0), r.get("total", 0)])
        return _xlsx_response(wb, "sale_report")

    return router
