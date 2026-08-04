#!/usr/bin/env python3
"""
Phase 2C Backend Testing — GO OIL DMS
All test credentials in /app/memory/test_credentials.md — passwords are GoOil@2026
Base URL from /app/frontend/.env: REACT_APP_BACKEND_URL
All endpoints prefixed with /api
"""
import requests
import json
import io
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Read base URL from frontend/.env
with open("/app/frontend/.env", "r") as f:
    for line in f:
        if line.startswith("REACT_APP_BACKEND_URL="):
            BASE_URL = line.split("=")[1].strip()
            break

API_BASE = f"{BASE_URL}/api"
PASSWORD = "GoOil@2026"

# Test accounts
ACCOUNTS = {
    "owner": "owner@gooil.com",
    "accountant": "accountant@gooil.com",
    "distributor1": "distributor1@gooil.com",
    "distributor2": "distributor2@gooil.com",
    "retailer1": "retailer1@gooil.com",
    "retailer2": "retailer2@gooil.com",
    "salesperson": "salesperson@gooil.com",
}

def login(email: str) -> str:
    """Login and return JWT token"""
    resp = requests.post(f"{API_BASE}/auth/login", json={"email": email, "password": PASSWORD})
    if resp.status_code != 200:
        raise Exception(f"Login failed for {email}: {resp.status_code} {resp.text}")
    return resp.json()["token"]

def headers(token: str) -> dict:
    """Return authorization headers"""
    return {"Authorization": f"Bearer {token}"}

# ============================================================================
# TEST SUITE
# ============================================================================
def main():
    print("=" * 80)
    print("PHASE 2C BACKEND TESTING — GO OIL DMS")
    print("=" * 80)
    
    # Login all accounts
    print("\n[SETUP] Logging in all test accounts...")
    tokens = {}
    for role, email in ACCOUNTS.items():
        try:
            tokens[role] = login(email)
            print(f"  ✅ {role}: {email}")
        except Exception as e:
            print(f"  ❌ {role}: {email} — {e}")
            return
    
    print(f"\n✅ All {len(tokens)} accounts logged in successfully\n")
    
    # Get IDs for testing
    print("[SETUP] Fetching test data IDs...")
    
    # Get distributor IDs
    resp = requests.get(f"{API_BASE}/dms/distributors", headers=headers(tokens["owner"]))
    distributors = resp.json()
    if isinstance(distributors, dict) and "data" in distributors:
        distributors = distributors["data"]
    dist1_id = next((d["id"] for d in distributors if d["email"] == "distributor1@gooil.com"), None)
    dist2_id = next((d["id"] for d in distributors if d["email"] == "distributor2@gooil.com"), None)
    print(f"  Distributor 1 ID: {dist1_id}")
    print(f"  Distributor 2 ID: {dist2_id}")
    
    # Get retailer IDs
    resp = requests.get(f"{API_BASE}/dms/retailers", headers=headers(tokens["owner"]))
    retailers = resp.json()
    if isinstance(retailers, dict) and "data" in retailers:
        retailers = retailers["data"]
    retailer1_id = next((r["id"] for r in retailers if r["email"] == "retailer1@gooil.com"), None)
    retailer2_id = next((r["id"] for r in retailers if r["email"] == "retailer2@gooil.com"), None)
    print(f"  Retailer 1 ID: {retailer1_id}")
    print(f"  Retailer 2 ID: {retailer2_id}")
    
    # Get product IDs
    resp = requests.get(f"{API_BASE}/dms/products", headers=headers(tokens["owner"]))
    products = resp.json()
    if isinstance(products, dict) and "data" in products:
        products = products["data"]
    product1_id = products[0]["id"] if products else None
    product2_id = products[1]["id"] if len(products) > 1 else None
    print(f"  Product 1 ID: {product1_id}")
    print(f"  Product 2 ID: {product2_id}")
    
    # Get godown IDs
    resp = requests.get(f"{API_BASE}/dms/godowns", headers=headers(tokens["owner"]))
    godowns = resp.json()
    if isinstance(godowns, dict) and "data" in godowns:
        godowns = godowns["data"]
    godown_id = godowns[0]["id"] if godowns else None
    print(f"  Godown ID: {godown_id}")
    
    # Get primary order ID (for PO PDF test)
    resp = requests.get(f"{API_BASE}/dms/primary-orders", headers=headers(tokens["owner"]))
    primary_orders = resp.json()
    primary_order_id = None
    if isinstance(primary_orders, dict) and "data" in primary_orders:
        primary_orders = primary_orders["data"]
    if primary_orders:
        primary_order_id = primary_orders[0]["id"]
    print(f"  Primary Order ID: {primary_order_id}")
    
    print()
    
    # ========================================================================
    # TEST 1: PARTIES EXPORT (owner only)
    # ========================================================================
    print("=" * 80)
    print("TEST 1: PARTIES EXPORT — GET /api/dms/parties/export")
    print("=" * 80)
    
    print("\n[1.1] Owner GET /api/dms/parties/export → 200 + xlsx content")
    resp = requests.get(f"{API_BASE}/dms/parties/export", headers=headers(tokens["owner"]))
    if resp.status_code == 200:
        content_type = resp.headers.get("content-type", "")
        size = len(resp.content)
        print(f"  ✅ PASS: HTTP 200, content-type={content_type}, size={size} bytes")
        
        # Verify it's a valid xlsx with 2 sheets
        try:
            wb = load_workbook(io.BytesIO(resp.content))
            sheet_names = wb.sheetnames
            print(f"  ✅ PASS: Valid xlsx with sheets: {sheet_names}")
            if "Distributors" in sheet_names and "Retailers" in sheet_names:
                print(f"  ✅ PASS: Both 'Distributors' and 'Retailers' sheets present")
            else:
                print(f"  ❌ FAIL: Expected 'Distributors' and 'Retailers' sheets, got {sheet_names}")
        except Exception as e:
            print(f"  ❌ FAIL: Invalid xlsx file: {e}")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
    
    print("\n[1.2] Salesperson GET /api/dms/parties/export → 403")
    resp = requests.get(f"{API_BASE}/dms/parties/export", headers=headers(tokens["salesperson"]))
    if resp.status_code == 403:
        print(f"  ✅ PASS: Salesperson blocked (403)")
    else:
        print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    print("\n[1.3] Retailer GET /api/dms/parties/export → 403")
    resp = requests.get(f"{API_BASE}/dms/parties/export", headers=headers(tokens["retailer1"]))
    if resp.status_code == 403:
        print(f"  ✅ PASS: Retailer blocked (403)")
    else:
        print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    # ========================================================================
    # TEST 2: PARTIES IMPORT (owner only)
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 2: PARTIES IMPORT — POST /api/dms/parties/import")
    print("=" * 80)
    
    print("\n[2.1] Owner POST /api/dms/parties/import with valid xlsx")
    
    # Create test xlsx with 2 sheets
    wb = Workbook()
    
    # Sheet 1: Distributors
    ws_dist = wb.active
    ws_dist.title = "Distributors"
    ws_dist.append(["name", "code", "email", "phone", "address", "gstin", "credit_limit", "active"])
    ws_dist.append(["QA Test Distributor", "QA-DIST-001", "qa_dist@gooil.com", "9876543210", "Test Address", "29ABCDE1234F1Z5", "100000", "TRUE"])
    ws_dist.append(["Anil Distributor — Delhi", "DIST-001", "distributor1@gooil.com", "9876543211", "Updated Address", "29ABCDE1234F1Z6", "200000", "TRUE"])
    
    # Sheet 2: Retailers
    ws_ret = wb.create_sheet("Retailers")
    ws_ret.append(["name", "code", "email", "phone", "address", "gstin", "distributor_email", "active"])
    ws_ret.append(["QA Test Retailer", "QA-RET-001", "qa_ret@gooil.com", "9876543220", "Test Retailer Address", "29ABCDE1234F1Z7", "qa_dist@gooil.com", "TRUE"])
    ws_ret.append(["Invalid Retailer", "INV-RET-001", "invalid_ret@gooil.com", "9876543221", "Invalid Address", "29ABCDE1234F1Z8", "unknown@gooil.com", "TRUE"])
    
    # Save to BytesIO
    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    
    files = {"file": ("parties_import.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = requests.post(f"{API_BASE}/dms/parties/import", headers=headers(tokens["owner"]), files=files)
    
    if resp.status_code == 200:
        result = resp.json()
        print(f"  ✅ PASS: HTTP 200")
        print(f"  Result: {json.dumps(result, indent=2)}")
        
        # Verify expected results
        dist_result = result.get("distributors", {})
        ret_result = result.get("retailers", {})
        
        if dist_result.get("created") == 1 and dist_result.get("updated") == 1:
            print(f"  ✅ PASS: Distributors — created=1, updated=1")
        else:
            print(f"  ❌ FAIL: Distributors — expected created=1, updated=1, got {dist_result}")
        
        if ret_result.get("created") == 1 and ret_result.get("skipped") == 1:
            print(f"  ✅ PASS: Retailers — created=1, skipped=1 (unknown distributor)")
        else:
            print(f"  ❌ FAIL: Retailers — expected created=1, skipped=1, got {ret_result}")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
    
    # ========================================================================
    # TEST 3: SALE BILLS EXPORT (owner only)
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 3: SALE BILLS EXPORT — GET /api/dms/sale-bills/export")
    print("=" * 80)
    
    print("\n[3.1] Owner GET /api/dms/sale-bills/export → 200 + xlsx with 2 sheets")
    resp = requests.get(f"{API_BASE}/dms/sale-bills/export", headers=headers(tokens["owner"]))
    if resp.status_code == 200:
        content_type = resp.headers.get("content-type", "")
        size = len(resp.content)
        print(f"  ✅ PASS: HTTP 200, content-type={content_type}, size={size} bytes")
        
        # Verify xlsx structure
        try:
            wb = load_workbook(io.BytesIO(resp.content))
            sheet_names = wb.sheetnames
            print(f"  ✅ PASS: Valid xlsx with sheets: {sheet_names}")
            if "Primary_eBills" in sheet_names and "Retailer_Bills" in sheet_names:
                print(f"  ✅ PASS: Both 'Primary_eBills' and 'Retailer_Bills' sheets present")
                
                # Check for sample bills
                ws_eb = wb["Primary_eBills"]
                ws_rb = wb["Retailer_Bills"]
                
                eb_rows = list(ws_eb.iter_rows(values_only=True))
                rb_rows = list(ws_rb.iter_rows(values_only=True))
                
                # Check for EB-SAMPLE and RB-SAMPLE
                eb_has_sample = any("EB-SAMPLE" in str(row) for row in eb_rows)
                rb_has_sample = any("RB-SAMPLE" in str(row) for row in rb_rows)
                
                if eb_has_sample:
                    print(f"  ✅ PASS: EB-SAMPLE found in Primary_eBills")
                else:
                    print(f"  ⚠️  WARNING: EB-SAMPLE not found in Primary_eBills")
                
                if rb_has_sample:
                    print(f"  ✅ PASS: RB-SAMPLE found in Retailer_Bills")
                else:
                    print(f"  ⚠️  WARNING: RB-SAMPLE not found in Retailer_Bills")
            else:
                print(f"  ❌ FAIL: Expected 'Primary_eBills' and 'Retailer_Bills' sheets, got {sheet_names}")
        except Exception as e:
            print(f"  ❌ FAIL: Invalid xlsx file: {e}")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
    
    print("\n[3.2] Distributor GET /api/dms/sale-bills/export → 403")
    resp = requests.get(f"{API_BASE}/dms/sale-bills/export", headers=headers(tokens["distributor1"]))
    if resp.status_code == 403:
        print(f"  ✅ PASS: Distributor blocked (403)")
    else:
        print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    print("\n[3.3] Retailer GET /api/dms/sale-bills/export → 403")
    resp = requests.get(f"{API_BASE}/dms/sale-bills/export", headers=headers(tokens["retailer1"]))
    if resp.status_code == 403:
        print(f"  ✅ PASS: Retailer blocked (403)")
    else:
        print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    # ========================================================================
    # TEST 4: PAYMENTS EXPORT (owner only)
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 4: PAYMENTS EXPORT — GET /api/dms/payments/export")
    print("=" * 80)
    
    print("\n[4.1] Owner GET /api/dms/payments/export → 200 + xlsx with 2 sheets")
    resp = requests.get(f"{API_BASE}/dms/payments/export", headers=headers(tokens["owner"]))
    if resp.status_code == 200:
        content_type = resp.headers.get("content-type", "")
        size = len(resp.content)
        print(f"  ✅ PASS: HTTP 200, content-type={content_type}, size={size} bytes")
        
        # Verify xlsx structure
        try:
            wb = load_workbook(io.BytesIO(resp.content))
            sheet_names = wb.sheetnames
            print(f"  ✅ PASS: Valid xlsx with sheets: {sheet_names}")
            if "Primary_Payments" in sheet_names and "Secondary_Payments" in sheet_names:
                print(f"  ✅ PASS: Both 'Primary_Payments' and 'Secondary_Payments' sheets present")
            else:
                print(f"  ❌ FAIL: Expected 'Primary_Payments' and 'Secondary_Payments' sheets, got {sheet_names}")
        except Exception as e:
            print(f"  ❌ FAIL: Invalid xlsx file: {e}")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
    
    # ========================================================================
    # TEST 5: DIRECT SALES (owner + distributor)
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 5: DIRECT SALES — POST /api/dms/direct-sales")
    print("=" * 80)
    
    if not all([dist1_id, retailer1_id, product1_id]):
        print("  ⚠️  SKIP: Missing required IDs for direct sales test")
    else:
        print("\n[5.1] Owner POST /api/dms/direct-sales → 200, creates retailer bill")
        direct_sale_data = {
            "distributor_id": dist1_id,
            "retailer_id": retailer1_id,
            "date": "2026-08-04",
            "items": [
                {"product_id": product1_id, "qty_boxes": 1, "box_price": 500}
            ],
            "notes": "QA direct sale"
        }
        resp = requests.post(f"{API_BASE}/dms/direct-sales", json=direct_sale_data, headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            result = resp.json()
            print(f"  ✅ PASS: HTTP 200, bill created")
            print(f"  Bill ID: {result.get('bill_id')}")
            print(f"  Bill No: {result.get('bill_no')}")
            
            # Verify source=direct_sale
            bill_id = result.get("bill_id")
            if bill_id:
                resp_bill = requests.get(f"{API_BASE}/dms/retailer-bills/{bill_id}", headers=headers(tokens["owner"]))
                if resp_bill.status_code == 200:
                    bill = resp_bill.json()
                    if bill.get("source") == "direct_sale":
                        print(f"  ✅ PASS: Bill source='direct_sale'")
                    else:
                        print(f"  ❌ FAIL: Expected source='direct_sale', got {bill.get('source')}")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
        
        print("\n[5.2] Distributor1 POST /api/dms/direct-sales (own retailer) → 200")
        direct_sale_data2 = {
            "retailer_id": retailer1_id,
            "date": "2026-08-04",
            "items": [
                {"product_id": product1_id, "qty_boxes": 1, "box_price": 500}
            ],
            "notes": "Distributor direct sale"
        }
        resp = requests.post(f"{API_BASE}/dms/direct-sales", json=direct_sale_data2, headers=headers(tokens["distributor1"]))
        if resp.status_code == 200:
            print(f"  ✅ PASS: Distributor can create direct sale for own retailer")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
        
        if retailer2_id:
            print("\n[5.3] Distributor1 POST /api/dms/direct-sales (other distributor's retailer) → 403")
            direct_sale_data3 = {
                "retailer_id": retailer2_id,
                "date": "2026-08-04",
                "items": [
                    {"product_id": product1_id, "qty_boxes": 1, "box_price": 500}
                ],
                "notes": "Cross-distributor direct sale"
            }
            resp = requests.post(f"{API_BASE}/dms/direct-sales", json=direct_sale_data3, headers=headers(tokens["distributor1"]))
            if resp.status_code == 403:
                print(f"  ✅ PASS: Distributor blocked from other distributor's retailer (403)")
            else:
                print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
        
        print("\n[5.4] Stop-sale test: Insufficient stock with stop_sale ON → 400")
        # First, enable stop_sale
        resp_settings = requests.put(f"{API_BASE}/dms/settings", json={"stop_sale_on_negative": True}, headers=headers(tokens["owner"]))
        
        # Try to create direct sale with huge quantity
        direct_sale_huge = {
            "distributor_id": dist1_id,
            "retailer_id": retailer1_id,
            "date": "2026-08-04",
            "items": [
                {"product_id": product1_id, "qty_boxes": 999999, "box_price": 500}
            ],
            "notes": "Huge quantity test"
        }
        resp = requests.post(f"{API_BASE}/dms/direct-sales", json=direct_sale_huge, headers=headers(tokens["owner"]))
        if resp.status_code == 400:
            print(f"  ✅ PASS: Insufficient stock blocked (400)")
            print(f"  Error: {resp.json().get('detail', resp.text)}")
        else:
            print(f"  ❌ FAIL: Expected 400, got {resp.status_code}")
        
        print("\n[5.5] Stop-sale test: Disable stop_sale → should succeed")
        resp_settings = requests.put(f"{API_BASE}/dms/settings", json={"stop_sale_on_negative": False}, headers=headers(tokens["owner"]))
        resp = requests.post(f"{API_BASE}/dms/direct-sales", json=direct_sale_huge, headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            print(f"  ✅ PASS: Direct sale succeeded with stop_sale OFF")
        else:
            print(f"  ⚠️  WARNING: Expected 200, got {resp.status_code} — {resp.text}")
        
        # Re-enable stop_sale
        requests.put(f"{API_BASE}/dms/settings", json={"stop_sale_on_negative": True}, headers=headers(tokens["owner"]))
    
    # ========================================================================
    # TEST 6: PO PDF (owner + distributor)
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 6: PO PDF — GET /api/dms/print/purchase-order/{oid}")
    print("=" * 80)
    
    if not primary_order_id:
        print("  ⚠️  SKIP: No primary order found for PO PDF test")
    else:
        print(f"\n[6.1] Owner GET /api/dms/print/purchase-order/{primary_order_id} → 200")
        resp = requests.get(f"{API_BASE}/dms/print/purchase-order/{primary_order_id}", headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            po_data = resp.json()
            print(f"  ✅ PASS: HTTP 200")
            
            # Verify required fields
            required_fields = ["company_name", "invoice_terms", "invoice_message", "doc_type"]
            for field in required_fields:
                if field in po_data:
                    print(f"  ✅ PASS: Field '{field}' present: {po_data[field][:50] if isinstance(po_data[field], str) else po_data[field]}")
                else:
                    print(f"  ❌ FAIL: Field '{field}' missing")
            
            if po_data.get("doc_type") == "Purchase Order":
                print(f"  ✅ PASS: doc_type='Purchase Order'")
            else:
                print(f"  ❌ FAIL: Expected doc_type='Purchase Order', got {po_data.get('doc_type')}")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
        
        print(f"\n[6.2] Distributor1 GET own PO → 200")
        resp = requests.get(f"{API_BASE}/dms/print/purchase-order/{primary_order_id}", headers=headers(tokens["distributor1"]))
        if resp.status_code in [200, 403]:
            if resp.status_code == 200:
                print(f"  ✅ PASS: Distributor can access own PO (200)")
            else:
                print(f"  ⚠️  INFO: Distributor blocked from PO (403) — may be expected")
        else:
            print(f"  ❌ FAIL: Unexpected status {resp.status_code}")
        
        print(f"\n[6.3] Retailer GET PO → 403")
        resp = requests.get(f"{API_BASE}/dms/print/purchase-order/{primary_order_id}", headers=headers(tokens["retailer1"]))
        if resp.status_code == 403:
            print(f"  ✅ PASS: Retailer blocked (403)")
        else:
            print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    # ========================================================================
    # TEST 7: DOCUMENT STUBS (5 types)
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 7: DOCUMENT STUBS — POST/GET /api/dms/documents")
    print("=" * 80)
    
    if not retailer1_id:
        print("  ⚠️  SKIP: Missing retailer ID for document stubs test")
    else:
        doc_types = [
            ("estimate", "EST-"),
            ("delivery_challan", "DC-"),
            ("sale_return", "SR-"),
            ("credit_note", "CN-"),
            ("debit_note", "DN-")
        ]
        
        created_docs = []
        
        for doc_type, prefix in doc_types:
            print(f"\n[7.{doc_types.index((doc_type, prefix)) + 1}] Owner POST /api/dms/documents (type={doc_type}) → 200")
            doc_data = {
                "type": doc_type,
                "party_type": "retailer",
                "party_id": retailer1_id,
                "items": [
                    {"description": "Oil Change Service", "qty": 2, "rate": 500}
                ],
                "gst_pct": 18
            }
            resp = requests.post(f"{API_BASE}/dms/documents", json=doc_data, headers=headers(tokens["owner"]))
            if resp.status_code == 200:
                doc = resp.json()
                doc_no = doc.get("doc_no", "")
                print(f"  ✅ PASS: HTTP 200, doc_no={doc_no}")
                
                if doc_no.startswith(prefix):
                    print(f"  ✅ PASS: doc_no starts with '{prefix}'")
                else:
                    print(f"  ❌ FAIL: Expected doc_no to start with '{prefix}', got {doc_no}")
                
                # Verify calculations
                subtotal = doc.get("subtotal", 0)
                gst_total = doc.get("gst_total", 0)
                total = doc.get("total", 0)
                
                expected_subtotal = 1000  # 2 * 500
                expected_gst = 180  # 1000 * 0.18
                expected_total = 1180
                
                if subtotal == expected_subtotal:
                    print(f"  ✅ PASS: subtotal={subtotal}")
                else:
                    print(f"  ❌ FAIL: Expected subtotal={expected_subtotal}, got {subtotal}")
                
                if gst_total == expected_gst:
                    print(f"  ✅ PASS: gst_total={gst_total}")
                else:
                    print(f"  ❌ FAIL: Expected gst_total={expected_gst}, got {gst_total}")
                
                if total == expected_total:
                    print(f"  ✅ PASS: total={total}")
                else:
                    print(f"  ❌ FAIL: Expected total={expected_total}, got {total}")
                
                created_docs.append((doc_type, doc.get("id"), doc_no))
            else:
                print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
        
        # Test duplicate doc_no
        if created_docs:
            print(f"\n[7.6] Owner POST /api/dms/documents with duplicate doc_no → 400")
            first_doc_type, first_doc_id, first_doc_no = created_docs[0]
            dup_doc_data = {
                "type": first_doc_type,
                "party_type": "retailer",
                "party_id": retailer1_id,
                "doc_no": first_doc_no,
                "items": [
                    {"description": "Duplicate test", "qty": 1, "rate": 100}
                ],
                "gst_pct": 18
            }
            resp = requests.post(f"{API_BASE}/dms/documents", json=dup_doc_data, headers=headers(tokens["owner"]))
            if resp.status_code == 400:
                print(f"  ✅ PASS: Duplicate doc_no rejected (400)")
            else:
                print(f"  ❌ FAIL: Expected 400, got {resp.status_code}")
        
        # Test GET with filter
        print(f"\n[7.7] Owner GET /api/dms/documents?type=estimate → filters correctly")
        resp = requests.get(f"{API_BASE}/dms/documents?type=estimate", headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            docs = resp.json()
            if isinstance(docs, dict) and "data" in docs:
                docs = docs["data"]
            
            estimate_docs = [d for d in docs if d.get("type") == "estimate"]
            if len(estimate_docs) > 0:
                print(f"  ✅ PASS: Filter working, found {len(estimate_docs)} estimate(s)")
            else:
                print(f"  ⚠️  WARNING: No estimates found (may be expected if none created)")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
        
        # Test print endpoint
        if created_docs:
            print(f"\n[7.8] Owner GET /api/dms/documents/{created_docs[0][1]}/print → 200")
            resp = requests.get(f"{API_BASE}/dms/documents/{created_docs[0][1]}/print", headers=headers(tokens["owner"]))
            if resp.status_code == 200:
                print_data = resp.json()
                print(f"  ✅ PASS: HTTP 200")
                
                # Verify required fields
                required_fields = ["party", "company_name", "invoice_terms", "invoice_message", "doc_type_label"]
                for field in required_fields:
                    if field in print_data:
                        print(f"  ✅ PASS: Field '{field}' present")
                    else:
                        print(f"  ❌ FAIL: Field '{field}' missing")
            else:
                print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
        
        # Test distributor scope
        print(f"\n[7.9] Distributor1 POST /api/dms/documents for own retailer → 200")
        doc_data_dist = {
            "type": "estimate",
            "party_type": "retailer",
            "party_id": retailer1_id,
            "items": [
                {"description": "Distributor estimate", "qty": 1, "rate": 1000}
            ],
            "gst_pct": 18
        }
        resp = requests.post(f"{API_BASE}/dms/documents", json=doc_data_dist, headers=headers(tokens["distributor1"]))
        if resp.status_code == 200:
            print(f"  ✅ PASS: Distributor can create document for own retailer")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
        
        if retailer2_id:
            print(f"\n[7.10] Distributor1 POST /api/dms/documents for other distributor's retailer → 403")
            doc_data_cross = {
                "type": "estimate",
                "party_type": "retailer",
                "party_id": retailer2_id,
                "items": [
                    {"description": "Cross-distributor estimate", "qty": 1, "rate": 1000}
                ],
                "gst_pct": 18
            }
            resp = requests.post(f"{API_BASE}/dms/documents", json=doc_data_cross, headers=headers(tokens["distributor1"]))
            if resp.status_code == 403:
                print(f"  ✅ PASS: Distributor blocked from other distributor's retailer (403)")
            else:
                print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
        
        print(f"\n[7.11] Retailer POST /api/dms/documents → 403")
        resp = requests.post(f"{API_BASE}/dms/documents", json=doc_data, headers=headers(tokens["retailer1"]))
        if resp.status_code == 403:
            print(f"  ✅ PASS: Retailer blocked (403)")
        else:
            print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    # ========================================================================
    # TEST 8: FINANCE SNAPSHOT (owner + accountant only)
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 8: FINANCE SNAPSHOT — GET /api/dms/dashboard/finance-snapshot")
    print("=" * 80)
    
    print("\n[8.1] Owner GET /api/dms/dashboard/finance-snapshot → 200")
    resp = requests.get(f"{API_BASE}/dms/dashboard/finance-snapshot", headers=headers(tokens["owner"]))
    if resp.status_code == 200:
        snapshot = resp.json()
        print(f"  ✅ PASS: HTTP 200")
        
        # Verify 5 numeric fields
        required_fields = ["cash_in_bank", "cash_in_hand", "outstanding_loans", "net_liquid", "net_position"]
        for field in required_fields:
            if field in snapshot:
                value = snapshot[field]
                if isinstance(value, (int, float)):
                    print(f"  ✅ PASS: Field '{field}' present: {value}")
                else:
                    print(f"  ❌ FAIL: Field '{field}' is not numeric: {value}")
            else:
                print(f"  ❌ FAIL: Field '{field}' missing")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
    
    print("\n[8.2] Owner Accountant GET /api/dms/dashboard/finance-snapshot → 200")
    resp = requests.get(f"{API_BASE}/dms/dashboard/finance-snapshot", headers=headers(tokens["accountant"]))
    if resp.status_code == 200:
        print(f"  ✅ PASS: Owner Accountant has access (200)")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
    
    print("\n[8.3] Salesperson GET /api/dms/dashboard/finance-snapshot → 403")
    resp = requests.get(f"{API_BASE}/dms/dashboard/finance-snapshot", headers=headers(tokens["salesperson"]))
    if resp.status_code == 403:
        print(f"  ✅ PASS: Salesperson blocked (403)")
    else:
        print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    print("\n[8.4] Retailer GET /api/dms/dashboard/finance-snapshot → 403")
    resp = requests.get(f"{API_BASE}/dms/dashboard/finance-snapshot", headers=headers(tokens["retailer1"]))
    if resp.status_code == 403:
        print(f"  ✅ PASS: Retailer blocked (403)")
    else:
        print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    print("\n[8.5] Distributor GET /api/dms/dashboard/finance-snapshot → 403")
    resp = requests.get(f"{API_BASE}/dms/dashboard/finance-snapshot", headers=headers(tokens["distributor1"]))
    if resp.status_code == 403:
        print(f"  ✅ PASS: Distributor blocked (403)")
    else:
        print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    # ========================================================================
    # TEST 9: GODOWN REORDER LEVEL + LOW-STOCK
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 9: GODOWN REORDER LEVEL + LOW-STOCK")
    print("=" * 80)
    
    if not all([godown_id, product1_id]):
        print("  ⚠️  SKIP: Missing godown or product ID for reorder level test")
    else:
        print(f"\n[9.1] Owner PUT /api/dms/godowns/{godown_id}/reorder-level → 200")
        reorder_data = {
            "product_id": product1_id,
            "reorder_level_boxes": 999
        }
        resp = requests.put(f"{API_BASE}/dms/godowns/{godown_id}/reorder-level", json=reorder_data, headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            print(f"  ✅ PASS: HTTP 200, reorder level set to 999")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code} — {resp.text}")
        
        print(f"\n[9.2] Owner GET /api/dms/godowns/{godown_id}/inventory → verify low_stock=true")
        resp = requests.get(f"{API_BASE}/dms/godowns/{godown_id}/inventory", headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            inventory = resp.json()
            if isinstance(inventory, dict) and "data" in inventory:
                inventory = inventory["data"]
            
            target_row = next((row for row in inventory if row.get("product_id") == product1_id), None)
            if target_row:
                reorder_level = target_row.get("reorder_level_boxes")
                low_stock = target_row.get("low_stock")
                
                if reorder_level == 999:
                    print(f"  ✅ PASS: reorder_level_boxes=999")
                else:
                    print(f"  ❌ FAIL: Expected reorder_level_boxes=999, got {reorder_level}")
                
                if low_stock is True:
                    print(f"  ✅ PASS: low_stock=true")
                else:
                    print(f"  ❌ FAIL: Expected low_stock=true, got {low_stock}")
            else:
                print(f"  ⚠️  WARNING: Product {product1_id} not found in godown inventory")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
        
        print(f"\n[9.3] Owner GET /api/dms/godowns/low-stock → includes target row")
        resp = requests.get(f"{API_BASE}/dms/godowns/low-stock", headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            low_stock_rows = resp.json()
            if isinstance(low_stock_rows, dict) and "data" in low_stock_rows:
                low_stock_rows = low_stock_rows["data"]
            
            print(f"  ✅ PASS: HTTP 200, found {len(low_stock_rows)} low-stock row(s)")
            
            # Check if our target row is in the list
            target_in_list = any(row.get("product_id") == product1_id and row.get("godown_id") == godown_id for row in low_stock_rows)
            if target_in_list:
                print(f"  ✅ PASS: Target row (godown={godown_id}, product={product1_id}) in low-stock list")
                
                # Verify godown_name and product_name are present
                target_row = next(row for row in low_stock_rows if row.get("product_id") == product1_id and row.get("godown_id") == godown_id)
                if "godown_name" in target_row and "product_name" in target_row:
                    print(f"  ✅ PASS: godown_name and product_name present")
                else:
                    print(f"  ❌ FAIL: godown_name or product_name missing")
            else:
                print(f"  ⚠️  WARNING: Target row not in low-stock list (may be expected if stock > reorder level)")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
        
        print(f"\n[9.4] Reset reorder level to 0 → low_stock should become false")
        reorder_data_reset = {
            "product_id": product1_id,
            "reorder_level_boxes": 0
        }
        resp = requests.put(f"{API_BASE}/dms/godowns/{godown_id}/reorder-level", json=reorder_data_reset, headers=headers(tokens["owner"]))
        if resp.status_code == 200:
            print(f"  ✅ PASS: Reorder level reset to 0")
            
            # Verify low_stock is now false
            resp_inv = requests.get(f"{API_BASE}/dms/godowns/{godown_id}/inventory", headers=headers(tokens["owner"]))
            if resp_inv.status_code == 200:
                inventory = resp_inv.json()
                if isinstance(inventory, dict) and "data" in inventory:
                    inventory = inventory["data"]
                
                target_row = next((row for row in inventory if row.get("product_id") == product1_id), None)
                if target_row and target_row.get("low_stock") is False:
                    print(f"  ✅ PASS: low_stock=false after reset")
                else:
                    print(f"  ⚠️  WARNING: low_stock not false after reset")
        else:
            print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
        
        print(f"\n[9.5] Salesperson PUT /api/dms/godowns/{godown_id}/reorder-level → 403")
        resp = requests.put(f"{API_BASE}/dms/godowns/{godown_id}/reorder-level", json=reorder_data, headers=headers(tokens["salesperson"]))
        if resp.status_code == 403:
            print(f"  ✅ PASS: Salesperson blocked (403)")
        else:
            print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
        
        print(f"\n[9.6] Retailer GET /api/dms/godowns/low-stock → 403")
        resp = requests.get(f"{API_BASE}/dms/godowns/low-stock", headers=headers(tokens["retailer1"]))
        if resp.status_code == 403:
            print(f"  ✅ PASS: Retailer blocked (403)")
        else:
            print(f"  ❌ FAIL: Expected 403, got {resp.status_code}")
    
    # ========================================================================
    # TEST 10: REGRESSION SANITY
    # ========================================================================
    print("\n" + "=" * 80)
    print("TEST 10: REGRESSION SANITY")
    print("=" * 80)
    
    print("\n[10.1] Phase 2A: POST /api/dms/expenses still works")
    expense_data = {
        "category": "Office Supplies",
        "amount": 100,
        "date": "2026-08-04",
        "description": "Regression test"
    }
    resp = requests.post(f"{API_BASE}/dms/expenses", json=expense_data, headers=headers(tokens["owner"]))
    if resp.status_code in [200, 201]:
        print(f"  ✅ PASS: Expenses endpoint working")
    else:
        print(f"  ❌ FAIL: Expected 200/201, got {resp.status_code}")
    
    print("\n[10.2] Phase 2B: GET /api/dms/bank-accounts still works")
    resp = requests.get(f"{API_BASE}/dms/bank-accounts", headers=headers(tokens["owner"]))
    if resp.status_code == 200:
        print(f"  ✅ PASS: Bank accounts endpoint working")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
    
    print("\n[10.3] Phase 2B: GET /api/dms/godowns still works")
    resp = requests.get(f"{API_BASE}/dms/godowns", headers=headers(tokens["owner"]))
    if resp.status_code == 200:
        godowns = resp.json()
        print(f"  ✅ PASS: Godowns endpoint working, found {len(godowns)} godown(s)")
    else:
        print(f"  ❌ FAIL: Expected 200, got {resp.status_code}")
    
    # ========================================================================
    # SUMMARY
    # ========================================================================
    print("\n" + "=" * 80)
    print("PHASE 2C BACKEND TESTING COMPLETE")
    print("=" * 80)
    print("\nAll Phase 2C endpoints have been tested.")
    print("Review the output above for detailed results.")
    print("\n" + "=" * 80)

if __name__ == "__main__":
    main()
